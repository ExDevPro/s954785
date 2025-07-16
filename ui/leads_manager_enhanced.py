# ui/leads_manager_enhanced.py
"""
Enhanced Leads Manager with Professional UI and Fixed Data Persistence
Uses the new enhanced base manager with all improvements
"""

from ui.enhanced_base_manager import EnhancedBaseManager
from PyQt6.QtWidgets import QMessageBox, QFileDialog, QHeaderView
from PyQt6.QtCore import Qt
import os
import csv
from openpyxl import Workbook, load_workbook
from core.utils.logger import get_module_logger

logger = get_module_logger(__name__)

class LeadsManagerEnhanced(EnhancedBaseManager):
    """Enhanced Leads Manager with professional UI and data persistence"""
    
    def __init__(self, parent=None):
        # Default headers for leads
        self.default_headers = [
            "Email", "First Name", "Last Name", "Company", 
            "Phone", "Country", "Industry", "Website", "Notes"
        ]
        
        super().__init__(
            manager_type="Leads",
            data_subdir="leads", 
            file_extension=".xlsx",
            parent=parent
        )
    
    def _create_new_list_structure(self, list_name: str):
        """Create new leads list structure with folder and Excel file"""
        try:
            # Create folder for the list
            list_folder = os.path.join(self.data_dir, list_name)
            os.makedirs(list_folder, exist_ok=True)
            
            # Create Excel file with default headers
            excel_path = os.path.join(list_folder, f"{list_name}.xlsx")
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(self.default_headers)
            workbook.save(excel_path)
            
            logger.info(f"Created leads list structure: {list_folder}")
            
        except Exception as e:
            logger.error(f"Failed to create leads list structure: {str(e)}")
            raise e
    
    def _get_list_data_path(self, list_name: str) -> str:
        """Get path to leads list Excel file"""
        return os.path.join(self.data_dir, list_name, f"{list_name}.xlsx")
    
    def _load_list_data(self, list_name: str):
        """Load leads data from Excel file"""
        try:
            data_path = self._get_list_data_path(list_name)
            
            if not os.path.exists(data_path):
                # Create empty list if file doesn't exist
                self._create_empty_list_data()
                return
            
            # Load from Excel
            workbook = load_workbook(data_path, read_only=True)
            worksheet = workbook.active
            
            # Read headers
            self.headers = []
            header_row = worksheet[1]
            for cell in header_row:
                if cell.value:
                    self.headers.append(str(cell.value))
                else:
                    break
            
            # Use default headers if none found
            if not self.headers:
                self.headers = self.default_headers.copy()
            
            # Read data
            self.current_data = []
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 1):
                if any(cell is not None for cell in row):  # Skip empty rows
                    # Convert to list and pad with empty strings
                    row_data = list(row[:len(self.headers)])
                    while len(row_data) < len(self.headers):
                        row_data.append("")
                    
                    # Convert None values to empty strings
                    row_data = [str(cell) if cell is not None else "" for cell in row_data]
                    self.current_data.append(row_data)
            
            workbook.close()
            
            # Update table display
            self._update_table_display()
            
            logger.info(f"Loaded leads list: {list_name} ({len(self.current_data)} leads)")
            
        except Exception as e:
            logger.error(f"Failed to load leads data: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to load leads data:\n{str(e)}")
            self._create_empty_list_data()
    
    def _create_empty_list_data(self):
        """Create empty leads data structure"""
        self.headers = self.default_headers.copy()
        self.current_data = []
        self._update_table_display()
    
    def _save_current_data(self):
        """Save current leads data to Excel file"""
        if not self.current_list_name:
            return
        
        try:
            data_path = self._get_list_data_path(self.current_list_name)
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(data_path), exist_ok=True)
            
            # Collect current data from table (in case of unsaved edits)
            self._collect_table_data()
            
            # Save to Excel
            workbook = Workbook()
            worksheet = workbook.active
            
            # Write headers
            for col_idx, header in enumerate(self.headers, 1):
                worksheet.cell(row=1, column=col_idx, value=header)
            
            # Write data
            for row_idx, row_data in enumerate(self.current_data, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    if col_idx <= len(self.headers):
                        worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
            
            workbook.save(data_path)
            workbook.close()
            
            logger.info(f"Saved leads data: {self.current_list_name} ({len(self.current_data)} leads)")
            
        except Exception as e:
            logger.error(f"Failed to save leads data: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to save leads data:\n{str(e)}")
    
    def _collect_table_data(self):
        """Collect current data from table widget"""
        if not self.table_widget:
            return
        
        self.current_data = []
        for row_idx in range(self.table_widget.rowCount()):
            row_data = []
            for col_idx in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row_idx, col_idx)
                cell_value = item.text() if item else ""
                row_data.append(cell_value)
            self.current_data.append(row_data)
    
    def _import_data(self):
        """Import leads data from CSV or Excel file"""
        if not self.current_list_name:
            QMessageBox.warning(self, "Error", "Please select a leads list first!")
            return
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Leads Data",
            "", "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, 0)  # Indeterminate progress
            
            imported_data = []
            imported_headers = []
            
            if file_path.lower().endswith('.csv'):
                # Import from CSV
                with open(file_path, 'r', encoding='utf-8', newline='') as csvfile:
                    # Detect dialect
                    sample = csvfile.read(1024)
                    csvfile.seek(0)
                    dialect = csv.Sniffer().sniff(sample)
                    
                    reader = csv.reader(csvfile, dialect)
                    
                    # Read headers
                    imported_headers = next(reader, [])
                    
                    # Read data
                    for row in reader:
                        if any(cell.strip() for cell in row):  # Skip empty rows
                            imported_data.append(row)
            
            else:
                # Import from Excel
                workbook = load_workbook(file_path, read_only=True)
                worksheet = workbook.active
                
                # Read headers
                header_row = worksheet[1]
                imported_headers = []
                for cell in header_row:
                    if cell.value:
                        imported_headers.append(str(cell.value))
                    else:
                        break
                
                # Read data
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        row_data = [str(cell) if cell is not None else "" for cell in row[:len(imported_headers)]]
                        imported_data.append(row_data)
                
                workbook.close()
            
            # Ask user about merge or replace
            if self.current_data:
                reply = QMessageBox.question(
                    self, "Import Options",
                    f"Found {len(imported_data)} leads to import.\n\n"
                    f"Current list has {len(self.current_data)} leads.\n\n"
                    f"How would you like to import?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
                )
                
                if reply == QMessageBox.StandardButton.Cancel:
                    return
                elif reply == QMessageBox.StandardButton.Yes:
                    # Merge data
                    pass  # Keep current data and append
                else:
                    # Replace data
                    self.current_data = []
                    self.headers = imported_headers if imported_headers else self.default_headers
            else:
                # No existing data, just import
                self.headers = imported_headers if imported_headers else self.default_headers
            
            # Process imported data to match current headers
            for row_data in imported_data:
                # Pad or trim to match headers length
                processed_row = row_data[:len(self.headers)]
                while len(processed_row) < len(self.headers):
                    processed_row.append("")
                self.current_data.append(processed_row)
            
            # Update display
            self._update_table_display()
            
            # Auto-save if enabled
            if self.auto_save_enabled:
                self._save_current_data()
            
            self.progress_bar.setVisible(False)
            
            QMessageBox.information(
                self, "Import Successful",
                f"Successfully imported {len(imported_data)} leads!"
            )
            
        except Exception as e:
            self.progress_bar.setVisible(False)
            logger.error(f"Failed to import leads: {str(e)}")
            QMessageBox.critical(self, "Import Error", f"Failed to import leads:\n{str(e)}")
    
    def _export_data(self):
        """Export leads data to CSV or Excel file"""
        if not self.current_data:
            QMessageBox.warning(self, "Error", "No data to export!")
            return
        
        file_path, file_filter = QFileDialog.getSaveFileName(
            self, "Export Leads Data",
            f"{self.current_list_name}_leads.xlsx",
            "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, len(self.current_data))
            
            if file_path.lower().endswith('.csv'):
                # Export to CSV
                with open(file_path, 'w', encoding='utf-8', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(self.headers)
                    
                    for idx, row_data in enumerate(self.current_data):
                        writer.writerow(row_data)
                        self.progress_bar.setValue(idx + 1)
            
            else:
                # Export to Excel
                workbook = Workbook()
                worksheet = workbook.active
                
                # Write headers
                for col_idx, header in enumerate(self.headers, 1):
                    worksheet.cell(row=1, column=col_idx, value=header)
                
                # Write data
                for row_idx, row_data in enumerate(self.current_data, 2):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    self.progress_bar.setValue(row_idx - 1)
                
                workbook.save(file_path)
                workbook.close()
            
            self.progress_bar.setVisible(False)
            
            QMessageBox.information(
                self, "Export Successful",
                f"Successfully exported {len(self.current_data)} leads to:\n{file_path}"
            )
            
        except Exception as e:
            self.progress_bar.setVisible(False)
            logger.error(f"Failed to export leads: {str(e)}")
            QMessageBox.critical(self, "Export Error", f"Failed to export leads:\n{str(e)}")
    
    def _list_exists(self, list_name: str) -> bool:
        """Check if leads list exists"""
        return os.path.exists(self._get_list_data_path(list_name))
    
    def _delete_list_structure(self, list_name: str):
        """Delete leads list structure"""
        import shutil
        list_folder = os.path.join(self.data_dir, list_name)
        if os.path.exists(list_folder):
            shutil.rmtree(list_folder)
            logger.info(f"Deleted leads list: {list_folder}")
    
    def _rename_list(self, old_name: str, new_name: str):
        """Rename leads list"""
        old_folder = os.path.join(self.data_dir, old_name)
        new_folder = os.path.join(self.data_dir, new_name)
        
        if os.path.exists(old_folder):
            os.rename(old_folder, new_folder)
            
            # Rename the Excel file inside
            old_file = os.path.join(new_folder, f"{old_name}.xlsx")
            new_file = os.path.join(new_folder, f"{new_name}.xlsx")
            if os.path.exists(old_file):
                os.rename(old_file, new_file)
            
            logger.info(f"Renamed leads list: {old_name} -> {new_name}")
    
    def _count_items_in_list(self, list_name: str) -> int:
        """Count leads in a list"""
        try:
            data_path = self._get_list_data_path(list_name)
            if not os.path.exists(data_path):
                return 0
            
            workbook = load_workbook(data_path, read_only=True)
            worksheet = workbook.active
            
            # Count non-empty rows (excluding header)
            count = 0
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row):
                    count += 1
            
            workbook.close()
            return count
            
        except Exception as e:
            logger.error(f"Failed to count leads in {list_name}: {str(e)}")
            return 0
    
    def _add_row(self):
        """Add a new leads row with email focus"""
        super()._add_row()
        
        # Focus on email field for new row
        if self.table_widget.rowCount() > 0:
            last_row = self.table_widget.rowCount() - 1
            email_column = 0  # Email is first column
            
            # Find email column
            for col_idx, header in enumerate(self.headers):
                if "email" in header.lower():
                    email_column = col_idx
                    break
            
            email_item = self.table_widget.item(last_row, email_column)
            if email_item:
                self.table_widget.setCurrentItem(email_item)
                self.table_widget.editItem(email_item)