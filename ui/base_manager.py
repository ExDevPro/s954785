# ui/base_manager.py
"""
Base class for all data managers providing consistent UI and threading
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QListWidget, QPushButton, 
    QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit, QProgressBar,
    QMessageBox, QInputDialog, QFileDialog, QAbstractItemView, QStyle,
    QFrame, QSplitter, QScrollArea
)
from PyQt6.QtCore import Qt, pyqtSignal, QTimer
from PyQt6.QtGui import QFont, QPalette, QCursor

import os
from workers.data_operation_worker import DataOperationWorker
from core.utils.logger import get_module_logger

logger = get_module_logger(__name__)

class BaseManager(QWidget):
    """Base class for all list managers with consistent UI and threading"""
    
    # Signals
    counts_changed = pyqtSignal(int, int)  # list_count, item_count
    list_selected = pyqtSignal(str)  # list_name
    operation_progress = pyqtSignal(int, int, str)  # current, total, operation
    
    def __init__(self, manager_type: str, data_subdir: str, file_extension: str = ".xlsx", parent=None):
        super().__init__(parent)
        
        # Manager configuration
        self.manager_type = manager_type  # e.g., "leads", "smtp", "subjects"
        self.data_subdir = data_subdir    # e.g., "leads", "smtps", "subjects"
        self.file_extension = file_extension  # Default file extension
        
        # Paths
        self.base_path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
        self.data_dir = os.path.join(self.base_path, 'data', self.data_subdir)
        os.makedirs(self.data_dir, exist_ok=True)
        
        # State
        self.current_list_name = None
        self.current_data = []
        self.headers = []
        
        # Threading
        self.worker = DataOperationWorker()
        self.worker.operation_completed.connect(self._on_operation_completed)
        self.worker.error_occurred.connect(self._on_operation_error)
        self.worker.progress_updated.connect(self._on_progress_updated)
        
        # UI Components (will be created by _build_ui)
        self.list_widget = None
        self.table_widget = None
        self.progress_bar = None
        self.search_bar = None
        
        # Build UI
        self._build_ui()
        self._apply_styling()
        self._load_lists()
        
        # Auto-refresh timer for counts
        self.refresh_timer = QTimer()
        self.refresh_timer.timeout.connect(self._update_counts)
        self.refresh_timer.start(5000)  # Refresh every 5 seconds
    
    def _build_ui(self):
        """Build the standard UI layout - override for custom layouts"""
        main_layout = QHBoxLayout(self)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Left Panel - List Management
        left_panel = self._create_left_panel()
        
        # Right Panel - Data Table
        right_panel = self._create_right_panel()
        
        # Create splitter for resizable panels
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([300, 700])  # Initial sizes
        
        main_layout.addWidget(splitter)
    
    def _create_left_panel(self):
        """Create the left panel with list management"""
        panel = QFrame()
        panel.setFrameStyle(QFrame.Shape.StyledPanel)
        panel.setMinimumWidth(250)
        panel.setMaximumWidth(400)
        
        layout = QVBoxLayout(panel)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Title
        title = QLabel(f"{self.manager_type.title()} Lists")
        title.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        layout.addWidget(title)
        
        # Search bar
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText(f"🔍 Search {self.manager_type} lists...")
        self.search_bar.textChanged.connect(self._filter_lists)
        layout.addWidget(self.search_bar)
        
        # List management buttons
        buttons_layout = QHBoxLayout()
        
        btn_new = QPushButton("➕ New List")
        btn_new.setToolTip(f"Create new {self.manager_type} list")
        btn_new.clicked.connect(self._create_new_list)
        btn_new.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        btn_delete = QPushButton("🗑 Delete List")
        btn_delete.setToolTip(f"Delete selected {self.manager_type} list")
        btn_delete.clicked.connect(self._delete_selected_list)
        btn_delete.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        
        buttons_layout.addWidget(btn_new)
        buttons_layout.addWidget(btn_delete)
        layout.addLayout(buttons_layout)
        
        # List widget
        self.list_widget = QListWidget()
        self.list_widget.setAlternatingRowColors(True)
        self.list_widget.itemClicked.connect(self._on_list_selected)
        self.list_widget.setMinimumHeight(200)
        layout.addWidget(self.list_widget)
        
        # List statistics
        self.list_stats_label = QLabel("No lists")
        self.list_stats_label.setStyleSheet("color: #666; font-style: italic;")
        layout.addWidget(self.list_stats_label)
        
        layout.addStretch()
        return panel
    
    def _create_right_panel(self):
        """Create the right panel with data table and tools"""
        panel = QFrame()
        panel.setFrameStyle(QFrame.Shape.StyledPanel)
        
        layout = QVBoxLayout(panel)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Top toolbar
        toolbar_layout = self._create_toolbar()
        layout.addLayout(toolbar_layout)
        
        # Progress bar (hidden by default)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Data table
        self.table_widget = QTableWidget()
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_widget.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        
        # Enable horizontal scrolling
        self.table_widget.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_widget.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        # Auto-resize columns but allow manual resize
        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(True)
        
        layout.addWidget(self.table_widget)
        
        # Data statistics
        self.data_stats_label = QLabel("No data")
        self.data_stats_label.setStyleSheet("color: #666; font-style: italic;")
        layout.addWidget(self.data_stats_label)
        
        return panel
    
    def _create_toolbar(self):
        """Create the data management toolbar - override for custom tools"""
        toolbar_layout = QHBoxLayout()
        
        # Import button
        btn_import = QPushButton("📥 Import Data")
        btn_import.setToolTip(f"Import {self.manager_type} from file")
        btn_import.clicked.connect(self._import_data)
        btn_import.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        
        # Export button
        btn_export = QPushButton("📤 Export Data")
        btn_export.setToolTip(f"Export {self.manager_type} to file")
        btn_export.clicked.connect(self._export_data)
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        """)
        
        # Add row button
        btn_add = QPushButton("➕ Add Row")
        btn_add.setToolTip(f"Add new {self.manager_type} entry")
        btn_add.clicked.connect(self._add_row)
        
        # Delete row button
        btn_delete_row = QPushButton("➖ Delete Row")
        btn_delete_row.setToolTip(f"Delete selected {self.manager_type} entry")
        btn_delete_row.clicked.connect(self._delete_row)
        
        # Horizontal scroll helper button
        btn_scroll = QPushButton("⇄ Scroll Table")
        btn_scroll.setToolTip("Scroll table to see all columns")
        btn_scroll.clicked.connect(self._toggle_table_scroll)
        
        toolbar_layout.addWidget(btn_import)
        toolbar_layout.addWidget(btn_export)
        toolbar_layout.addWidget(btn_add)
        toolbar_layout.addWidget(btn_delete_row)
        toolbar_layout.addStretch()
        toolbar_layout.addWidget(btn_scroll)
        
        return toolbar_layout
    
    def _apply_styling(self):
        """Apply consistent styling"""
        self.setStyleSheet("""
            QFrame {
                border: 1px solid #ddd;
                border-radius: 5px;
                background-color: #fafafa;
            }
            QListWidget {
                border: 1px solid #ccc;
                border-radius: 3px;
                background-color: white;
                selection-background-color: #3498db;
            }
            QTableWidget {
                border: 1px solid #ccc;
                border-radius: 3px;
                background-color: white;
                gridline-color: #ddd;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QLineEdit {
                border: 1px solid #ccc;
                border-radius: 3px;
                padding: 5px;
                background-color: white;
            }
            QProgressBar {
                border: 1px solid #ccc;
                border-radius: 3px;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 2px;
            }
        """)
    
    # =============================================================================
    # Core Methods (Override these in subclasses)
    # =============================================================================
    
    def _load_lists(self):
        """Load available lists - override in subclasses"""
        self.list_widget.clear()
        
        try:
            if os.path.exists(self.data_dir):
                items = os.listdir(self.data_dir)
                lists = []
                
                for item in items:
                    item_path = os.path.join(self.data_dir, item)
                    if os.path.isdir(item_path):
                        lists.append(item)
                    elif item.endswith(self.file_extension):
                        # Handle file-based lists (like SMTP)
                        list_name = os.path.splitext(item)[0]
                        lists.append(list_name)
                
                lists.sort()
                self.list_widget.addItems(lists)
                
                # Update statistics
                self._update_counts()
                
        except Exception as e:
            logger.error(f"Failed to load {self.manager_type} lists: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to load {self.manager_type} lists:\n{str(e)}")
    
    def _get_list_data_path(self, list_name: str) -> str:
        """Get the path to list data - override in subclasses"""
        if self.file_extension == ".xlsx":
            return os.path.join(self.data_dir, f"{list_name}.xlsx")
        else:
            return os.path.join(self.data_dir, list_name)
    
    def _get_item_count(self, list_name: str) -> int:
        """Get the number of items in a list - override in subclasses"""
        try:
            data_path = self._get_list_data_path(list_name)
            
            if self.file_extension == ".xlsx" and os.path.exists(data_path):
                from openpyxl import load_workbook
                workbook = load_workbook(data_path, read_only=True)
                count = workbook.active.max_row - 1  # Subtract header
                workbook.close()
                return max(0, count)
            elif self.file_extension == ".txt" and os.path.exists(data_path):
                with open(data_path, 'r', encoding='utf-8') as f:
                    return sum(1 for line in f if line.strip())
            elif os.path.isdir(data_path):
                # Handle folder-based lists
                return len([f for f in os.listdir(data_path) if os.path.isfile(os.path.join(data_path, f))])
            
        except Exception as e:
            logger.error(f"Failed to count items in {list_name}: {str(e)}")
        
        return 0
    
    # =============================================================================
    # UI Event Handlers
    # =============================================================================
    
    def _filter_lists(self, text: str):
        """Filter lists based on search text"""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            item.setHidden(text.lower() not in item.text().lower())
    
    def _on_list_selected(self, item):
        """Handle list selection"""
        if item:
            list_name = item.text()
            self.current_list_name = list_name
            self.list_selected.emit(list_name)
            self._load_list_data(list_name)
    
    def _update_counts(self):
        """Update list and item counts"""
        try:
            list_count = self.list_widget.count()
            
            # Calculate total items across all lists
            total_items = 0
            for i in range(list_count):
                item = self.list_widget.item(i)
                if item and not item.isHidden():
                    list_name = item.text()
                    total_items += self._get_item_count(list_name)
            
            # Update statistics labels
            self.list_stats_label.setText(f"{list_count} lists, {total_items} total items")
            
            if hasattr(self, 'current_data'):
                self.data_stats_label.setText(f"{len(self.current_data)} items in current list")
            
            # Emit signal for dashboard
            self.counts_changed.emit(list_count, total_items)
            
        except Exception as e:
            logger.error(f"Failed to update counts: {str(e)}")
    
    # =============================================================================
    # Data Operations (Threaded)
    # =============================================================================
    
    def _load_list_data(self, list_name: str):
        """Load data for a specific list"""
        try:
            data_path = self._get_list_data_path(list_name)
            
            if os.path.exists(data_path):
                self.progress_bar.setVisible(True)
                self.progress_bar.setRange(0, 0)  # Indeterminate
                
                self.worker.set_operation(
                    'load_data',
                    file_path=data_path,
                    manager_type=self.manager_type
                )
                self.worker.start()
            else:
                # Create empty list
                self._create_empty_list_data()
                
        except Exception as e:
            logger.error(f"Failed to load list data: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to load list data:\n{str(e)}")
    
    def _create_empty_list_data(self):
        """Create empty data structure - override in subclasses"""
        self.headers = ["Column 1"]
        self.current_data = []
        self._update_table()
    
    def _update_table(self):
        """Update the table widget with current data"""
        try:
            self.table_widget.setRowCount(len(self.current_data))
            self.table_widget.setColumnCount(len(self.headers))
            self.table_widget.setHorizontalHeaderLabels(self.headers)
            
            for row_idx, row_data in enumerate(self.current_data):
                for col_idx, cell_value in enumerate(row_data):
                    if col_idx < len(self.headers):
                        item = QTableWidgetItem(str(cell_value) if cell_value is not None else "")
                        self.table_widget.setItem(row_idx, col_idx, item)
            
            # Auto-resize columns to content
            self.table_widget.resizeColumnsToContents()
            
        except Exception as e:
            logger.error(f"Failed to update table: {str(e)}")
    
    # =============================================================================
    # Threading Callbacks
    # =============================================================================
    
    def _on_operation_completed(self, result: dict):
        """Handle completed operations"""
        self.progress_bar.setVisible(False)
        
        try:
            if 'headers' in result:
                self.headers = result['headers']
                self.current_data = result['data']
                self._update_table()
                
            elif 'success' in result:
                # Save operation completed
                QMessageBox.information(self, "Success", f"Data saved successfully!")
                self._load_lists()  # Refresh lists
                
            self._update_counts()
            
        except Exception as e:
            logger.error(f"Failed to handle operation result: {str(e)}")
    
    def _on_operation_error(self, error_msg: str, operation_type: str):
        """Handle operation errors"""
        self.progress_bar.setVisible(False)
        logger.error(f"{operation_type} error: {error_msg}")
        QMessageBox.critical(self, f"{operation_type.title()} Error", error_msg)
    
    def _on_progress_updated(self, current: int, total: int, operation: str):
        """Handle progress updates"""
        if total > 0:
            self.progress_bar.setRange(0, total)
            self.progress_bar.setValue(current)
        else:
            self.progress_bar.setRange(0, 0)  # Indeterminate
        
        self.progress_bar.setFormat(f"{operation}: {current}/{total}")
    
    # =============================================================================
    # Action Handlers (Override these for custom behavior)
    # =============================================================================
    
    def _create_new_list(self):
        """Create a new list"""
        dialog = QInputDialog(self)
        dialog.setWindowTitle(f"Create New {self.manager_type.title()} List")
        dialog.setLabelText(f"Enter name for new {self.manager_type} list:")
        dialog.setTextEchoMode(QLineEdit.EchoMode.Normal)
        
        if dialog.exec() == QInputDialog.DialogCode.Accepted:
            list_name = dialog.textValue().strip()
            
            if not list_name:
                QMessageBox.warning(self, "Error", "List name cannot be empty!")
                return
            
            # Check if list already exists
            data_path = self._get_list_data_path(list_name)
            if os.path.exists(data_path):
                QMessageBox.warning(self, "Error", f"List '{list_name}' already exists!")
                return
            
            try:
                # Create the list
                self._create_new_list_structure(list_name)
                
                # Refresh lists and select the new one
                self._load_lists()
                
                # Select the new list
                for i in range(self.list_widget.count()):
                    item = self.list_widget.item(i)
                    if item.text() == list_name:
                        self.list_widget.setCurrentItem(item)
                        self._on_list_selected(item)
                        break
                
                QMessageBox.information(self, "Success", f"List '{list_name}' created successfully!")
                
            except Exception as e:
                logger.error(f"Failed to create list: {str(e)}")
                QMessageBox.critical(self, "Error", f"Failed to create list:\n{str(e)}")
    
    def _create_new_list_structure(self, list_name: str):
        """Create the actual list structure - override in subclasses"""
        data_path = self._get_list_data_path(list_name)
        
        if self.file_extension == ".xlsx":
            # Create empty Excel file
            from openpyxl import Workbook
            workbook = Workbook()
            worksheet = workbook.active
            # Add default headers - override in subclasses
            worksheet.append(["Column 1"])
            workbook.save(data_path)
        elif os.path.dirname(data_path) != data_path:  # It's a folder
            os.makedirs(data_path, exist_ok=True)
        else:
            # Create empty text file
            with open(data_path, 'w', encoding='utf-8') as f:
                f.write("")
    
    def _delete_selected_list(self):
        """Delete the selected list"""
        current_item = self.list_widget.currentItem()
        if not current_item:
            QMessageBox.warning(self, "Error", "Please select a list to delete!")
            return
        
        list_name = current_item.text()
        
        # Confirmation dialog
        reply = QMessageBox.question(
            self, 
            "Confirm Deletion",
            f"Are you sure you want to delete the list '{list_name}' and all its data?\n\nThis action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                data_path = self._get_list_data_path(list_name)
                
                if os.path.isfile(data_path):
                    os.remove(data_path)
                elif os.path.isdir(data_path):
                    import shutil
                    shutil.rmtree(data_path)
                
                # Refresh lists
                self._load_lists()
                
                # Clear current selection
                self.current_list_name = None
                self.current_data = []
                self.headers = []
                self._update_table()
                
                QMessageBox.information(self, "Success", f"List '{list_name}' deleted successfully!")
                
            except Exception as e:
                logger.error(f"Failed to delete list: {str(e)}")
                QMessageBox.critical(self, "Error", f"Failed to delete list:\n{str(e)}")
    
    def _import_data(self):
        """Import data from file"""
        if not self.current_list_name:
            QMessageBox.warning(self, "Error", "Please select a list first!")
            return
        
        # File dialog
        file_filter = "Excel Files (*.xlsx);;CSV Files (*.csv);;Text Files (*.txt);;All Files (*)"
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            f"Import {self.manager_type.title()} Data", 
            "", 
            file_filter
        )
        
        if file_path:
            # Determine operation type based on file extension
            if file_path.endswith('.xlsx'):
                operation_type = 'import_excel'
            elif file_path.endswith('.csv'):
                operation_type = 'import_csv'
            elif file_path.endswith('.txt'):
                operation_type = 'import_txt'
            else:
                QMessageBox.warning(self, "Error", "Unsupported file format!")
                return
            
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, 0)  # Indeterminate
            
            self.worker.set_operation(
                operation_type,
                file_path=file_path,
                manager_type=self.manager_type
            )
            self.worker.start()
    
    def _export_data(self):
        """Export current data to file"""
        if not self.current_data:
            QMessageBox.warning(self, "Error", "No data to export!")
            return
        
        # File dialog
        file_filter = "Excel Files (*.xlsx);;CSV Files (*.csv);;Text Files (*.txt)"
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            f"Export {self.manager_type.title()} Data", 
            f"{self.current_list_name or 'export'}.xlsx", 
            file_filter
        )
        
        if file_path:
            # Determine operation type
            if file_path.endswith('.xlsx'):
                operation_type = 'save_excel'
            elif file_path.endswith('.csv'):
                operation_type = 'save_csv'
            elif file_path.endswith('.txt'):
                operation_type = 'save_txt'
            else:
                QMessageBox.warning(self, "Error", "Unsupported file format!")
                return
            
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, 0)
            
            self.worker.set_operation(
                operation_type,
                data_to_save=self.current_data,
                save_path=file_path,
                headers=self.headers
            )
            self.worker.start()
    
    def _add_row(self):
        """Add a new row to the current data"""
        if not self.current_list_name:
            QMessageBox.warning(self, "Error", "Please select a list first!")
            return
        
        # Create empty row with same number of columns as headers
        new_row = [""] * len(self.headers)
        self.current_data.append(new_row)
        
        # Update table
        self._update_table()
        
        # Select the new row for editing
        new_row_index = len(self.current_data) - 1
        self.table_widget.selectRow(new_row_index)
        self.table_widget.scrollToItem(self.table_widget.item(new_row_index, 0))
    
    def _delete_row(self):
        """Delete selected row"""
        current_row = self.table_widget.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Error", "Please select a row to delete!")
            return
        
        # Confirmation
        reply = QMessageBox.question(
            self,
            "Confirm Deletion",
            f"Are you sure you want to delete row {current_row + 1}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            del self.current_data[current_row]
            self._update_table()
    
    def _toggle_table_scroll(self):
        """Toggle table horizontal scroll position"""
        horizontal_bar = self.table_widget.horizontalScrollBar()
        
        if horizontal_bar.value() == horizontal_bar.minimum():
            # Scroll to end
            horizontal_bar.setValue(horizontal_bar.maximum())
        else:
            # Scroll to beginning
            horizontal_bar.setValue(horizontal_bar.minimum())