# ui/main_window.py (Modified for Signal Connection)
import os
import glob
import json
import openpyxl

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QScrollArea, QFrame, QLabel, QSizePolicy, QGridLayout,
    QSplitter, QListWidget, QListWidgetItem, QStackedWidget,
    QComboBox, QMessageBox, QPushButton
)
from PyQt6.QtCore import Qt, QTimer, QSize, pyqtSignal
from PyQt6.QtGui import QIcon, QFont, QPalette, QPixmap, QCursor, QMovie
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas

# Import UI Managers - using enhanced versions
def import_managers():
    """Import managers with enhanced versions as priority"""
    try:
        # Try importing enhanced versions first
        from ui.leads_manager_enhanced import LeadsManagerEnhanced as LeadsManager
        from ui.smtp_manager_enhanced import SMTPManagerEnhanced as SMTPManager
        from ui.subjects_manager_enhanced import SubjectsManagerEnhanced as SubjectManager
        from ui.proxy_manager_enhanced import ProxyManagerEnhanced as ProxyManager
        print("✅ Using enhanced leads, SMTP, subjects, and proxy managers")
        
        # Fallback to original versions for message and attachment managers
        try:
            from ui.message_manager import MessageManager
            from ui.attachment_manager import AttachmentManager
        except:
            # Create placeholder managers if needed
            MessageManager = LeadsManager  # Temporary fallback
            AttachmentManager = LeadsManager  # Temporary fallback
            
        print("✅ Using enhanced managers with original fallbacks")
        return LeadsManager, SMTPManager, ProxyManager, SubjectManager, MessageManager, AttachmentManager, True
        
    except Exception as e:
        print(f"⚠️ Fallback to original managers: {e}")
        try:
            from ui.leads_manager import LeadsManager
            from ui.smtp_manager import SMTPManager
            from ui.proxy_manager import ProxyManager
            from ui.subject_manager import SubjectManager
            from ui.message_manager import MessageManager
            from ui.attachment_manager import AttachmentManager
            return LeadsManager, SMTPManager, ProxyManager, SubjectManager, MessageManager, AttachmentManager, False
        except Exception as e2:
            print(f"❌ Error importing original managers: {e2}")
            raise e2

# Import managers
LeadsManager, SMTPManager, ProxyManager, SubjectManager, MessageManager, AttachmentManager, using_improved = import_managers()

# These remain unchanged for now
from ui.campaign_builder import CampaignBuilder
from ui.settings_panel import SettingsPanel

# Import foundation components for logging
from core.utils.logger import get_module_logger

logger = get_module_logger(__name__)

# --- StatCard ---
# (Remains the same - no changes needed here)
class StatCard(QFrame):
    clicked = pyqtSignal()
    # Keep track of the label for identification
    def __init__(self, icon: QIcon, list_count: int, item_count: int, label: str, parent=None):
        super().__init__(parent); self.setObjectName("statCard"); self.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed); self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.card_label = label # Store the label
        layout = QVBoxLayout(self); layout.setSpacing(6); layout.setContentsMargins(12, 12, 12, 12); layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ico_lbl = QLabel();
        if not icon.isNull(): ico_lbl.setPixmap(icon.pixmap(QSize(48, 48)))
        else: ico_lbl.setText("?"); ico_lbl.setFixedSize(48, 48); ico_lbl.setStyleSheet("border: 1px solid gray; font-size: 24px;"); ico_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ico_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter); ico_lbl.setFixedSize(48, 48); layout.addWidget(ico_lbl)
        self.cnt_lbl = QLabel(); self.cnt_lbl.setObjectName("statCountLabel"); self.cnt_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.update_counts(list_count, item_count); layout.addWidget(self.cnt_lbl)
        self.lab_lbl = QLabel(label); self.lab_lbl.setObjectName("statTextLabel"); self.lab_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter); layout.addWidget(self.lab_lbl)
    def update_counts(self, list_count: int, item_count: int):
        plain_text = f"{list_count} ({item_count})"; self.cnt_lbl.setText(plain_text)
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton: self.clicked.emit(); event.accept()
        else: super().mousePressEvent(event)

# --- DashboardChart ---
# (Remains the same - no changes needed here)
class DashboardChart(FigureCanvas):
    def __init__(self, data_dir: str, parent=None):
        self.fig, self.ax = plt.subplots(figsize=(6, 3), tight_layout=True); super().__init__(self.fig); self.data_dir = data_dir; self.setParent(parent)
        self.fig.patch.set_alpha(0.0); self.ax.patch.set_alpha(0.0); self.plot()
    def plot(self):
        self.fig.patch.set_alpha(0.0); self.ax.patch.set_alpha(0.0); stats = {
            'Leads': len(glob.glob(os.path.join(self.data_dir,'leads','*.xlsx'))) if os.path.exists(os.path.join(self.data_dir,'leads')) else 0,
            'SMTPs': len(glob.glob(os.path.join(self.data_dir,'smtps','*.xlsx'))) if os.path.exists(os.path.join(self.data_dir,'smtps')) else 0,
            'Subjects': len(glob.glob(os.path.join(self.data_dir,'subjects','*.txt'))) if os.path.exists(os.path.join(self.data_dir,'subjects')) else 0,
            # *** NOTE: This chart still counts list folders, not message folders/files ***
            # This chart logic is separate from the StatCard update logic.
            # Fixing the chart requires changing this part specifically if desired.
            'Messages': len([d for d in os.listdir(os.path.join(self.data_dir,'messages')) if os.path.isdir(os.path.join(self.data_dir,'messages',d))]) if os.path.exists(os.path.join(self.data_dir,'messages')) else 0,
            'Attachments': len([d for d in os.listdir(os.path.join(self.data_dir,'attachments')) if os.path.isdir(os.path.join(self.data_dir,'attachments',d))]) if os.path.exists(os.path.join(self.data_dir,'attachments')) else 0,
            'Proxies': len(glob.glob(os.path.join(self.data_dir,'proxies','*.txt'))) if os.path.exists(os.path.join(self.data_dir,'proxies')) else 0 }
        running = scheduled = 0; campaigns_dir = os.path.join(self.data_dir,'campaigns')
        if os.path.isdir(campaigns_dir):
            for name in os.listdir(campaigns_dir):
                 camp_path = os.path.join(campaigns_dir,name); summary_file = os.path.join(camp_path,'summary.json')
                 if os.path.isdir(camp_path) and os.path.isfile(summary_file):
                     try:
                         with open(summary_file,'r',encoding='utf-8') as f: st = json.load(f).get('status','').lower()
                         if st == 'running': running += 1
                         elif st == 'scheduled': scheduled += 1
                     except Exception as e: print(f"W: Sum {summary_file}: {e}")
        stats['Running'] = running; stats['Scheduled'] = scheduled; keys, vals = list(stats.keys()), list(stats.values())
        self.ax.clear(); self.ax.patch.set_alpha(0.0)
        if keys and vals:
            bars = self.ax.bar(keys, vals, color='#4A90E2'); self.ax.set_ylabel('Count', fontdict={'fontsize':12, 'color': '#333333'})
            self.ax.tick_params(axis='x', rotation=30, labelsize=10, colors='#333333'); self.ax.tick_params(axis='y', labelsize=10, colors='#333333')
            spine_color = '#BCC3D4'; self.ax.spines['top'].set_color(spine_color); self.ax.spines['right'].set_color(spine_color)
            self.ax.spines['bottom'].set_color(spine_color); self.ax.spines['left'].set_color(spine_color)
            for bar in bars:
                yval = bar.get_height();
                if yval > 0: self.ax.text(bar.get_x()+bar.get_width()/2.0, yval, int(yval), va='bottom', ha='center', fontsize=9, color='#333333')
        else: self.ax.text(0.5,0.5,'No data available',ha='center',va='center',transform=self.ax.transAxes,color='#888888')
        self.ax.margins(y=0.1); self.draw()

# --- Dashboard Widget ---
class DashboardWidget(QWidget):
    card_clicked = pyqtSignal(str)
    refreshStarted = pyqtSignal()
    refreshFinished = pyqtSignal()
    def __init__(self, base_path: str, data_dir: str, parent=None):
        super().__init__(parent); self.base_path = base_path; self.data_dir = data_dir; self.setObjectName("dashboardWidget")
        # *** Store cards in a dictionary for easy access by label ***
        self.cards = {}
        self._build_ui()

    # (Helper functions _count_excel_rows, _count_text_lines remain the same)
    def _count_excel_rows(self, filepath: str) -> int:
        count = 0;
        try: wb = openpyxl.load_workbook(filepath, read_only=True); sheet = wb.active; count = sheet.max_row - 1 if sheet.max_row > 0 else 0; wb.close()
        except FileNotFoundError: print(f"W: File {filepath}"); pass
        except Exception as e: print(f"W: Excel {os.path.basename(filepath)}: {e}"); pass
        return max(0, count)
    def _count_text_lines(self, filepath: str) -> int:
        count = 0;
        try:
            with open(filepath, 'r', encoding='utf-8') as f: count = sum(1 for line in f if line.strip())
        except FileNotFoundError: print(f"W: File {filepath}"); pass
        except Exception as e: print(f"W: Text {os.path.basename(filepath)}: {e}"); pass
        return count

    # *** MODIFIED: _count_folder_items is NO LONGER USED by Messages card ***
    # It might still be used by Attachments, so we keep it for now.
    def _count_folder_items(self, folderpath: str) -> int:
        """Counts files directly inside a folder (not recursive)."""
        count = 0;
        try:
            # Count only files, not subdirectories
            count = sum(1 for item in os.listdir(folderpath) if os.path.isfile(os.path.join(folderpath, item)))
        except FileNotFoundError: print(f"W: Folder {folderpath}"); pass
        except Exception as e: print(f"W: Folder {os.path.basename(folderpath)}: {e}"); pass
        return count

    def _build_ui(self):
        layout = QVBoxLayout(self); layout.setContentsMargins(15, 15, 15, 15); layout.setSpacing(15)
        base_assets = os.path.join(self.base_path, 'assets', 'icons')
        if not os.path.isdir(self.data_dir): layout.addWidget(QLabel("Error: Data directory not found!")); return

        # Define specs including a placeholder for messages count function initially
        self.specs_data = [
            ("leads.ico", "Leads", lambda: (l:=glob.glob(os.path.join(self.data_dir,'leads','*.xlsx')), sum(self._count_excel_rows(f) for f in l))),
            ("smtp.ico", "SMTPs", lambda: (l:=glob.glob(os.path.join(self.data_dir,'smtps','*.xlsx')), sum(self._count_excel_rows(f) for f in l))),
            ("subject.ico", "Subjects", lambda: (l:=glob.glob(os.path.join(self.data_dir,'subjects','*.txt')), sum(self._count_text_lines(f) for f in l))),
            ("message.ico", "Messages", None), # <<< Use None initially, will be updated by signal
            ("attachment.ico", "Attachments", lambda: (l:=[os.path.join(self.data_dir,'attachments',d) for d in os.listdir(os.path.join(self.data_dir,'attachments')) if os.path.isdir(os.path.join(self.data_dir,'attachments',d))] if os.path.exists(os.path.join(self.data_dir,'attachments')) else [], sum(self._count_folder_items(f) for f in l))), # Attachments still uses _count_folder_items
            ("proxy.ico", "Proxies", lambda: (l:=glob.glob(os.path.join(self.data_dir,'proxies','*.txt')), sum(self._count_text_lines(f) for f in l))),
        ]

        cards_container = QWidget(); cards_container.setObjectName("cardsContainer"); grid = QGridLayout(cards_container); grid.setSpacing(15);
        # self.cards dictionary was initialized in __init__
        num_cols = 3
        for idx, (icon_file, label, count_fn) in enumerate(self.specs_data):
            icon_path = os.path.join(base_assets, icon_file); icon = QIcon(icon_path) if os.path.exists(icon_path) else QIcon();
            list_count, item_count = 0, 0 # Default counts
            # Calculate initial counts only if count_fn is provided
            if count_fn:
                try: list_data, item_count = count_fn(); list_count = len(list_data)
                except Exception as e: print(f"E: Calc initial counts {label}: {e}")

            card = StatCard(icon, list_count, item_count, label); card.clicked.connect(lambda lbl=label: self._on_stat_card_clicked(lbl));
            # *** Store card in dictionary using its label as key ***
            self.cards[label] = card
            r, c = divmod(idx, num_cols); grid.addWidget(card, r, c)

        for c in range(num_cols): grid.setColumnStretch(c, 1)
        layout.addWidget(cards_container)

        chart_container = QWidget(); chart_container.setObjectName("chartContainer"); chart_layout = QVBoxLayout(chart_container); chart_layout.setContentsMargins(0,0,0,0)
        self.chart = DashboardChart(self.data_dir, parent=chart_container); chart_layout.addWidget(self.chart); layout.addWidget(chart_container)

    def _on_stat_card_clicked(self, label: str): print(f"Stat card clicked: {label}"); self.card_clicked.emit(label)

    def refresh(self):
        """Refreshes stats, EXCLUDING Messages card (updated by signal)."""
        print("Dashboard refresh started...")
        self.refreshStarted.emit(); QApplication.processEvents()
        if hasattr(self, 'cards'):
             # Iterate through specs to find the right function for each card
             for icon_file, label, count_fn in self.specs_data:
                 # *** Skip the 'Messages' card - it's updated by signal ***
                 if label == "Messages" or not count_fn:
                     continue

                 card = self.cards.get(label)
                 if card:
                     try:
                         list_data, item_count = count_fn()
                         list_count = len(list_data)
                         card.update_counts(list_count, item_count)
                     except Exception as e:
                         print(f"E: Refresh card counts for '{label}': {e}")
                 else:
                     print(f"W: Card not found for label '{label}' during refresh.")

        if hasattr(self, 'chart'):
            try: self.chart.plot()
            except Exception as e: print(f"E: Refresh chart: {e}")
        print("Dashboard refresh finished.")
        self.refreshFinished.emit()

    # *** ADDED: Method to specifically update a card by label ***
    def update_card_by_label(self, label: str, list_count: int, item_count: int):
        """Updates counts for a specific StatCard identified by its label."""
        card = self.cards.get(label)
        if card:
            print(f"Updating card '{label}' with counts: {list_count}, {item_count}")
            card.update_counts(list_count, item_count)
        else:
            print(f"W: Attempted to update non-existent card with label: {label}")


# --- MainWindow ---
class MainWindow(QMainWindow):
    NAV_MAP = { "Dashboard": 0, "Leads": 1, "SMTPs": 2, "Subjects": 3, "Messages": 4, "Attachments": 5, "Proxies": 6, "Campaigns": 7, "Settings": 8 }
    def __init__(self, base_path: str, config: dict):
        super().__init__(); self.base_path = base_path; self.config = config; self.data_dir = os.path.join(self.base_path, 'data')
        self.config_path = os.path.join(self.data_dir, 'config', 'settings.json'); self.themes_dir = os.path.join(self.base_path, 'assets', 'themes')
        os.makedirs(self.themes_dir, exist_ok=True); self.setWindowTitle("Bulk Email Sender"); self.resize(1600, 950); self._set_icon();

        # --- Build UI FIRST ---
        self._build_ui() # This creates self.stack, self.dashboard_widget, etc.

        # --- Connect Signals AFTER UI is built ---
        if hasattr(self, 'dashboard_widget') and self.dashboard_widget:
             self.dashboard_widget.card_clicked.connect(self._navigate_from_dashboard)
             self.dashboard_widget.refreshStarted.connect(self._show_loading_indicator)
             self.dashboard_widget.refreshFinished.connect(self._hide_loading_indicator)
        else: print("E: Dashboard widget not found post-build.")

        # *** Connect manager signals if available ***
        if hasattr(self, 'message_manager') and self.message_manager:
            logger.info("Connecting MessageManager signal if available")
            # Signal is already connected above during widget creation if available
        else:
            logger.warning("MessageManager not found, cannot connect signal")

        # Setup refresh timer for dashboard (excluding messages card now)
        if isinstance(self.dashboard_widget, DashboardWidget):
            self.refresh_timer = QTimer(self); self.refresh_interval_ms = 15000 # 15 seconds
            self.refresh_timer.timeout.connect(self.dashboard_widget.refresh)
            self.refresh_timer.start(self.refresh_interval_ms); print(f"Dashboard refresh timer started ({self.refresh_interval_ms} ms).")
        else: print(f"W: Dashboard widget not found. Refresh timer not started.")

    # (_set_icon, _navigate_from_dashboard, _show_loading_indicator,
    #  _hide_loading_indicator remain the same)
    def _set_icon(self):
        icon_path = os.path.join(self.base_path, 'assets', 'icons', 'logo.ico')
        if os.path.exists(icon_path): self.setWindowIcon(QIcon(icon_path))
        else: self.setWindowIcon(QApplication.style().standardIcon(QApplication.style().StandardPixmap.SP_ComputerIcon))
    def _navigate_from_dashboard(self, card_label: str):
        print(f"Navigating from dashboard click: {card_label}"); target_index = self.NAV_MAP.get(card_label)
        if target_index is not None and 0 <= target_index < self.nav.count():
             if self.nav.currentRow() == target_index: print(f"Already on {card_label} tab."); return
             self.nav.setCurrentRow(target_index); print(f"Switched navigation to index {target_index} ({card_label})")
        else: print(f"W: Could not find navigation index for card label '{card_label}'.")
    def _show_loading_indicator(self):
        if hasattr(self, 'loading_label') and self.loading_label:
             self.loading_label.show()
             if hasattr(self, 'loading_movie') and self.loading_movie: self.loading_movie.start()
        if hasattr(self, 'refresh_button'): self.refresh_button.setEnabled(False)
    def _hide_loading_indicator(self):
        if hasattr(self, 'loading_label') and self.loading_label:
             self.loading_label.hide()
             if hasattr(self, 'loading_movie') and self.loading_movie: self.loading_movie.stop()
        if hasattr(self, 'refresh_button'): self.refresh_button.setEnabled(True)

    def _trigger_global_refresh(self):
        """Triggers refresh for dashboard (excluding messages) and current manager list."""
        print("Global refresh triggered.")
        if hasattr(self, 'dashboard_widget') and self.dashboard_widget:
            # Dashboard refresh now skips messages internally
            self.dashboard_widget.refresh()
        else:
            print("W: Cannot refresh dashboard, widget not found.")

        # Refresh Currently Visible Data Manager List (if applicable)
        current_widget = self.stack.currentWidget()
        if hasattr(current_widget, '_refresh_list') and callable(current_widget._refresh_list):
            print(f"Refreshing list for: {current_widget.__class__.__name__}")
            try:
                current_widget._refresh_list() # This will trigger counts_changed if it's MessageManager
            except Exception as e:
                print(f"E: Failed to call _refresh_list for {current_widget.__class__.__name__}: {e}")
        else:
            print(f"Current widget ({current_widget.__class__.__name__}) has no _refresh_list method.")

    def _build_ui(self):
        # Create responsive main layout
        root = QWidget()
        self.setCentralWidget(root)
        layout = QHBoxLayout(root)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Create main splitter for responsive design
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(1)
        
        # Enhanced sidebar with responsive sizing
        sidebar_widget = QWidget()
        sidebar_widget.setObjectName("sidebarWidget")
        sidebar_layout = QVBoxLayout(sidebar_widget)
        sidebar_layout.setContentsMargins(10, 10, 10, 10)
        sidebar_layout.setSpacing(8)
        
        # Enhanced navigation list with proper sizing
        self.nav = QListWidget()
        self.nav.setObjectName("navigationList")
        self.nav.setMinimumWidth(200)
        self.nav.setMaximumWidth(350)
        self.nav.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)
        self.nav.setIconSize(QSize(20, 20))
        # Navigation items with better organization
        self.nav_items_in_order = [
            "Dashboard", "Leads", "SMTPs", "Subjects", 
            "Messages", "Attachments", "Proxies", "Campaigns", "Settings"
        ]
        icon_files = [
            "dashboard.ico", "leads.ico", "smtp.ico", "subject.ico", 
            "message.ico", "attachment.ico", "proxy.ico", "campaign.ico", "settings.ico"
        ]
        self.NAV_MAP = {label: index for index, label in enumerate(self.nav_items_in_order)}
        icon_base = os.path.join(self.base_path, 'assets', 'icons')
        
        # Add navigation items with proper sizing
        for i, text in enumerate(self.nav_items_in_order):
            icon_path = os.path.join(icon_base, icon_files[i])
            if not os.path.exists(icon_path):
                print(f"W: Icon missing '{icon_path}'. Using default.")
                icon = QApplication.style().standardIcon(
                    QApplication.style().StandardPixmap.SP_FileDialogDetailedView 
                    if text == "Settings" else QApplication.style().StandardPixmap.SP_FileIcon
                )
            else:
                icon = QIcon(icon_path)
            
            item = QListWidgetItem(icon, f"  {text}")  # Add spacing for better appearance
            item.setSizeHint(QSize(0, 44))  # Consistent height, auto width
            item.setToolTip(text)
            self.nav.addItem(item)
        
        sidebar_layout.addWidget(self.nav)
        
        # Enhanced refresh section with better layout
        refresh_section = QFrame()
        refresh_section.setFrameStyle(QFrame.Shape.StyledPanel)
        refresh_section.setMaximumHeight(60)
        refresh_layout = QHBoxLayout(refresh_section)
        refresh_layout.setContentsMargins(8, 8, 8, 8)
        refresh_layout.setSpacing(8)
        
        # Refresh button with better styling
        self.refresh_button = QPushButton("🔄 Refresh")
        self.refresh_button.setObjectName("sidebarRefreshButton")
        self.refresh_button.setToolTip("Refresh dashboard counts and current list view")
        self.refresh_button.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.refresh_button.clicked.connect(self._trigger_global_refresh)
        self.refresh_button.setMinimumHeight(32)
        
        # Loading indicator
        self.loading_label = QLabel()
        self.loading_label.setObjectName("loadingIndicator")
        loading_gif_path = os.path.join(self.base_path, 'assets', 'icons', 'loading.gif')
        if os.path.exists(loading_gif_path):
            self.loading_movie = QMovie(loading_gif_path)
            self.loading_movie.setScaledSize(QSize(16, 16))
            self.loading_label.setMovie(self.loading_movie)
        else:
            print("W: loading.gif not found. Loading indicator disabled.")
            self.loading_movie = None
        self.loading_label.setFixedSize(16, 16)
        self.loading_label.hide()
        
        refresh_layout.addWidget(self.refresh_button)
        refresh_layout.addWidget(self.loading_label)
        refresh_layout.addStretch()
        
        sidebar_layout.addWidget(refresh_section)
        
        # Set minimum sidebar width to prevent text cutoff
        sidebar_widget.setMinimumWidth(220)
        sidebar_widget.setMaximumWidth(380)
        
        # Add sidebar to splitter
        splitter.addWidget(sidebar_widget)
        
        # Create content area with proper responsive design
        self.stack = QStackedWidget()
        self.stack.setObjectName("contentStack")
        self.stack.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        # Ensure dashboard widget is created
        self.dashboard_widget = DashboardWidget(self.base_path, self.data_dir)
        self.stack.addWidget(self.dashboard_widget)
        
        # Add content area to splitter
        splitter.addWidget(self.stack)
        
        # Set splitter proportions for responsive design
        splitter.setSizes([280, 1000])  # Sidebar: content ratio
        splitter.setStretchFactor(0, 0)  # Sidebar doesn't stretch
        splitter.setStretchFactor(1, 1)  # Content area stretches
        
        # Add splitter to main layout
        layout.addWidget(splitter)
        
        # Add managers using improved threaded versions
        logger.info("Initializing UI managers")
        
        # Leads Manager (Improved with threading)
        self.leads_manager = LeadsManager()
        if hasattr(self.leads_manager, 'counts_changed'):
            self.leads_manager.counts_changed.connect(self._update_leads_stats)
        self.stack.addWidget(self.leads_manager)
        
        # SMTP Manager (Improved with threading)
        self.smtp_manager = SMTPManager()
        if hasattr(self.smtp_manager, 'counts_changed'):
            self.smtp_manager.counts_changed.connect(self._update_smtp_stats)
        self.stack.addWidget(self.smtp_manager)
        
        # Subject Manager (Improved with threading)
        self.subject_manager = SubjectManager()
        if hasattr(self.subject_manager, 'counts_changed'):
            self.subject_manager.counts_changed.connect(self._update_subject_stats)
        self.stack.addWidget(self.subject_manager)
        
        # Message Manager (Improved with threading)
        self.message_manager = MessageManager()
        if hasattr(self.message_manager, 'counts_changed'):
            self.message_manager.counts_changed.connect(self._update_message_dashboard_count)
        self.stack.addWidget(self.message_manager)
        
        # Attachment Manager (Improved with threading)
        self.attachment_manager = AttachmentManager()
        if hasattr(self.attachment_manager, 'counts_changed'):
            self.attachment_manager.counts_changed.connect(self._update_attachment_stats)
        self.stack.addWidget(self.attachment_manager)
        
        # Proxy Manager (Improved with threading)
        self.proxy_manager = ProxyManager()
        if hasattr(self.proxy_manager, 'counts_changed'):
            self.proxy_manager.counts_changed.connect(self._update_proxy_stats)
        self.stack.addWidget(self.proxy_manager)
        
        # Campaign Builder (Original)
        self.campaign_builder = CampaignBuilder()
        self.stack.addWidget(self.campaign_builder)
        
        # Settings Panel
        self.settings_panel = SettingsPanel(self.base_path, self.config, self.config_path)
        self.stack.addWidget(self.settings_panel)
        
        # Connect navigation to stack
        self.nav.currentRowChanged.connect(self.stack.setCurrentIndex)
        self.nav.setCurrentRow(0)

    # *** ADDED: Slot to receive signal from MessageManager ***
    def _update_message_dashboard_count(self, list_count: int, total_messages: int):
        """Receives counts from MessageManager and updates the dashboard card."""
        logger.debug("Received message counts", lists=list_count, messages=total_messages)
        if hasattr(self, 'dashboard_widget') and self.dashboard_widget:
            # Call the new update method in DashboardWidget
            self.dashboard_widget.update_card_by_label("Messages", list_count, total_messages)
        else:
            logger.warning("Dashboard widget not available to update message count")
    
    # *** ADDED: Slots for manager stats (some managers may not have these signals) ***
    def _update_leads_stats(self, list_count: int, total_leads: int):
        """Receives counts from LeadsManager and updates the dashboard card."""
        logger.debug("Received leads counts", lists=list_count, leads=total_leads)
        if hasattr(self, 'dashboard_widget') and self.dashboard_widget:
            self.dashboard_widget.update_card_by_label("Leads", list_count, total_leads)
        else:
            logger.warning("Dashboard widget not available to update leads count")
    
    def _update_smtp_stats(self, list_count: int, total_smtps: int):
        """Receives counts from SMTPManager and updates the dashboard card."""
        logger.debug("Received SMTP counts", lists=list_count, smtps=total_smtps)
        if hasattr(self, 'dashboard_widget') and self.dashboard_widget:
            self.dashboard_widget.update_card_by_label("SMTPs", list_count, total_smtps)
        else:
            logger.warning("Dashboard widget not available to update SMTP count")
    
    def _update_subject_stats(self, list_count: int, total_subjects: int):
        """Receives counts from SubjectManager and updates the dashboard card."""
        logger.debug("Received subject counts", lists=list_count, subjects=total_subjects)
        if hasattr(self, 'dashboard_widget') and self.dashboard_widget:
            self.dashboard_widget.update_card_by_label("Subjects", list_count, total_subjects)
        else:
            logger.warning("Dashboard widget not available to update subject count")
    
    def _update_attachment_stats(self, list_count: int, total_attachments: int):
        """Receives counts from AttachmentManager and updates the dashboard card."""
        logger.debug("Received attachment counts", lists=list_count, attachments=total_attachments)
        if hasattr(self, 'dashboard_widget') and self.dashboard_widget:
            self.dashboard_widget.update_card_by_label("Attachments", list_count, total_attachments)
        else:
            logger.warning("Dashboard widget not available to update attachment count")
    
    def _update_proxy_stats(self, list_count: int, total_proxies: int):
        """Receives counts from ProxyManager and updates the dashboard card."""
        logger.debug("Received proxy counts", lists=list_count, proxies=total_proxies)
        if hasattr(self, 'dashboard_widget') and self.dashboard_widget:
            self.dashboard_widget.update_card_by_label("Proxies", list_count, total_proxies)
        else:
            logger.warning("Dashboard widget not available to update proxy count")

    # (closeEvent remains the same)
    def closeEvent(self, event):
         if hasattr(self, 'refresh_timer') and self.refresh_timer.isActive(): self.refresh_timer.stop()
         print("Stopping timers and closing.")
         if hasattr(QApplication.instance(), 'tray_icon') and QApplication.instance().tray_icon: QApplication.instance().tray_icon.hide()
         event.accept()

# (main.py execution part remains the same, ensure it imports this modified MainWindow)