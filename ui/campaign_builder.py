# ui/campaign_builder.py
import os
import json
import shutil
import random # <-- Ensures import is present
from datetime import datetime
from PyQt6.QtWidgets import (
    QWidget, QLabel, QPushButton, QComboBox, QRadioButton, QGroupBox,
    QHBoxLayout, QVBoxLayout, QGridLayout, QMessageBox, QSpinBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QProgressBar, QTextEdit,
    QSplitter, QListWidget, QInputDialog, QSpacerItem, QSizePolicy,
    QFormLayout # <-- Ensures import is present
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from engine.utils import replace_placeholders
from engine.sender import send_email
from engine.scheduler import (
    generate_schedule_no_delay, generate_schedule_custom_delay,
    generate_schedule_batch, generate_schedule_spike
)

BASE_PATH = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
DATA_DIR = os.path.join(BASE_PATH, 'data')
CAMPAIGNS_DIR = os.path.join(DATA_DIR, 'campaigns')
CONFIG_FILENAME = "campaign_config.json"

# SendWorker class (no changes from previous version)
class SendWorker(QThread):
    progress = pyqtSignal(int, int); finished = pyqtSignal(); log = pyqtSignal(str)
    def __init__(self, tasks):
        super().__init__()
        self.tasks = tasks if tasks is not None else []
    def run(self):
        total = len(self.tasks); sent = 0
        if not self.tasks:
            self.log.emit("No tasks to process.")
            self.finished.emit()
            return
        for task in self.tasks:
            if not task or 'args' not in task:
                self.log.emit(f"Skipping invalid task: {task}")
                continue
            args = task['args']
            if not args or 'smtp' not in args or 'msg' not in args:
                self.log.emit(f"Skipping task with invalid args: {args}")
                continue
            to_addr = args['msg'].get('to', 'Unknown')
            timestamp = datetime.now().strftime("%H:%M:%S")
            try:
                success = send_email(
                    smtp_conf=args['smtp'],
                    msg_conf=args['msg'],
                    proxy_conf=args.get('proxy')
                )
                sent += 1; status = "✅ OK" if success else "❌ FAIL"
                self.log.emit(f"[{timestamp}] {status}: to={to_addr}")
            except Exception as e:
                sent += 1
                status = "❌ ERROR"
                self.log.emit(f"[{timestamp}] {status}: to={to_addr} - Error: {e}")
            finally:
                self.progress.emit(sent, total)
        self.finished.emit()

class CampaignBuilder(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        os.makedirs(CAMPAIGNS_DIR, exist_ok=True)
        self.current_campaign_name = None
        self.worker = None
        # --- Store the Save Config button instance ---
        self.btn_save_config = None
        self._build_ui()
        self._refresh_campaign_list()

    def _build_ui(self):
        main_layout = QHBoxLayout(self); main_layout.setContentsMargins(0, 0, 0, 0)
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # --- Left Panel (Campaign List & Management) ---
        left_widget = QWidget(); left_layout = QVBoxLayout(left_widget); left_layout.setContentsMargins(5, 5, 5, 5); left_layout.setSpacing(5)
        left_layout.addWidget(QLabel("<b>Campaigns</b>"))
        self.campaign_list = QListWidget(); self.campaign_list.setObjectName("campaignListWidget")
        self.campaign_list.currentItemChanged.connect(self._on_campaign_selected)
        left_layout.addWidget(self.campaign_list)

        list_button_layout = QGridLayout()
        btn_new = QPushButton("＋ New"); btn_new.setToolTip("Create new campaign"); btn_new.clicked.connect(self._new_campaign)
        # MOVED: Save Config button is no longer created or added here
        # btn_save_config = QPushButton("💾 Save Config"); btn_save_config.setToolTip("Save current campaign configuration"); btn_save_config.clicked.connect(self._save_campaign)
        btn_del = QPushButton("🗑️ Delete"); btn_del.setToolTip("Delete selected campaign"); btn_del.clicked.connect(self._delete_campaign)
        list_button_layout.addWidget(btn_new, 0, 0);
        # ADDED: Delete button now next to New
        list_button_layout.addWidget(btn_del, 0, 1);
        # Removed row 1 as Save Config is gone
        left_layout.addLayout(list_button_layout)
        splitter.addWidget(left_widget)

        # --- Right Panel (Configuration & Actions) ---
        right_widget = QWidget(); right_layout = QVBoxLayout(right_widget); right_layout.setSpacing(12); right_layout.setContentsMargins(15, 10, 15, 10); right_widget.setObjectName("campaignConfigPanel")
        right_layout.addWidget(QLabel("<h2>Configure Campaign</h2>"))

        # Data Lists Group (using QFormLayout)
        data_list_group = QGroupBox("Data Lists")
        form_layout = QFormLayout(data_list_group)
        form_layout.setRowWrapPolicy(QFormLayout.RowWrapPolicy.WrapAllRows)
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        form_layout.setContentsMargins(10, 15, 10, 10)
        form_layout.setSpacing(10)
        cats = ["leads", "smtps", "subjects", "messages", "attachments", "proxies"]; labels = ["Lead List:", "SMTP List:", "Subject List:", "Message List:", "Attachment List (opt):", "Proxy List (opt):"]; self.combos = {}
        for i, cat in enumerate(cats):
            cb = QComboBox(); cb.setMinimumWidth(200)
            items = self._get_lists(cat); is_optional = cat in ('attachments', 'proxies')
            if is_optional: cb.addItem("None")
            cb.addItems(items); self.combos[cat] = cb
            if not items and not is_optional: cb.addItem("No lists found!")
            form_layout.addRow(QLabel(labels[i]), cb)
        right_layout.addWidget(data_list_group)

        # Sending Mode (using QComboBox)
        sending_mode_layout = QHBoxLayout()
        sending_mode_layout.addWidget(QLabel("<b>Sending Mode:</b>"))
        self.sending_mode_combo = QComboBox()
        self.sending_mode_combo.addItems(["No Delay", "Custom Delay", "Batch Mode", "Spike Mode"])
        self.sending_mode_combo.setToolTip("Select how emails should be scheduled")
        self.sending_mode_combo.currentTextChanged.connect(self._mode_toggled)
        sending_mode_layout.addWidget(self.sending_mode_combo)
        sending_mode_layout.addStretch(1)
        right_layout.addLayout(sending_mode_layout)

        # --- Parameter Group Boxes ---
        # Custom Delay Parameters
        self.custom_delay_group = QGroupBox("Custom Delay Parameters"); delay_layout = QHBoxLayout(self.custom_delay_group)
        self.delay_min = QSpinBox(); self.delay_min.setRange(0, 3600); self.delay_min.setPrefix("Min Delay: "); self.delay_min.setSuffix(" s"); delay_layout.addWidget(self.delay_min)
        self.delay_max = QSpinBox(); self.delay_max.setRange(0, 3600); self.delay_max.setPrefix("Max Delay: "); self.delay_max.setSuffix(" s"); delay_layout.addWidget(self.delay_max)
        right_layout.addWidget(self.custom_delay_group)
        # Batch Mode Parameters
        self.batch_mode_group = QGroupBox("Batch Mode Parameters"); batch_layout = QHBoxLayout(self.batch_mode_group)
        self.batch_min = QSpinBox(); self.batch_min.setRange(1, 1000); self.batch_min.setPrefix("Min Batch Size: "); batch_layout.addWidget(self.batch_min)
        self.batch_max = QSpinBox(); self.batch_max.setRange(1, 1000); self.batch_max.setPrefix("Max Batch Size: "); batch_layout.addWidget(self.batch_max)
        self.batch_delay_min = QSpinBox(); self.batch_delay_min.setRange(0, 86400); self.batch_delay_min.setPrefix("Min Delay (Batches): "); self.batch_delay_min.setSuffix(" s"); batch_layout.addWidget(self.batch_delay_min)
        self.batch_delay_max = QSpinBox(); self.batch_delay_max.setRange(0, 86400); self.batch_delay_max.setPrefix("Max Delay (Batches): "); self.batch_delay_max.setSuffix(" s"); batch_layout.addWidget(self.batch_delay_max)
        right_layout.addWidget(self.batch_mode_group)
        # Spike Mode Configuration
        self.spike_group = QGroupBox("Spike Mode Configuration"); sg_layout = QVBoxLayout(self.spike_group)
        self.day_table = QTableWidget(0, 2); self.day_table.setHorizontalHeaderLabels(["Day", "Count"]); self.day_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        sg_layout.addWidget(self.day_table); btn_add = QPushButton("＋ Add Day"); btn_add.clicked.connect(self._add_day); sg_layout.addWidget(btn_add)
        right_layout.addWidget(self.spike_group)
        self._mode_toggled() # Call initially

        # --- Stretch and Bottom Elements ---
        right_layout.addStretch(1)

        # --- Action Buttons (Save Config, Launch) ---
        action_btns_layout = QHBoxLayout();

        # MOVED: Create and add Save Config button here
        self.btn_save_config = QPushButton("💾 Save Config") # Store instance if needed elsewhere
        self.btn_save_config.setToolTip("Save current campaign configuration")
        self.btn_save_config.clicked.connect(self._save_campaign)
        action_btns_layout.addWidget(self.btn_save_config) # Add Save Config first

        # Add some spacing between buttons
        action_btns_layout.addSpacing(10)

        # Launch Button
        btn_send = QPushButton("🚀 Launch Campaign"); btn_send.setObjectName("launchButton"); btn_send.setToolTip("Start sending emails for the selected campaign"); btn_send.clicked.connect(self._on_send)
        action_btns_layout.addWidget(btn_send) # Add Launch next

        action_btns_layout.addStretch(1) # Pushes buttons left
        right_layout.addLayout(action_btns_layout)

        # Progress Bar and Log View
        self.progress = QProgressBar(); self.progress.setVisible(False); self.progress.setTextVisible(True); right_layout.addWidget(self.progress)
        self.log_view = QTextEdit(); self.log_view.setObjectName("logView"); self.log_view.setReadOnly(True); self.log_view.setPlaceholderText("Campaign log will appear here..."); right_layout.addWidget(self.log_view)

        # --- Add Right Panel to Splitter ---
        splitter.addWidget(right_widget)
        splitter.setStretchFactor(0, 0); splitter.setStretchFactor(1, 1);
        splitter.setSizes([200, 800]) # Adjusted initial sizes slightly
        main_layout.addWidget(splitter)

    # --- Methods (_refresh_campaign_list, _get_lists, _update_list_combos, etc.) ---
    # --- No changes to the rest of the methods from the previous version ---
    # --- They remain the same as provided in the last code block ---

    def _refresh_campaign_list(self):
        self.campaign_list.blockSignals(True); self.campaign_list.clear()
        try:
            if os.path.isdir(CAMPAIGNS_DIR):
                campaign_names = sorted([d for d in os.listdir(CAMPAIGNS_DIR) if os.path.isdir(os.path.join(CAMPAIGNS_DIR, d))])
                self.campaign_list.addItems(campaign_names)
                print(f"Found campaigns: {campaign_names}")
            else:
                print(f"Campaign directory does not exist: {CAMPAIGNS_DIR}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not read campaign directory:\n{CAMPAIGNS_DIR}\n\n{e}")
        self.campaign_list.blockSignals(False)
        if self.campaign_list.count() == 0:
             self._clear_config_ui()
             self.current_campaign_name = None

    def _get_lists(self, cat):
        path = os.path.join(DATA_DIR, cat); items = []
        if os.path.isdir(path):
            try:
                for name in sorted(os.listdir(path)):
                    full_path = os.path.join(path, name);
                    is_list_file = cat in ('leads', 'smtps') and name.lower().endswith('.xlsx');
                    is_text_file = cat in ('subjects',) and name.lower().endswith('.txt');
                    is_folder = cat in ('messages', 'attachments', 'proxies') and os.path.isdir(full_path)
                    if is_list_file or is_text_file: items.append(os.path.splitext(name)[0])
                    elif is_folder: items.append(name)
            except Exception as e: print(f"W: Could not read {cat} list directory: {e}")
        return items

    def _update_list_combos(self):
        print("Updating list comboboxes...");
        selections = {cat: combo.currentText() for cat, combo in self.combos.items()}
        for cat, combo in self.combos.items():
            current_selection = selections.get(cat)
            combo.blockSignals(True)
            combo.clear()
            items = self._get_lists(cat)
            is_optional = cat in ('attachments', 'proxies')
            if is_optional: combo.addItem("None")
            if items: combo.addItems(items)
            elif not is_optional: combo.addItem("No lists found!")
            index = combo.findText(current_selection)
            if index != -1: combo.setCurrentIndex(index)
            elif combo.count() > 0: combo.setCurrentIndex(0)
            combo.blockSignals(False)

    def showEvent(self, event):
        print("CampaignBuilder showEvent triggered.")
        self._update_list_combos()
        self._refresh_campaign_list()
        if self.current_campaign_name:
            items = self.campaign_list.findItems(self.current_campaign_name, Qt.MatchFlag.MatchExactly)
            if items:
                 if self.campaign_list.currentItem() != items[0]:
                      print(f"Reselecting campaign: {self.current_campaign_name}")
                      self.campaign_list.blockSignals(True)
                      self.campaign_list.setCurrentItem(items[0])
                      self.campaign_list.blockSignals(False)
            else:
                 print(f"Previously selected campaign '{self.current_campaign_name}' not found. Clearing UI.")
                 self.current_campaign_name = None
                 self._clear_config_ui()
        else:
            if self.campaign_list.count() > 0:
                 print("No campaign selected, selecting first in list.")
                 self.campaign_list.setCurrentRow(0)
            else:
                 print("No campaigns available, clearing UI.")
                 self._clear_config_ui()
        super().showEvent(event)

    def _on_campaign_selected(self, current_item, previous_item):
        if current_item:
            name = current_item.text()
            print(f"Campaign selected: {name}")
            if name != self.current_campaign_name:
                self._load_campaign(name)
        else:
            print("Campaign selection cleared.")
            self._clear_config_ui()
            self.current_campaign_name = None

    def _new_campaign(self):
        name, ok = QInputDialog.getText(self, "New Campaign", "Enter campaign name:")
        if ok and name and name.strip():
            clean_name = name.strip(); path = os.path.join(CAMPAIGNS_DIR, clean_name)
            if os.path.exists(path):
                QMessageBox.warning(self, "Exists", f"Campaign '{clean_name}' already exists."); return
            try:
                os.makedirs(path)
                default_config = {
                     "leads": None, "smtps": None, "subjects": None, "messages": None,
                     "attachments": None, "proxies": None, "sending_mode": "No Delay"
                 }
                config_path = os.path.join(path, CONFIG_FILENAME)
                with open(config_path, 'w', encoding='utf-8') as f:
                     json.dump(default_config, f, indent=4)
                print(f"Created campaign '{clean_name}' with default config.")
                self._refresh_campaign_list();
                items = self.campaign_list.findItems(clean_name, Qt.MatchFlag.MatchExactly)
                if items: self.campaign_list.setCurrentItem(items[0])
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Could not create campaign folder/config:\n{e}")

    def _load_campaign(self, name: str):
        if not name: return
        print(f"Loading campaign: {name}")
        config_path = os.path.join(CAMPAIGNS_DIR, name, CONFIG_FILENAME); config_data = {}
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f: config_data = json.load(f)
            except Exception as e:
                QMessageBox.critical(self, "Load Error", f"Could not read config:\n{config_path}\n\n{e}")
                self._clear_config_ui(); self.current_campaign_name = name; return
        else:
            print(f"Config file not found for '{name}'. Using defaults and clearing UI.")
            self._clear_config_ui()
        self.current_campaign_name = name;
        print(f"Applying loaded config: {config_data}")
        for combo in self.combos.values(): combo.blockSignals(True)
        self.sending_mode_combo.blockSignals(True)
        for cat, combo in self.combos.items():
            value = config_data.get(cat, "None" if cat in ('attachments', 'proxies') else None)
            if value is None and cat in ('attachments', 'proxies'): value = "None"
            if value:
                 index = combo.findText(value)
                 if index != -1: combo.setCurrentIndex(index)
                 else: print(f"W: Saved value '{value}' for '{cat}' not found. Setting to default."); combo.setCurrentIndex(0)
            elif combo.count() > 0: combo.setCurrentIndex(0)
        mode = config_data.get("sending_mode", "No Delay");
        mode_index = self.sending_mode_combo.findText(mode)
        if mode_index != -1: self.sending_mode_combo.setCurrentIndex(mode_index)
        else: print(f"W: Saved sending mode '{mode}' not recognized. Defaulting to 'No Delay'."); self.sending_mode_combo.setCurrentIndex(0)
        self.delay_min.setValue(config_data.get("delay_min", 0));
        self.delay_max.setValue(config_data.get("delay_max", 5))
        self.batch_min.setValue(config_data.get("batch_min", 10));
        self.batch_max.setValue(config_data.get("batch_max", 20))
        self.batch_delay_min.setValue(config_data.get("batch_delay_min", 60));
        self.batch_delay_max.setValue(config_data.get("batch_delay_max", 120))
        spike_data = config_data.get("spike_days", []);
        self.day_table.setRowCount(0)
        self.day_table.setRowCount(len(spike_data))
        for r, count in enumerate(spike_data):
            self.day_table.setItem(r, 0, QTableWidgetItem(f"Day {r+1}"))
            try: count_str = str(int(count))
            except (ValueError, TypeError): count_str = "0"
            item = QTableWidgetItem(count_str);
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable);
            self.day_table.setItem(r, 1, item)
        self.day_table.resizeColumnsToContents()
        for combo in self.combos.values(): combo.blockSignals(False)
        self.sending_mode_combo.blockSignals(False)
        self._mode_toggled()

    def _clear_config_ui(self):
        print("Clearing config UI")
        for combo in self.combos.values(): combo.blockSignals(True)
        self.sending_mode_combo.blockSignals(True)
        for combo in self.combos.values():
             if combo.count() > 0: combo.setCurrentIndex(0)
        self.sending_mode_combo.setCurrentIndex(0)
        self.delay_min.setValue(0); self.delay_max.setValue(5)
        self.batch_min.setValue(10); self.batch_max.setValue(20)
        self.batch_delay_min.setValue(60); self.batch_delay_max.setValue(120)
        self.day_table.setRowCount(0)
        for combo in self.combos.values(): combo.blockSignals(False)
        self.sending_mode_combo.blockSignals(False)
        self._mode_toggled()
        self.log_view.clear()
        self.progress.setVisible(False)

    def _get_current_config_from_ui(self) -> dict:
        config = {}
        for cat, combo in self.combos.items():
            text = combo.currentText()
            if text == "None" or text == "No lists found!": config[cat] = None
            else: config[cat] = text
        mode = self.sending_mode_combo.currentText()
        config["sending_mode"] = mode
        if mode == "Custom Delay":
             config["delay_min"] = self.delay_min.value()
             config["delay_max"] = self.delay_max.value()
        elif mode == "Batch Mode":
             config["batch_min"] = self.batch_min.value()
             config["batch_max"] = self.batch_max.value()
             config["batch_delay_min"] = self.batch_delay_min.value()
             config["batch_delay_max"] = self.batch_delay_max.value()
        elif mode == "Spike Mode":
             spike_days = []
             for r in range(self.day_table.rowCount()):
                  item = self.day_table.item(r, 1)
                  count = 0
                  if item and item.text():
                      try: count = int(item.text())
                      except ValueError: count = 0
                  spike_days.append(max(0, count))
             config["spike_days"] = spike_days
        return config

    def _save_campaign(self):
        if not self.current_campaign_name:
            QMessageBox.warning(self, "No Campaign Selected", "Please select or create a campaign before saving."); return
        config_data = self._get_current_config_from_ui();
        config_path = os.path.join(CAMPAIGNS_DIR, self.current_campaign_name, CONFIG_FILENAME)
        try:
             os.makedirs(os.path.dirname(config_path), exist_ok=True)
             with open(config_path, 'w', encoding='utf-8') as f:
                  json.dump(config_data, f, indent=4)
             QMessageBox.information(self, "Saved", f"Campaign '{self.current_campaign_name}' configuration saved successfully.");
             print(f"Saved config to {config_path}")
        except Exception as e:
             QMessageBox.critical(self, "Save Error", f"Could not save configuration for '{self.current_campaign_name}':\n{e}")

    def _delete_campaign(self):
        if not self.current_campaign_name:
             QMessageBox.warning(self, "No Campaign Selected", "Please select the campaign you want to delete."); return
        name = self.current_campaign_name; path = os.path.join(CAMPAIGNS_DIR, name)
        if QMessageBox.question(self, "Confirm Delete", f"Are you sure you want to permanently delete the campaign '{name}' and its configuration?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            try:
                if os.path.isdir(path):
                    shutil.rmtree(path);
                    print(f"Deleted campaign folder: {path}");
                    self.current_campaign_name = None;
                    self._refresh_campaign_list();
                else:
                    QMessageBox.warning(self, "Not Found", f"Campaign folder not found (might have been deleted already):\n{path}");
                    self._refresh_campaign_list();
            except Exception as e:
                QMessageBox.critical(self, "Delete Error", f"Could not delete campaign '{name}':\n{e}")

    def _mode_toggled(self):
        if not hasattr(self, 'sending_mode_combo'): return
        selected_mode = self.sending_mode_combo.currentText()
        is_custom = (selected_mode == "Custom Delay")
        is_batch = (selected_mode == "Batch Mode")
        is_spike = (selected_mode == "Spike Mode")
        if hasattr(self, 'custom_delay_group'): self.custom_delay_group.setVisible(is_custom or is_batch)
        if hasattr(self, 'batch_mode_group'): self.batch_mode_group.setVisible(is_batch)
        if hasattr(self, 'spike_group'): self.spike_group.setVisible(is_spike)

    def _add_day(self):
        r = self.day_table.rowCount(); self.day_table.insertRow(r);
        day_item = QTableWidgetItem(f"Day {r+1}")
        day_item.setFlags(day_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.day_table.setItem(r, 0, day_item)
        count_item = QTableWidgetItem("0");
        count_item.setFlags(count_item.flags() | Qt.ItemFlag.ItemIsEditable);
        self.day_table.setItem(r, 1, count_item)

    # --- Data Loading --- (No changes from previous version)
    def _load_data_from_selection(self, list_name, category):
        print(f"--- Placeholder: Loading data for '{list_name}' ({category}) ---")
        data_path_base = os.path.join(DATA_DIR, category)
        if not list_name or list_name == "None": return []
        if category in ('leads', 'smtps'):
            file_path = os.path.join(data_path_base, f"{list_name}.xlsx")
            if not os.path.exists(file_path): print(f"W: File not found: {file_path}"); return []
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename=file_path, read_only=True); ws = wb.active; header = [cell.value for cell in ws[1]]; data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                     if not any(row): continue; item_dict = dict(zip(header, row))
                     if category == 'smtps' and 'Port' in item_dict:
                         try: item_dict['Port'] = int(item_dict['Port'])
                         except (ValueError, TypeError): item_dict['Port'] = None
                     data.append(item_dict)
                wb.close(); print(f"Loaded {len(data)} items from {file_path}"); return data
            except Exception as e: print(f"E: Failed to load Excel file {file_path}: {e}"); return []
        elif category in ('subjects', 'proxies'):
            file_path = os.path.join(data_path_base, f"{list_name}.txt")
            if not os.path.exists(file_path): print(f"W: File not found: {file_path}"); return []
            try:
                with open(file_path, 'r', encoding='utf-8') as f: lines = [line.strip() for line in f if line.strip()]
                print(f"Loaded {len(lines)} lines from {file_path}"); return lines
            except Exception as e: print(f"E: Failed to load Text file {file_path}: {e}"); return []
        elif category in ('messages', 'attachments'):
            folder_path = os.path.join(data_path_base, list_name)
            if not os.path.isdir(folder_path): print(f"W: Folder not found: {folder_path}"); return []
            try:
                items = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
                print(f"Found {len(items)} files in {folder_path}"); return items
            except Exception as e: print(f"E: Failed to list folder {folder_path}: {e}"); return []
        return []

    def _collect_tasks(self):
        if not self.current_campaign_name: QMessageBox.warning(self, "Error", "No campaign selected. Cannot collect tasks."); return []
        print(f"Collecting tasks for campaign: {self.current_campaign_name}"); config_data = self._get_current_config_from_ui(); print(f"Using configuration: {config_data}")
        try:
            leads = self._load_data_from_selection(config_data.get('leads'), 'leads'); smtps = self._load_data_from_selection(config_data.get('smtps'), 'smtps')
            subjects = self._load_data_from_selection(config_data.get('subjects'), 'subjects'); message_files = self._load_data_from_selection(config_data.get('messages'), 'messages')
            attachment_files = self._load_data_from_selection(config_data.get('attachments'), 'attachments'); proxies = self._load_data_from_selection(config_data.get('proxies'), 'proxies')
        except Exception as e: QMessageBox.critical(self, "Data Loading Error", f"Failed to load data lists for campaign:\n{e}"); return []
        if not leads: QMessageBox.warning(self, "Missing Data", "Lead list is empty or failed to load."); return [];
        if not smtps: QMessageBox.warning(self, "Missing Data", "SMTP list is empty or failed to load."); return [];
        if not subjects: QMessageBox.warning(self, "Missing Data", "Subject list is empty or failed to load."); return [];
        if not message_files: QMessageBox.warning(self, "Missing Data", "Message list (folder) is empty or failed to load."); return []
        messages = [];
        for msg_path in message_files:
             try:
                 with open(msg_path, 'r', encoding='utf-8') as f: messages.append(f.read())
             except Exception as e: print(f"W: Could not read message file {msg_path}: {e}")
        if not messages: QMessageBox.warning(self, "Missing Data", "Could not read any message files."); return []
        count = len(leads); mode = config_data.get("sending_mode", "No Delay"); schedule = []
        try:
            if mode == "Custom Delay": schedule = generate_schedule_custom_delay(count, config_data.get("delay_min",0), config_data.get("delay_max",5))
            elif mode == "Batch Mode": schedule = generate_schedule_batch(count, config_data.get("batch_min",10), config_data.get("batch_max",20), config_data.get("batch_delay_min",60), config_data.get("batch_delay_max",120))
            elif mode == "Spike Mode":
                 day_counts = config_data.get("spike_days", []);
                 if not day_counts or sum(day_counts) == 0: QMessageBox.warning(self, "Spike Mode Error", "No day counts provided or all counts are zero."); return []
                 if sum(day_counts) > count: QMessageBox.warning(self, "Spike Mode Error", f"Total spike count ({sum(day_counts)}) exceeds lead count ({count})."); return []
                 schedule = generate_schedule_spike(day_counts); count = len(schedule)
                 if count > len(leads): print("W: Spike schedule length exceeds leads. Trimming."); schedule = schedule[:len(leads)]; count = len(leads)
            else: schedule = generate_schedule_no_delay(count)
        except Exception as e: QMessageBox.critical(self, "Scheduling Error", f"Failed to generate schedule for mode '{mode}':\n{e}"); return []
        if not schedule: QMessageBox.critical(self, "Scheduling Error", "Generated schedule is empty."); return []
        tasks = [];
        for i in range(count):
            if i >= len(leads) or i >= len(schedule): break
            lead_data = leads[i];
            if not isinstance(lead_data, dict) or 'email' not in lead_data: print(f"W: Skipping invalid lead data at index {i}: {lead_data}"); continue
            try:
                smtp_config = random.choice(smtps); subject_template = random.choice(subjects); message_template = random.choice(messages)
                proxy_config = random.choice(proxies) if proxies else None
                selected_attachments = random.sample(attachment_files, k=min(len(attachment_files), 1)) if attachment_files else []
            except IndexError: QMessageBox.critical(self, "Data Error", "Could not randomly select from required lists."); return []
            except Exception as e: QMessageBox.critical(self, "Task Creation Error", f"Error during random selection: {e}"); return []
            try: final_subject = replace_placeholders(subject_template, lead_data); final_body = replace_placeholders(message_template, lead_data)
            except Exception as e: print(f"W: Error replacing placeholders for {lead_data.get('email')}: {e}"); final_subject = subject_template; final_body = message_template
            msg_args = {'to': lead_data['email'], 'subject': final_subject, 'body': final_body, 'attachments': selected_attachments };
            task_args = {'smtp': smtp_config, 'msg': msg_args, 'proxy': proxy_config };
            tasks.append({'send_time': schedule[i], 'args': task_args});
        print(f"Collected {len(tasks)} tasks successfully."); return tasks

    def _on_preview(self):
        QMessageBox.information(self, "Preview", "Campaign preview feature is not yet implemented.")

    def _on_send(self):
        if self.worker and self.worker.isRunning(): QMessageBox.warning(self, "Already Running", "A campaign is already in progress."); return
        if not self.current_campaign_name: QMessageBox.warning(self, "No Campaign Selected", "Please select a campaign to launch."); return
        if QMessageBox.question(self, "Confirm Launch", f"Are you sure you want to launch the campaign '{self.current_campaign_name}'?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No) == QMessageBox.StandardButton.No:
             print("Campaign launch cancelled by user."); return
        tasks = self._collect_tasks();
        if not tasks: print("No tasks generated. Campaign not started."); self.progress.setVisible(False); return
        total = len(tasks); self.progress.setRange(0, total); self.progress.setValue(0); self.progress.setFormat(f"Sending... %v/%m (%p%)"); self.progress.setVisible(True);
        self.log_view.clear(); self.log_view.append(f"🚀 Starting campaign '{self.current_campaign_name}' ({total} emails)...")
        QApplication.processEvents()
        self.worker = SendWorker(tasks);
        self.worker.progress.connect(lambda sent, tot: self.progress.setValue(sent))
        self.worker.log.connect(lambda msg: self.log_view.append(msg))
        self.worker.finished.connect(self._on_campaign_finished)
        self.worker.start()
        print("SendWorker thread started.")

    def _on_campaign_finished(self):
        print(f"Campaign '{self.current_campaign_name}' finished processing.")
        if self.progress.value() == self.progress.maximum(): self.progress.setFormat("✅ Completed %m/%m (100%)")
        else: self.progress.setFormat(f"Finished %v/%m (%p%)")
        QMessageBox.information(self, "Campaign Finished", f"Campaign '{self.current_campaign_name}' has finished processing.")
        self.worker = None

    def closeEvent(self, event):
         if self.worker and self.worker.isRunning():
             reply = QMessageBox.question(self, 'Confirm Close', 'A campaign is running. Stop it and close?', QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
             if reply == QMessageBox.StandardButton.Yes: print("Terminating SendWorker..."); self.worker.terminate(); self.worker.wait(); event.accept()
             else: event.ignore()
         else: event.accept()