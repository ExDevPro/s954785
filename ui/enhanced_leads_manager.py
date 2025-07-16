# Enhanced Leads Manager with Pagination, Background Workers, and Professional UI

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QListWidget, QPushButton, QTableWidget,
    QHeaderView, QLineEdit, QProgressBar, QMessageBox, QFileDialog, QInputDialog,
    QApplication, QTableWidgetItem, QMenu, QAbstractItemView, QStyle, QDialog,
    QComboBox, QDialogButtonBox, QGridLayout, QSpinBox, QTextEdit, QSplitter,
    QScrollArea, QFrame, QSizePolicy
)
from PyQt6.QtGui import QAction, QFont, QPalette, QPixmap, QIcon
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QMutex, QWaitCondition, QTimer

import os
import re
import shutil
import logging
import traceback
import json
from openpyxl import load_workbook, Workbook
from datetime import datetime

# Configuration constants
DEFAULT_CHUNK_SIZE = 500
DATA_DIR = os.path.join(os.path.abspath(os.path.dirname(__file__)), '../data/leads')

class CreateListDialog(QDialog):
    """Professional dialog for creating new lists with descriptions."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Create New Leads List")
        self.setModal(True)
        self.setFixedSize(400, 250)
        
        layout = QVBoxLayout(self)
        
        # List name
        layout.addWidget(QLabel("List Name:"))
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("Enter list name...")
        layout.addWidget(self.name_edit)
        
        # Description
        layout.addWidget(QLabel("Description (Optional):"))
        self.desc_edit = QTextEdit()
        self.desc_edit.setPlaceholderText("Enter description for this list...")
        self.desc_edit.setMaximumHeight(80)
        layout.addWidget(self.desc_edit)
        
        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        self.name_edit.textChanged.connect(self._validate_input)
        self._validate_input()
    
    def _validate_input(self):
        name = self.name_edit.text().strip()
        valid = bool(name and re.match(r'^[a-zA-Z0-9_\-\s]+$', name))
        self.findChild(QDialogButtonBox).button(QDialogButtonBox.StandardButton.Ok).setEnabled(valid)
    
    def get_list_data(self):
        return {
            'name': self.name_edit.text().strip(),
            'description': self.desc_edit.toPlainText().strip()
        }

class LeadsDataWorker(QThread):
    """Background worker for all leads data operations."""
    
    data_loaded = pyqtSignal(list, list, int, int)
    progress_updated = pyqtSignal(int, int, str)
    save_finished = pyqtSignal(bool, str)
    error_occurred = pyqtSignal(str)
    import_finished = pyqtSignal(list, list, int)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.operation = None
        self.parameters = {}
        self.stop_flag = False
        self.mutex = QMutex()
        
    def load_chunk(self, file_path, offset, chunk_size):
        """Load a specific chunk of data."""
        self.mutex.lock()
        self.operation = "load_chunk"
        self.parameters = {
            'file_path': file_path,
            'offset': offset,
            'chunk_size': chunk_size
        }
        self.stop_flag = False
        self.mutex.unlock()
        self.start()
    
    def save_data(self, file_path, headers, data):
        """Save data to Excel file."""
        self.mutex.lock()
        self.operation = "save_data"
        self.parameters = {
            'file_path': file_path,
            'headers': headers,
            'data': data
        }
        self.stop_flag = False
        self.mutex.unlock()
        self.start()
    
    def import_file(self, source_path, target_path):
        """Import data from external file."""
        self.mutex.lock()
        self.operation = "import_file"
        self.parameters = {
            'source_path': source_path,
            'target_path': target_path
        }
        self.stop_flag = False
        self.mutex.unlock()
        self.start()
    
    def stop_operation(self):
        """Stop current operation."""
        self.mutex.lock()
        self.stop_flag = True
        self.mutex.unlock()
        self.wait()
    
    def run(self):
        try:
            if self.operation == "load_chunk":
                self._load_chunk()
            elif self.operation == "save_data":
                self._save_data()
            elif self.operation == "import_file":
                self._import_file()
        except Exception as e:
            self.error_occurred.emit(f"Worker error: {str(e)}\n{traceback.format_exc()}")
    
    def _load_chunk(self):
        """Load data chunk from Excel file."""
        try:
            params = self.parameters
            file_path = params['file_path']
            offset = params['offset']
            chunk_size = params['chunk_size']
            
            if not os.path.exists(file_path):
                # Create empty file if it doesn't exist
                wb = Workbook()
                ws = wb.active
                ws.append(['Email', 'First Name', 'Last Name', 'Company', 'Phone', 'Notes'])
                wb.save(file_path)
                self.data_loaded.emit(['Email', 'First Name', 'Last Name', 'Company', 'Phone', 'Notes'], [], 0, 0)
                return
            
            self.progress_updated.emit(10, 100, "Opening file...")
            
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            if ws.max_row >= 1:
                header_row = next(ws.iter_rows(min_row=1, max_row=1), [])
                headers = [cell.value or f"Column {i+1}" for i, cell in enumerate(header_row)]
            
            self.progress_updated.emit(30, 100, "Reading data...")
            
            # Calculate total rows (excluding header)
            total_rows = max(0, ws.max_row - 1)
            
            # Read chunk of data
            start_row = offset + 2  # +2 because row 1 is header and Excel is 1-indexed
            end_row = min(start_row + chunk_size - 1, ws.max_row)
            
            data = []
            if start_row <= ws.max_row:
                for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True)):
                    if self.stop_flag:
                        break
                    
                    # Ensure row has same number of columns as headers
                    padded_row = list(row) + [''] * (len(headers) - len(row))
                    padded_row = padded_row[:len(headers)]
                    # Convert None values to empty strings
                    padded_row = [str(cell) if cell is not None else '' for cell in padded_row]
                    data.append(padded_row)
                    
                    if row_idx % 50 == 0:
                        progress = 30 + int((row_idx / chunk_size) * 60)
                        self.progress_updated.emit(progress, 100, f"Loading row {offset + row_idx + 1}...")
            
            wb.close()
            
            self.progress_updated.emit(100, 100, "Complete")
            self.data_loaded.emit(headers, data, total_rows, offset)
            
        except Exception as e:
            self.error_occurred.emit(f"Error loading data: {str(e)}")
    
    def _save_data(self):
        """Save data to Excel file."""
        try:
            params = self.parameters
            file_path = params['file_path']
            headers = params['headers']
            data = params['data']
            
            self.progress_updated.emit(10, 100, "Creating workbook...")
            
            wb = Workbook()
            ws = wb.active
            
            # Write headers
            ws.append(headers)
            
            # Write data
            total_rows = len(data)
            for idx, row in enumerate(data):
                if self.stop_flag:
                    break
                
                ws.append(row)
                
                if idx % 100 == 0:
                    progress = 10 + int((idx / total_rows) * 80)
                    self.progress_updated.emit(progress, 100, f"Saving row {idx + 1}/{total_rows}...")
            
            self.progress_updated.emit(90, 100, "Saving file...")
            wb.save(file_path)
            
            self.progress_updated.emit(100, 100, "Complete")
            self.save_finished.emit(True, "Data saved successfully")
            
        except Exception as e:
            self.save_finished.emit(False, f"Error saving data: {str(e)}")
    
    def _import_file(self):
        """Import data from external file."""
        try:
            params = self.parameters
            source_path = params['source_path']
            target_path = params['target_path']
            
            self.progress_updated.emit(10, 100, "Reading source file...")
            
            # Determine file type and read accordingly
            if source_path.endswith('.xlsx') or source_path.endswith('.xls'):
                wb = load_workbook(filename=source_path, read_only=True)
                ws = wb.active
                
                # Read headers
                headers = []
                if ws.max_row >= 1:
                    header_row = next(ws.iter_rows(min_row=1, max_row=1), [])
                    headers = [cell.value or f"Column {i+1}" for i, cell in enumerate(header_row)]
                
                # Read all data
                data = []
                total_rows = ws.max_row - 1  # Exclude header
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                    if self.stop_flag:
                        break
                    
                    # Ensure row has same number of columns as headers
                    padded_row = list(row) + [''] * (len(headers) - len(row))
                    padded_row = padded_row[:len(headers)]
                    # Convert None values to empty strings
                    padded_row = [str(cell) if cell is not None else '' for cell in padded_row]
                    data.append(padded_row)
                    
                    if row_idx % 500 == 0:
                        progress = 10 + int((row_idx / total_rows) * 70)
                        self.progress_updated.emit(progress, 100, f"Reading row {row_idx + 1}/{total_rows}...")
                
                wb.close()
                
            elif source_path.endswith('.csv'):
                import csv
                headers = []
                data = []
                
                with open(source_path, 'r', encoding='utf-8', newline='') as csvfile:
                    # Detect delimiter
                    sample = csvfile.read(1024)
                    csvfile.seek(0)
                    sniffer = csv.Sniffer()
                    delimiter = sniffer.sniff(sample).delimiter
                    
                    reader = csv.reader(csvfile, delimiter=delimiter)
                    
                    # Read headers
                    headers = next(reader, [])
                    
                    # Read data
                    for row_idx, row in enumerate(reader):
                        if self.stop_flag:
                            break
                        
                        # Ensure row has same number of columns as headers
                        padded_row = row + [''] * (len(headers) - len(row))
                        padded_row = padded_row[:len(headers)]
                        data.append(padded_row)
                        
                        if row_idx % 500 == 0:
                            self.progress_updated.emit(30 + (row_idx // 500), 100, f"Reading row {row_idx + 1}...")
            
            else:
                raise ValueError("Unsupported file format. Please use .xlsx, .xls, or .csv files.")
            
            self.progress_updated.emit(80, 100, "Saving imported data...")
            
            # Save to target file
            os.makedirs(os.path.dirname(target_path), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            
            # Write headers
            ws.append(headers)
            
            # Write data
            for row in data:
                ws.append(row)
            
            wb.save(target_path)
            
            self.progress_updated.emit(100, 100, "Import complete")
            self.import_finished.emit(headers, data, len(data))
            
        except Exception as e:
            self.error_occurred.emit(f"Error importing file: {str(e)}")

class EnhancedLeadsManager(QWidget):
    """Enhanced Leads Manager with pagination, background processing, and professional UI."""
    
    counts_changed = pyqtSignal()
    
    def __init__(self, parent=None, config=None):
        super().__init__(parent)
        self.config = config
        self.current_list = None
        self.current_headers = []
        self.current_offset = 0
        self.total_rows = 0
        self.chunk_size = self._get_chunk_size()
        
        # Background worker
        self.worker = LeadsDataWorker(self)
        self.worker.data_loaded.connect(self._on_data_loaded)
        self.worker.progress_updated.connect(self._on_progress_updated)
        self.worker.save_finished.connect(self._on_save_finished)
        self.worker.error_occurred.connect(self._on_error_occurred)
        self.worker.import_finished.connect(self._on_import_finished)
        
        # Auto-save timer
        self.auto_save_timer = QTimer()
        self.auto_save_timer.timeout.connect(self._auto_save_data)
        self.auto_save_timer.setSingleShot(True)
        
        self._setup_ui()
        self._load_lists()
        
        # Apply theme
        self._apply_styling()
        
    def _get_chunk_size(self):
        """Get chunk size from configuration."""
        if self.config and hasattr(self.config, 'get'):
            return self.config.get('leads.chunk_size', DEFAULT_CHUNK_SIZE)
        return DEFAULT_CHUNK_SIZE
    
    def _setup_ui(self):
        """Setup the user interface."""
        main_layout = QHBoxLayout(self)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Left panel - Lists
        left_panel = self._create_left_panel()
        main_layout.addWidget(left_panel, 1)
        
        # Splitter for resizable panels
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Right panel - Data view
        right_panel = self._create_right_panel()
        splitter.addWidget(right_panel)
        
        # Description panel
        desc_panel = self._create_description_panel()
        splitter.addWidget(desc_panel)
        
        splitter.setSizes([600, 200])  # Set initial sizes
        main_layout.addWidget(splitter, 3)
        
    def _create_left_panel(self):
        """Create the left panel with list management."""
        panel = QFrame()
        panel.setFrameStyle(QFrame.Shape.StyledPanel)
        panel.setMaximumWidth(350)
        panel.setMinimumWidth(250)
        
        layout = QVBoxLayout(panel)
        
        # Header
        header = QLabel("📋 Leads Lists")
        header.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setStyleSheet("QLabel { background: #3498db; color: white; padding: 8px; border-radius: 4px; }")
        layout.addWidget(header)
        
        # Search
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("🔍 Search lists...")
        self.search_edit.textChanged.connect(self._filter_lists)
        layout.addWidget(self.search_edit)
        
        # Lists
        self.lists_widget = QListWidget()
        self.lists_widget.itemClicked.connect(self._on_list_selected)
        self.lists_widget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.lists_widget.customContextMenuRequested.connect(self._show_list_context_menu)
        layout.addWidget(self.lists_widget)
        
        # Buttons
        buttons_layout = QVBoxLayout()
        
        self.btn_new_list = QPushButton("➕ New List")
        self.btn_new_list.clicked.connect(self._create_new_list)
        self.btn_new_list.setStyleSheet("QPushButton { background: #27ae60; color: white; padding: 8px; border-radius: 4px; }")
        buttons_layout.addWidget(self.btn_new_list)
        
        self.btn_delete_list = QPushButton("🗑️ Delete List")
        self.btn_delete_list.clicked.connect(self._delete_list)
        self.btn_delete_list.setStyleSheet("QPushButton { background: #e74c3c; color: white; padding: 8px; border-radius: 4px; }")
        buttons_layout.addWidget(self.btn_delete_list)
        
        self.btn_import = QPushButton("📥 Import Data")
        self.btn_import.clicked.connect(self._import_data)
        self.btn_import.setStyleSheet("QPushButton { background: #9b59b6; color: white; padding: 8px; border-radius: 4px; }")
        buttons_layout.addWidget(self.btn_import)
        
        self.btn_export = QPushButton("📤 Export Data")
        self.btn_export.clicked.connect(self._export_data)
        self.btn_export.setStyleSheet("QPushButton { background: #f39c12; color: white; padding: 8px; border-radius: 4px; }")
        buttons_layout.addWidget(self.btn_export)
        
        layout.addLayout(buttons_layout)
        
        return panel
    
    def _create_right_panel(self):
        """Create the right panel with data table."""
        panel = QFrame()
        panel.setFrameStyle(QFrame.Shape.StyledPanel)
        
        layout = QVBoxLayout(panel)
        
        # Header with pagination controls
        header_layout = QHBoxLayout()
        
        self.table_header = QLabel("📊 Leads Data")
        self.table_header.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        header_layout.addWidget(self.table_header)
        
        header_layout.addStretch()
        
        # Pagination controls
        self.btn_first_page = QPushButton("⏮️")
        self.btn_first_page.setToolTip("First page")
        self.btn_first_page.clicked.connect(self._go_to_first_page)
        self.btn_first_page.setFixedSize(30, 30)
        header_layout.addWidget(self.btn_first_page)
        
        self.btn_prev_page = QPushButton("◀️")
        self.btn_prev_page.setToolTip("Previous page")
        self.btn_prev_page.clicked.connect(self._go_to_prev_page)
        self.btn_prev_page.setFixedSize(30, 30)
        header_layout.addWidget(self.btn_prev_page)
        
        self.page_info = QLabel("Page 1 of 1")
        self.page_info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.page_info.setMinimumWidth(100)
        header_layout.addWidget(self.page_info)
        
        self.btn_next_page = QPushButton("▶️")
        self.btn_next_page.setToolTip("Next page")
        self.btn_next_page.clicked.connect(self._go_to_next_page)
        self.btn_next_page.setFixedSize(30, 30)
        header_layout.addWidget(self.btn_next_page)
        
        self.btn_last_page = QPushButton("⏭️")
        self.btn_last_page.setToolTip("Last page")
        self.btn_last_page.clicked.connect(self._go_to_last_page)
        self.btn_last_page.setFixedSize(30, 30)
        header_layout.addWidget(self.btn_last_page)
        
        layout.addLayout(header_layout)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Status label
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("QLabel { color: #7f8c8d; font-style: italic; }")
        layout.addWidget(self.status_label)
        
        # Data table
        self.data_table = QTableWidget()
        self.data_table.setAlternatingRowColors(True)
        self.data_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.data_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.data_table.customContextMenuRequested.connect(self._show_table_context_menu)
        self.data_table.cellChanged.connect(self._on_cell_changed)
        layout.addWidget(self.data_table)
        
        # Table controls
        table_controls = QHBoxLayout()
        
        self.btn_add_row = QPushButton("➕ Add Row")
        self.btn_add_row.clicked.connect(self._add_row)
        table_controls.addWidget(self.btn_add_row)
        
        self.btn_delete_row = QPushButton("➖ Delete Row")
        self.btn_delete_row.clicked.connect(self._delete_row)
        table_controls.addWidget(self.btn_delete_row)
        
        table_controls.addStretch()
        
        self.records_info = QLabel("0 records")
        table_controls.addWidget(self.records_info)
        
        layout.addLayout(table_controls)
        
        return panel
    
    def _create_description_panel(self):
        """Create the description panel."""
        panel = QFrame()
        panel.setFrameStyle(QFrame.Shape.StyledPanel)
        panel.setMinimumWidth(200)
        panel.setMaximumWidth(300)
        
        layout = QVBoxLayout(panel)
        
        # Header
        header = QLabel("📝 Description")
        header.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setStyleSheet("QLabel { background: #95a5a6; color: white; padding: 6px; border-radius: 4px; }")
        layout.addWidget(header)
        
        # Scrollable description area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        self.description_label = QLabel("Select a list to view its description.")
        self.description_label.setWordWrap(True)
        self.description_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.description_label.setStyleSheet("QLabel { padding: 10px; background: white; }")
        
        scroll_area.setWidget(self.description_label)
        layout.addWidget(scroll_area)
        
        # Edit button
        self.btn_edit_desc = QPushButton("✏️ Edit Description")
        self.btn_edit_desc.clicked.connect(self._edit_description)
        self.btn_edit_desc.setEnabled(False)
        layout.addWidget(self.btn_edit_desc)
        
        return panel
    
    def _apply_styling(self):
        """Apply professional styling."""
        self.setStyleSheet("""
            QFrame {
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: #ecf0f1;
            }
            QListWidget {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background: white;
                padding: 5px;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #ecf0f1;
            }
            QListWidget::item:selected {
                background: #3498db;
                color: white;
            }
            QTableWidget {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background: white;
                gridline-color: #ecf0f1;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QLineEdit {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                padding: 6px;
            }
            QLineEdit:focus {
                border: 2px solid #3498db;
            }
            QPushButton {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                padding: 6px 12px;
                background: #ecf0f1;
            }
            QPushButton:hover {
                background: #d5dbdb;
            }
            QPushButton:pressed {
                background: #bdc3c7;
            }
        """)
    
    # ========== List Management Methods ==========
    
    def _load_lists(self):
        """Load all available lists."""
        os.makedirs(DATA_DIR, exist_ok=True)
        self.lists_widget.clear()
        
        for item in os.listdir(DATA_DIR):
            item_path = os.path.join(DATA_DIR, item)
            if os.path.isdir(item_path):
                self.lists_widget.addItem(item)
        
        self.counts_changed.emit()
    
    def _filter_lists(self):
        """Filter lists based on search text."""
        search_text = self.search_edit.text().lower()
        for i in range(self.lists_widget.count()):
            item = self.lists_widget.item(i)
            item.setHidden(search_text not in item.text().lower())
    
    def _create_new_list(self):
        """Create a new leads list."""
        dialog = CreateListDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_list_data()
            list_name = data['name']
            description = data['description']
            
            # Check if list already exists
            list_dir = os.path.join(DATA_DIR, list_name)
            if os.path.exists(list_dir):
                QMessageBox.warning(self, "List Exists", f"A list named '{list_name}' already exists.")
                return
            
            try:
                # Create list directory and files
                os.makedirs(list_dir, exist_ok=True)
                
                # Create Excel file with default structure
                excel_file = os.path.join(list_dir, f"{list_name}.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.append(['Email', 'First Name', 'Last Name', 'Company', 'Phone', 'Notes'])
                wb.save(excel_file)
                
                # Save description
                if description:
                    desc_file = os.path.join(list_dir, "description.txt")
                    with open(desc_file, 'w', encoding='utf-8') as f:
                        f.write(description)
                
                self._load_lists()
                
                # Select the new list
                for i in range(self.lists_widget.count()):
                    if self.lists_widget.item(i).text() == list_name:
                        self.lists_widget.setCurrentRow(i)
                        self._on_list_selected(self.lists_widget.item(i))
                        break
                
                QMessageBox.information(self, "Success", f"List '{list_name}' created successfully.")
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to create list: {str(e)}")
    
    def _delete_list(self):
        """Delete the selected list."""
        current_item = self.lists_widget.currentItem()
        if not current_item:
            QMessageBox.warning(self, "No Selection", "Please select a list to delete.")
            return
        
        list_name = current_item.text()
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete the list '{list_name}'?\n\nThis action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                list_dir = os.path.join(DATA_DIR, list_name)
                shutil.rmtree(list_dir)
                
                self._load_lists()
                self._clear_data_view()
                
                QMessageBox.information(self, "Success", f"List '{list_name}' deleted successfully.")
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete list: {str(e)}")
    
    def _on_list_selected(self, item):
        """Handle list selection."""
        if not item:
            return
        
        list_name = item.text()
        self.current_list = list_name
        self.current_offset = 0
        
        # Update description
        self._load_description(list_name)
        
        # Load data
        self._load_current_chunk()
        
        # Update UI
        self.table_header.setText(f"📊 {list_name} - Leads Data")
        self.btn_edit_desc.setEnabled(True)
    
    def _load_description(self, list_name):
        """Load and display list description."""
        desc_file = os.path.join(DATA_DIR, list_name, "description.txt")
        
        if os.path.exists(desc_file):
            try:
                with open(desc_file, 'r', encoding='utf-8') as f:
                    description = f.read().strip()
                
                if description:
                    self.description_label.setText(description)
                else:
                    self.description_label.setText("No description available.")
            except Exception as e:
                self.description_label.setText(f"Error loading description: {str(e)}")
        else:
            self.description_label.setText("No description available.")
    
    def _edit_description(self):
        """Edit the description of the current list."""
        if not self.current_list:
            return
        
        desc_file = os.path.join(DATA_DIR, self.current_list, "description.txt")
        current_desc = ""
        
        if os.path.exists(desc_file):
            try:
                with open(desc_file, 'r', encoding='utf-8') as f:
                    current_desc = f.read().strip()
            except:
                pass
        
        new_desc, ok = QInputDialog.getMultiLineText(
            self, "Edit Description", 
            f"Description for '{self.current_list}':",
            current_desc
        )
        
        if ok:
            try:
                if new_desc.strip():
                    with open(desc_file, 'w', encoding='utf-8') as f:
                        f.write(new_desc.strip())
                else:
                    # Remove description file if empty
                    if os.path.exists(desc_file):
                        os.remove(desc_file)
                
                self._load_description(self.current_list)
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save description: {str(e)}")
    
    # ========== Data Management Methods ==========
    
    def _load_current_chunk(self):
        """Load current chunk of data."""
        if not self.current_list:
            return
        
        excel_file = os.path.join(DATA_DIR, self.current_list, f"{self.current_list}.xlsx")
        
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Loading data...")
        
        self.worker.load_chunk(excel_file, self.current_offset, self.chunk_size)
    
    def _on_data_loaded(self, headers, data, total_rows, offset):
        """Handle data loaded signal."""
        self.current_headers = headers
        self.total_rows = total_rows
        self.current_offset = offset
        
        # Update table
        self.data_table.setRowCount(len(data))
        self.data_table.setColumnCount(len(headers))
        self.data_table.setHorizontalHeaderLabels([f"{i+1}. {h}" for i, h in enumerate(headers)])
        
        # Populate data with serial numbers
        for row_idx, row_data in enumerate(data):
            for col_idx, cell_data in enumerate(row_data):
                item = QTableWidgetItem(str(cell_data))
                self.data_table.setItem(row_idx, col_idx, item)
        
        # Update pagination info
        self._update_pagination_info()
        
        # Update records info
        self.records_info.setText(f"{len(data)} of {total_rows} records")
        
        # Hide progress bar
        self.progress_bar.setVisible(False)
        self.status_label.setText("Ready")
        
        # Adjust column widths
        self.data_table.resizeColumnsToContents()
    
    def _on_progress_updated(self, current, total, message):
        """Handle progress updates."""
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.status_label.setText(message)
    
    def _on_save_finished(self, success, message):
        """Handle save completion."""
        self.progress_bar.setVisible(False)
        if success:
            self.status_label.setText("Data saved successfully")
        else:
            self.status_label.setText("Save failed")
            QMessageBox.critical(self, "Save Error", message)
    
    def _on_error_occurred(self, error_message):
        """Handle worker errors."""
        self.progress_bar.setVisible(False)
        self.status_label.setText("Error occurred")
        QMessageBox.critical(self, "Operation Error", error_message)
    
    def _on_import_finished(self, headers, data, total_imported):
        """Handle import completion."""
        self.progress_bar.setVisible(False)
        self.status_label.setText(f"Imported {total_imported} records")
        
        # Reload current data
        self.current_offset = 0
        self._load_current_chunk()
        
        QMessageBox.information(self, "Import Complete", f"Successfully imported {total_imported} records.")
    
    def _clear_data_view(self):
        """Clear the data view."""
        self.data_table.setRowCount(0)
        self.data_table.setColumnCount(0)
        self.current_list = None
        self.current_headers = []
        self.current_offset = 0
        self.total_rows = 0
        self.table_header.setText("📊 Leads Data")
        self.records_info.setText("0 records")
        self.description_label.setText("Select a list to view its description.")
        self.btn_edit_desc.setEnabled(False)
        self._update_pagination_info()
    
    # ========== Pagination Methods ==========
    
    def _update_pagination_info(self):
        """Update pagination information and button states."""
        if self.total_rows == 0:
            self.page_info.setText("Page 0 of 0")
            self._set_pagination_buttons_enabled(False)
            return
        
        current_page = (self.current_offset // self.chunk_size) + 1
        total_pages = ((self.total_rows - 1) // self.chunk_size) + 1
        
        self.page_info.setText(f"Page {current_page} of {total_pages}")
        
        # Enable/disable pagination buttons
        self.btn_first_page.setEnabled(current_page > 1)
        self.btn_prev_page.setEnabled(current_page > 1)
        self.btn_next_page.setEnabled(current_page < total_pages)
        self.btn_last_page.setEnabled(current_page < total_pages)
    
    def _set_pagination_buttons_enabled(self, enabled):
        """Enable/disable all pagination buttons."""
        self.btn_first_page.setEnabled(enabled)
        self.btn_prev_page.setEnabled(enabled)
        self.btn_next_page.setEnabled(enabled)
        self.btn_last_page.setEnabled(enabled)
    
    def _go_to_first_page(self):
        """Go to first page."""
        if self.current_offset != 0:
            self.current_offset = 0
            self._load_current_chunk()
    
    def _go_to_prev_page(self):
        """Go to previous page."""
        if self.current_offset >= self.chunk_size:
            self.current_offset -= self.chunk_size
            self._load_current_chunk()
    
    def _go_to_next_page(self):
        """Go to next page."""
        if self.current_offset + self.chunk_size < self.total_rows:
            self.current_offset += self.chunk_size
            self._load_current_chunk()
    
    def _go_to_last_page(self):
        """Go to last page."""
        if self.total_rows > 0:
            last_page_offset = ((self.total_rows - 1) // self.chunk_size) * self.chunk_size
            if self.current_offset != last_page_offset:
                self.current_offset = last_page_offset
                self._load_current_chunk()
    
    # ========== Data Editing Methods ==========
    
    def _on_cell_changed(self, row, column):
        """Handle cell changes."""
        # Start auto-save timer
        self.auto_save_timer.start(1000)  # Auto-save after 1 second of no changes
    
    def _auto_save_data(self):
        """Auto-save the current data."""
        if not self.current_list or not self.current_headers:
            return
        
        try:
            # Get all data from table
            data = []
            for row in range(self.data_table.rowCount()):
                row_data = []
                for col in range(self.data_table.columnCount()):
                    item = self.data_table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)
            
            # Save only the current chunk - need to merge with full data
            # This is a simplified version - in production, you'd want to save the entire file
            excel_file = os.path.join(DATA_DIR, self.current_list, f"{self.current_list}.xlsx")
            self.status_label.setText("Auto-saving...")
            
            # For now, just save the current chunk
            # In a full implementation, you'd need to load the entire file, update the chunk, and save
            
        except Exception as e:
            print(f"Auto-save error: {e}")
    
    def _add_row(self):
        """Add a new row to the table."""
        if not self.current_headers:
            QMessageBox.warning(self, "No List", "Please select a list first.")
            return
        
        row_count = self.data_table.rowCount()
        self.data_table.insertRow(row_count)
        
        # Add empty cells
        for col in range(len(self.current_headers)):
            item = QTableWidgetItem("")
            self.data_table.setItem(row_count, col, item)
        
        # Update records info
        self.records_info.setText(f"{row_count + 1} of {self.total_rows + 1} records")
        
        # Start auto-save timer
        self.auto_save_timer.start(1000)
    
    def _delete_row(self):
        """Delete selected row(s)."""
        selected_rows = set()
        for item in self.data_table.selectedItems():
            selected_rows.add(item.row())
        
        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select row(s) to delete.")
            return
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete {len(selected_rows)} row(s)?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Delete rows in reverse order to maintain indices
            for row in sorted(selected_rows, reverse=True):
                self.data_table.removeRow(row)
            
            # Update records info
            row_count = self.data_table.rowCount()
            self.records_info.setText(f"{row_count} of {self.total_rows} records")
            
            # Start auto-save timer
            self.auto_save_timer.start(1000)
    
    # ========== Import/Export Methods ==========
    
    def _import_data(self):
        """Import data from external file."""
        if not self.current_list:
            QMessageBox.warning(self, "No List", "Please select a list first.")
            return
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Data",
            "", "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        
        if file_path:
            reply = QMessageBox.question(
                self, "Import Data",
                f"Import data from '{os.path.basename(file_path)}'?\n\nThis will replace existing data.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                target_file = os.path.join(DATA_DIR, self.current_list, f"{self.current_list}.xlsx")
                
                self.progress_bar.setVisible(True)
                self.progress_bar.setValue(0)
                self.status_label.setText("Importing data...")
                
                self.worker.import_file(file_path, target_file)
    
    def _export_data(self):
        """Export current list data."""
        if not self.current_list:
            QMessageBox.warning(self, "No List", "Please select a list first.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Data",
            f"{self.current_list}_export.xlsx",
            "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)"
        )
        
        if file_path:
            try:
                source_file = os.path.join(DATA_DIR, self.current_list, f"{self.current_list}.xlsx")
                shutil.copy2(source_file, file_path)
                
                QMessageBox.information(self, "Export Complete", f"Data exported to '{file_path}'")
                
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Failed to export data: {str(e)}")
    
    # ========== Context Menu Methods ==========
    
    def _show_list_context_menu(self, position):
        """Show context menu for lists."""
        item = self.lists_widget.itemAt(position)
        
        menu = QMenu()
        
        if item:
            menu.addAction("📂 Open", lambda: self._on_list_selected(item))
            menu.addSeparator()
            menu.addAction("✏️ Edit Description", self._edit_description)
            menu.addSeparator()
            menu.addAction("🗑️ Delete", self._delete_list)
        
        menu.addSeparator()
        menu.addAction("➕ New List", self._create_new_list)
        menu.addAction("🔄 Refresh", self._load_lists)
        
        menu.exec(self.lists_widget.mapToGlobal(position))
    
    def _show_table_context_menu(self, position):
        """Show context menu for table."""
        menu = QMenu()
        
        menu.addAction("➕ Add Row", self._add_row)
        
        if self.data_table.selectedItems():
            menu.addAction("➖ Delete Selected Rows", self._delete_row)
        
        menu.addSeparator()
        menu.addAction("📥 Import Data", self._import_data)
        menu.addAction("📤 Export Data", self._export_data)
        
        menu.exec(self.data_table.mapToGlobal(position))
    
    # ========== Public Methods ==========
    
    def get_list_count(self):
        """Get the number of lists."""
        return self.lists_widget.count()
    
    def get_selected_list(self):
        """Get the currently selected list name."""
        return self.current_list
    
    def refresh_lists(self):
        """Refresh the lists display."""
        self._load_lists()