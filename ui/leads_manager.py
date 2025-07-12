# leads_manager.py - FINAL FULL VERSION (Part 1 of 2)

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QListWidget, QPushButton, QTableWidget,
    QHeaderView, QLineEdit, QProgressBar, QMessageBox, QFileDialog, QInputDialog,
    QApplication, QTableWidgetItem, QMenu, QAbstractItemView, QStyle, QDialog,
    QComboBox, QDialogButtonBox, QGridLayout
)
from PyQt6.QtGui import QAction
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QMutex, QWaitCondition

import os
import re
import shutil
import logging
import traceback
from openpyxl import load_workbook, Workbook

CHUNK_SIZE = 500
DATA_DIR = os.path.join(os.path.abspath(os.path.dirname(__file__)), '../data/leads')

class LeadsDataThread(QThread):
    data_loaded = pyqtSignal(list, list, int, int)
    load_progress = pyqtSignal(int, int)
    save_finished = pyqtSignal(bool, str)
    error_occurred = pyqtSignal(str)
    import_loaded = pyqtSignal(list, list)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.path = None
        self.operation = None
        self.data_to_save = None
        self.offset = 0
        self.stop_flag = False
        self.mutex = QMutex()
        self.wait_condition = QWaitCondition()

    def load_data(self, path: str, offset: int = 0):
        self.mutex.lock()
        self.path = path
        self.operation = "load"
        self.offset = offset
        self.stop_flag = False
        self.mutex.unlock()
        self.start()

    def save_data(self, path: str, data: list):
        self.mutex.lock()
        self.path = path
        self.operation = "save"
        self.data_to_save = data
        self.stop_flag = False
        self.mutex.unlock()
        self.start()

    def import_excel(self, path: str):
        self.mutex.lock()
        self.path = path
        self.operation = "import"
        self.stop_flag = False
        self.mutex.unlock()
        self.start()

    def stop(self):
        self.mutex.lock()
        self.stop_flag = True
        self.wait_condition.wakeAll()
        self.mutex.unlock()
        self.wait()

    def run(self):
        try:
            if self.operation == "load":
                self._load_from_excel()
            elif self.operation == "save":
                self._save_to_excel()
            elif self.operation == "import":
                self._import_excel()
        except Exception:
            self.error_occurred.emit(traceback.format_exc())

    def _load_from_excel(self):
        try:
            wb = load_workbook(filename=self.path, read_only=True)
            ws = wb.active

            if ws.max_row < 1:
                headers = []
                data = []
                total_rows = 0
            else:
                headers = [cell.value or "" for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
                data = [[cell if cell else "" for cell in row] for row in ws.iter_rows(min_row=2, values_only=True)]
                total_rows = len(data)

            wb.close()
            self.data_loaded.emit(headers, data, total_rows, 0)

        except Exception:
            self.error_occurred.emit(traceback.format_exc())

    def _save_to_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(self.data_to_save[0])
            for row in self.data_to_save[1:]:
                ws.append(row)
            wb.save(self.path)
            self.save_finished.emit(True, "Leads saved.")
        except Exception:
            self.error_occurred.emit(traceback.format_exc())

    def _import_excel(self):
        try:
            wb = load_workbook(filename=self.path, read_only=True)
            ws = wb.active
            headers = [cell.value or "" for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
            data = [[cell if cell else "" for cell in row] for row in ws.iter_rows(min_row=2, values_only=True)]
            wb.close()
            self.import_loaded.emit(headers, data)
        except Exception:
            self.error_occurred.emit(traceback.format_exc())

class LeadsManager(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_list = None
        self.all_data = []
        self.all_headers = []
        self.current_offset = 0
        self.total_rows = 0
        self.dirty_rows = set()
        self.thread = LeadsDataThread(self)
        self.thread.data_loaded.connect(self._on_data_loaded)
        self.thread.load_progress.connect(self._on_progress)
        self.thread.save_finished.connect(self._on_save_finished)
        self.thread.error_occurred.connect(self._on_error)
        self.thread.import_loaded.connect(self._on_import_loaded)

        self._build_ui()
        self._refresh_list()
        
    def _build_ui(self):
        layout = QHBoxLayout(self)

        # Left side: list of lead files
        left_layout = QVBoxLayout()
        self.list_widget = QListWidget()
        self.list_widget.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.list_widget.currentTextChanged.connect(self._on_list_selection_changed)
        left_layout.addWidget(QLabel("<b>Lead Lists</b>"))
        left_layout.addWidget(self.list_widget)

        btn_new = QPushButton()
        btn_new.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_FileIcon))
        btn_new.clicked.connect(self._new_list)
        btn_del = QPushButton()
        btn_del.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_TrashIcon))
        btn_del.clicked.connect(self._delete_list)

        left_buttons = QHBoxLayout()
        left_buttons.addWidget(btn_new)
        left_buttons.addWidget(btn_del)
        left_layout.addLayout(left_buttons)

        # Right side: table view, import/save etc.
        right_layout = QVBoxLayout()

        btn_import = QPushButton()
        btn_import.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogOpenButton))
        btn_import.clicked.connect(self._import)
        btn_save = QPushButton()
        btn_save.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        btn_save.clicked.connect(self._save)
        self.btn_save = btn_save
        self.btn_save.setEnabled(False)

        btn_dedupe = QPushButton()
        btn_dedupe.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogResetButton))
        btn_dedupe.setToolTip("Remove duplicate rows")
        btn_dedupe.clicked.connect(self._remove_duplicates)

        btn_nav_prev = QPushButton()
        btn_nav_prev.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ArrowBack))
        btn_nav_prev.clicked.connect(self._prev_chunk)

        btn_nav_next = QPushButton()
        btn_nav_next.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ArrowForward))
        btn_nav_next.clicked.connect(self._next_chunk)

        action_layout = QHBoxLayout()
        action_layout.addWidget(btn_import)
        action_layout.addWidget(btn_save)
        action_layout.addWidget(btn_dedupe)
        action_layout.addStretch(1)
        action_layout.addWidget(btn_nav_prev)
        action_layout.addWidget(btn_nav_next)
        right_layout.addLayout(action_layout)

        self.status_label = QLabel()
        right_layout.addWidget(self.status_label)

        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("Filter leads...")
        self.filter_input.textChanged.connect(self._filter_table)
        right_layout.addWidget(self.filter_input)

        self.table = QTableWidget()
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.horizontalHeader().setSectionsMovable(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.AllEditTriggers)
        self.table.itemChanged.connect(self._track_dirty)
        right_layout.addWidget(self.table)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        right_layout.addWidget(self.progress_bar)

        layout.addLayout(left_layout, 1)
        layout.addLayout(right_layout, 3)

    def _filter_table(self, text):
        text = text.lower()
        for r in range(self.table.rowCount()):
            match = any(
                text in (self.table.item(r, c).text().lower() if self.table.item(r, c) else "")
                for c in range(self.table.columnCount())
            )
            self.table.setRowHidden(r, not match)
            
    def _refresh_list(self):
        self.list_widget.clear()
        for fname in sorted(os.listdir(DATA_DIR)):
            if fname.endswith(".xlsx"):
                self.list_widget.addItem(os.path.splitext(fname)[0])

    def _on_list_selection_changed(self, name):
        if name:
            self._load_list(name)

    def _load_list(self, name: str):
        path = os.path.join(DATA_DIR, f"{name}.xlsx")
        if os.path.exists(path):
            self.current_list = path
            self.current_offset = 0
            self.thread.load_data(path)
            self.progress_bar.setVisible(True)
            self.progress_bar.setMaximum(0)

    def _on_data_loaded(self, headers, data, total_rows, offset):
        self.progress_bar.setVisible(False)
        self.all_headers = headers
        self.all_data = data or []  # Make sure it's not None
        self.total_rows = total_rows or 0
        self.current_offset = offset
        self._render_page()


    def _render_page(self):
        self.table.setRowCount(0)
        self.table.setColumnCount(len(self.all_headers))
        self.table.setHorizontalHeaderLabels(self.all_headers)

        start = self.current_offset
        end = min(start + CHUNK_SIZE, len(self.all_data))
        for r_idx, row in enumerate(self.all_data[start:end]):
            self.table.insertRow(r_idx)
            for c_idx, value in enumerate(row):
                item = QTableWidgetItem(str(value) if value else "")
                self.table.setItem(r_idx, c_idx, item)
        self.status_label.setText(f"Showing {start + 1} - {end} of {len(self.all_data)}")
        self.btn_save.setEnabled(True)

    def _next_chunk(self):
        if self.current_offset + CHUNK_SIZE < len(self.all_data):
            self.current_offset += CHUNK_SIZE
            self._render_page()

    def _prev_chunk(self):
        if self.current_offset - CHUNK_SIZE >= 0:
            self.current_offset -= CHUNK_SIZE
            self._render_page()

    def closeEvent(self, event):
        self.thread.stop()
        event.accept()
    def _track_dirty(self, item):
        self.dirty_rows.add(self.current_offset + item.row())

    def _on_progress(self, current, total):
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)

    def _on_save_finished(self, success, message):
        self.progress_bar.setVisible(False)
        QMessageBox.information(self, "Save", message)

    def _on_error(self, message):
        self.progress_bar.setVisible(False)
        QMessageBox.critical(self, "Error", message)

    def _save(self):
        if not self.current_list or not self.all_headers:
            return
        headers = self.all_headers
        data = [headers]
        for row in self.all_data:
            data.append([str(cell) if cell is not None else "" for cell in row])
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(0)
        self.thread.save_data(self.current_list, data)

    def _show_context_menu(self, pos):
        menu = QMenu(self)
        action_copy = QAction("Copy Row", self)
        action_delete = QAction("Delete Selected Rows", self)
        action_copy.triggered.connect(self._copy_row)
        action_delete.triggered.connect(self._delete_selected_rows)
        menu.addAction(action_copy)
        menu.addAction(action_delete)
        menu.exec(self.table.mapToGlobal(pos))

    def _copy_row(self):
        row = self.table.currentRow()
        if row >= 0:
            values = [self.table.item(row, c).text() if self.table.item(row, c) else "" for c in range(self.table.columnCount())]
            QApplication.clipboard().setText("\t".join(values))

    def _delete_selected_rows(self):
        selected = sorted(set(i.row() for i in self.table.selectedIndexes()), reverse=True)
        for view_row in selected:
            data_idx = self.current_offset + view_row
            if 0 <= data_idx < len(self.all_data):
                self.all_data.pop(data_idx)

        # re-render the table so the UI updates
        self._render_page()

        # immediately persist the change to disk
        self._save()

    def _remove_duplicates(self):
        seen = set()
        filtered = []
        for row in self.all_data:
            key = tuple(row)
            if key not in seen:
                seen.add(key)
                filtered.append(row)
        self.all_data = filtered
        self._render_page()
        QMessageBox.information(self, "Duplicates Removed", "Duplicates removed successfully.")

    def _new_list(self):
        name, ok = QInputDialog.getText(self, "New List", "Enter name:")
        if ok and name:
            name = re.sub(r"[\\/:*?<>|]+", "_", name.strip())
            path = os.path.join(DATA_DIR, f"{name}.xlsx")
            if not os.path.exists(path):
                Workbook().save(path)
                self._refresh_list()

    def _delete_list(self):
        item = self.list_widget.currentItem()
        if item:
            path = os.path.join(DATA_DIR, f"{item.text()}.xlsx")
            if os.path.exists(path):
                confirm = QMessageBox.question(self, "Delete", f"Are you sure to delete {item.text()}?")
                if confirm == QMessageBox.StandardButton.Yes:
                    os.remove(path)
                    self._refresh_list()
                    
    # … later in class LeadsManager …

    def _import(self):
        current = self.list_widget.currentItem()
        if not current:
            return
        file, _ = QFileDialog.getOpenFileName(self, "Import", "", "Excel Files (*.xlsx)")
        if not file:
            return
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(0)
        self._import_file_path = file
        self.thread.import_excel(file)

    # ← replace this entire method with the new synchronous-save version ↓
    def _on_import_loaded(self, new_headers, new_data):
        # Hide progress indicator
        self.progress_bar.setVisible(False)

        # If we already have headers and they differ, prompt for column mapping
        if self.all_headers and new_headers != self.all_headers:
            dialog = QDialog(self)
            dialog.setWindowTitle("Merge Columns")
            layout = QGridLayout(dialog)
            dropdowns = []

            for i, header in enumerate(new_headers):
                src_label = QLabel(f"Source: {header}")
                dst_combo = QComboBox()
                dst_combo.addItems(self.all_headers)
                dropdowns.append((i, dst_combo))
                layout.addWidget(src_label, i, 0)
                layout.addWidget(dst_combo, i, 1)

            buttons = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
            )
            layout.addWidget(buttons, len(new_headers), 0, 1, 2)
            buttons.accepted.connect(dialog.accept)
            buttons.rejected.connect(dialog.reject)

            if dialog.exec() == QDialog.DialogCode.Accepted:
                mapping = [combo.currentIndex() for _, combo in dropdowns]
                for row in new_data:
                    mapped = [""] * len(self.all_headers)
                    for src_idx, dst_idx in enumerate(mapping):
                        if src_idx < len(row):
                            mapped[dst_idx] = row[src_idx]
                    self.all_data.append(mapped)
        else:
            # No headers yet, or headers match: just append directly
            self.all_data.extend(new_data)

        # If this was the first import, adopt its headers
        if not self.all_headers and new_headers:
            self.all_headers = new_headers

        # Refresh the table view
        self._render_page()

        # **Immediately save the merged data to disk (synchronous)**
        try:
            wb = Workbook()
            ws = wb.active

            # Write headers
            ws.append(self.all_headers)

            # Write every row of merged data
            for row in self.all_data:
                ws.append([str(cell) if cell is not None else "" for cell in row])

            wb.save(self.current_list)
            QMessageBox.information(self, "Save", "Leads saved.")
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save merged leads:\n{e}")

if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    win = LeadsManager()
    win.show()
    sys.exit(app.exec())
