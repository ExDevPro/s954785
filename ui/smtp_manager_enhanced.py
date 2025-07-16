# ui/smtp_manager_enhanced.py
"""
Enhanced SMTP Manager with Professional UI and Fixed Data Persistence
"""

from ui.enhanced_base_manager import EnhancedBaseManager
from PyQt6.QtWidgets import QMessageBox, QFileDialog
import os
import csv
from openpyxl import Workbook, load_workbook
from core.utils.logger import get_module_logger

logger = get_module_logger(__name__)

class SMTPManagerEnhanced(EnhancedBaseManager):
    """Enhanced SMTP Manager with professional UI and data persistence"""
    
    def __init__(self, parent=None):
        # Default headers for SMTP
        self.default_headers = [
            "SMTP Server", "Port", "Username", "Password", "Use TLS", 
            "Use SSL", "From Name", "From Email", "Notes"
        ]
        
        super().__init__(
            manager_type="SMTP",
            data_subdir="smtps", 
            file_extension=".xlsx",
            parent=parent
        )
    
    def _create_new_list_structure(self, list_name: str):
        """Create new SMTP list structure with folder and Excel file"""
        try:
            # Create folder for the list
            list_folder = os.path.join(self.data_dir, list_name)
            os.makedirs(list_folder, exist_ok=True)
            
            # Create Excel file with default headers
            excel_path = os.path.join(list_folder, f"{list_name}.xlsx")
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(self.default_headers)
            workbook.save(excel_path)
            
            logger.info(f"Created SMTP list structure: {list_folder}")
            
        except Exception as e:
            logger.error(f"Failed to create SMTP list structure: {str(e)}")
            raise e
    
    def _get_list_data_path(self, list_name: str) -> str:
        """Get path to SMTP list Excel file"""
        return os.path.join(self.data_dir, list_name, f"{list_name}.xlsx")
    
    def _load_list_data(self, list_name: str):
        """Load SMTP data from Excel file"""
        try:
            data_path = self._get_list_data_path(list_name)
            
            if not os.path.exists(data_path):
                self._create_empty_list_data()
                return
            
            # Load from Excel
            workbook = load_workbook(data_path, read_only=True)
            worksheet = workbook.active
            
            # Read headers
            self.headers = []
            header_row = worksheet[1]
            for cell in header_row:
                if cell.value:
                    self.headers.append(str(cell.value))
                else:
                    break
            
            if not self.headers:
                self.headers = self.default_headers.copy()
            
            # Read data
            self.current_data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    row_data = list(row[:len(self.headers)])
                    while len(row_data) < len(self.headers):
                        row_data.append("")
                    
                    row_data = [str(cell) if cell is not None else "" for cell in row_data]
                    self.current_data.append(row_data)
            
            workbook.close()
            self._update_table_display()
            
            logger.info(f"Loaded SMTP list: {list_name} ({len(self.current_data)} SMTPs)")
            
        except Exception as e:
            logger.error(f"Failed to load SMTP data: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to load SMTP data:\n{str(e)}")
            self._create_empty_list_data()
    
    def _create_empty_list_data(self):
        """Create empty SMTP data structure"""
        self.headers = self.default_headers.copy()
        self.current_data = []
        self._update_table_display()
    
    def _save_current_data(self):
        """Save current SMTP data to Excel file"""
        if not self.current_list_name:
            return
        
        try:
            data_path = self._get_list_data_path(self.current_list_name)
            os.makedirs(os.path.dirname(data_path), exist_ok=True)
            
            self._collect_table_data()
            
            workbook = Workbook()
            worksheet = workbook.active
            
            # Write headers
            for col_idx, header in enumerate(self.headers, 1):
                worksheet.cell(row=1, column=col_idx, value=header)
            
            # Write data
            for row_idx, row_data in enumerate(self.current_data, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    if col_idx <= len(self.headers):
                        worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
            
            workbook.save(data_path)
            workbook.close()
            
            logger.info(f"Saved SMTP data: {self.current_list_name} ({len(self.current_data)} SMTPs)")
            
        except Exception as e:
            logger.error(f"Failed to save SMTP data: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to save SMTP data:\n{str(e)}")
    
    def _collect_table_data(self):
        """Collect current data from table widget"""
        if not self.table_widget:
            return
        
        self.current_data = []
        for row_idx in range(self.table_widget.rowCount()):
            row_data = []
            for col_idx in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row_idx, col_idx)
                cell_value = item.text() if item else ""
                row_data.append(cell_value)
            self.current_data.append(row_data)
    
    def _import_data(self):
        """Import SMTP data from file"""
        if not self.current_list_name:
            QMessageBox.warning(self, "Error", "Please select an SMTP list first!")
            return
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import SMTP Data",
            "", "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, 0)
            
            imported_data = []
            imported_headers = []
            
            if file_path.lower().endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8', newline='') as csvfile:
                    sample = csvfile.read(1024)
                    csvfile.seek(0)
                    dialect = csv.Sniffer().sniff(sample)
                    reader = csv.reader(csvfile, dialect)
                    imported_headers = next(reader, [])
                    
                    for row in reader:
                        if any(cell.strip() for cell in row):
                            imported_data.append(row)
            else:
                workbook = load_workbook(file_path, read_only=True)
                worksheet = workbook.active
                
                header_row = worksheet[1]
                imported_headers = []
                for cell in header_row:
                    if cell.value:
                        imported_headers.append(str(cell.value))
                    else:
                        break
                
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        row_data = [str(cell) if cell is not None else "" for cell in row[:len(imported_headers)]]
                        imported_data.append(row_data)
                
                workbook.close()
            
            # Merge or replace logic
            if self.current_data:
                reply = QMessageBox.question(
                    self, "Import Options",
                    f"Found {len(imported_data)} SMTP configs to import.\n\n"
                    f"Current list has {len(self.current_data)} configs.\n\n"
                    f"Replace existing data?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
                )
                
                if reply == QMessageBox.StandardButton.Cancel:
                    return
                elif reply == QMessageBox.StandardButton.Yes:
                    self.current_data = []
                    self.headers = imported_headers if imported_headers else self.default_headers
            else:
                self.headers = imported_headers if imported_headers else self.default_headers
            
            # Process imported data
            for row_data in imported_data:
                processed_row = row_data[:len(self.headers)]
                while len(processed_row) < len(self.headers):
                    processed_row.append("")
                self.current_data.append(processed_row)
            
            self._update_table_display()
            
            if self.auto_save_enabled:
                self._save_current_data()
            
            self.progress_bar.setVisible(False)
            
            QMessageBox.information(
                self, "Import Successful",
                f"Successfully imported {len(imported_data)} SMTP configs!"
            )
            
        except Exception as e:
            self.progress_bar.setVisible(False)
            logger.error(f"Failed to import SMTP data: {str(e)}")
            QMessageBox.critical(self, "Import Error", f"Failed to import SMTP data:\n{str(e)}")
    
    def _export_data(self):
        """Export SMTP data to file"""
        if not self.current_data:
            QMessageBox.warning(self, "Error", "No data to export!")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export SMTP Data",
            f"{self.current_list_name}_smtp.xlsx",
            "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, len(self.current_data))
            
            if file_path.lower().endswith('.csv'):
                with open(file_path, 'w', encoding='utf-8', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(self.headers)
                    
                    for idx, row_data in enumerate(self.current_data):
                        writer.writerow(row_data)
                        self.progress_bar.setValue(idx + 1)
            else:
                workbook = Workbook()
                worksheet = workbook.active
                
                for col_idx, header in enumerate(self.headers, 1):
                    worksheet.cell(row=1, column=col_idx, value=header)
                
                for row_idx, row_data in enumerate(self.current_data, 2):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    self.progress_bar.setValue(row_idx - 1)
                
                workbook.save(file_path)
                workbook.close()
            
            self.progress_bar.setVisible(False)
            
            QMessageBox.information(
                self, "Export Successful",
                f"Successfully exported {len(self.current_data)} SMTP configs to:\n{file_path}"
            )
            
        except Exception as e:
            self.progress_bar.setVisible(False)
            logger.error(f"Failed to export SMTP data: {str(e)}")
            QMessageBox.critical(self, "Export Error", f"Failed to export SMTP data:\n{str(e)}")
    
    def _list_exists(self, list_name: str) -> bool:
        """Check if SMTP list exists"""
        return os.path.exists(self._get_list_data_path(list_name))
    
    def _delete_list_structure(self, list_name: str):
        """Delete SMTP list structure"""
        import shutil
        list_folder = os.path.join(self.data_dir, list_name)
        if os.path.exists(list_folder):
            shutil.rmtree(list_folder)
            logger.info(f"Deleted SMTP list: {list_folder}")
    
    def _rename_list(self, old_name: str, new_name: str):
        """Rename SMTP list"""
        old_folder = os.path.join(self.data_dir, old_name)
        new_folder = os.path.join(self.data_dir, new_name)
        
        if os.path.exists(old_folder):
            os.rename(old_folder, new_folder)
            
            old_file = os.path.join(new_folder, f"{old_name}.xlsx")
            new_file = os.path.join(new_folder, f"{new_name}.xlsx")
            if os.path.exists(old_file):
                os.rename(old_file, new_file)
            
            logger.info(f"Renamed SMTP list: {old_name} -> {new_name}")
    
    def _count_items_in_list(self, list_name: str) -> int:
        """Count SMTP configs in a list"""
        try:
            data_path = self._get_list_data_path(list_name)
            if not os.path.exists(data_path):
                return 0
            
            workbook = load_workbook(data_path, read_only=True)
            worksheet = workbook.active
            
            count = 0
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row):
                    count += 1
            
            workbook.close()
            return count
            
        except Exception as e:
            logger.error(f"Failed to count SMTP configs in {list_name}: {str(e)}")
            return 0