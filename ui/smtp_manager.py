from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QListWidget, QTableWidget, QTableWidgetItem,
    QToolButton, QComboBox, QLabel, QFileDialog, QMessageBox, QMenu, QInputDialog,
    QAbstractItemView, QApplication, QStyle, QHeaderView, QDialog, QLineEdit, QPushButton
)
from PyQt6.QtCore import Qt, QDateTime, QThreadPool, QRunnable, pyqtSlot
import os
from openpyxl import load_workbook, Workbook
import traceback
from engine.smtp_worker import test_smtp

DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '../data/smtps'))

class TestSMTPTask(QRunnable):
    def __init__(self, row, smtp_data, callback):
        super().__init__()
        self.row = row
        self.smtp_data = smtp_data
        self.callback = callback

    @pyqtSlot()
    def run(self):
        result = test_smtp(self.smtp_data)
        self.callback(self.row, result)

class SMTPManager(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("SMTP Manager")
        self.thread_pool = QThreadPool()
        self.thread_pools = {}  # key: list file name, value: QThreadPool
        self.smtp_lists = QListWidget()
        self.smtp_lists.itemClicked.connect(self.load_smtp_list)

        self.smtp_table = QTableWidget()
        self.smtp_table.setColumnCount(10)
        self.smtp_table.setHorizontalHeaderLabels([
            "Host", "Port", "Security", "User", "Password", "From Name",
            "From Email", "Status", "Imported Date", "Last Update"
        ])
        self.smtp_table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self.smtp_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.smtp_table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.smtp_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.smtp_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.smtp_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.smtp_table.customContextMenuRequested.connect(self.show_context_menu)
        self.smtp_table.cellChanged.connect(self.update_last_modified)

        self.thread_selector = QComboBox()
        self.thread_selector.addItems([str(i) for i in [1,2,3,4,5,6,7,8,9,10,20,50,100,200,500]])
        self.thread_selector.setCurrentText("1")

        self.toolbar = self.create_toolbar()

        layout = QVBoxLayout()
        top_layout = QHBoxLayout()
        list_buttons = QVBoxLayout()

        self.btn_create_list = QPushButton("➕ New List")
        self.btn_create_list.clicked.connect(self.create_new_list)
        list_buttons.addWidget(self.btn_create_list)

        self.btn_delete_list = QPushButton("🗑 Delete List")
        self.btn_delete_list.clicked.connect(self.delete_selected_list)
        list_buttons.addWidget(self.btn_delete_list)

        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("🔍 Search SMTP Lists")
        self.search_bar.textChanged.connect(self.filter_list_items)
        list_buttons.insertWidget(0, self.search_bar)

        self.smtp_lists.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.smtp_lists.customContextMenuRequested.connect(self.show_list_context_menu)

        left_layout = QVBoxLayout()
        left_layout.addLayout(list_buttons)
        left_layout.addWidget(self.smtp_lists)

        top_layout.addLayout(left_layout, 1)
        top_layout.addWidget(self.smtp_table, 3)

        layout.addLayout(self.toolbar)
        layout.addLayout(top_layout)

        self.scroll_button = QPushButton("⇄ View Hidden Columns")
        self.scroll_button.clicked.connect(self.toggle_scroll_position)
        layout.addWidget(self.scroll_button)

        layout.addWidget(QLabel("Thread Count:"))
        layout.addWidget(self.thread_selector)

        self.setLayout(layout)
        os.makedirs(DATA_DIR, exist_ok=True)
        self.load_smtp_files()
        self.apply_theme()

    def apply_theme(self):
        qss_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../assets/themes/Default.qss'))
        if os.path.exists(qss_path):
            with open(qss_path, 'r', encoding='utf-8') as f:
                self.setStyleSheet(f.read())

    def create_toolbar(self):
        layout = QHBoxLayout()
        self.btn_import = QToolButton()
        self.btn_import.setIcon(QApplication.style().standardIcon(QStyle.StandardPixmap.SP_DialogOpenButton))
        self.btn_import.setToolTip("Import SMTP File")
        self.btn_import.clicked.connect(self.import_smtp_file)
        layout.addWidget(self.btn_import)

        self.btn_add = QToolButton()
        self.btn_add.setIcon(QApplication.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogNewFolder))
        self.btn_add.setToolTip("Add New SMTP")
        self.btn_add.clicked.connect(self.add_new_smtp)
        layout.addWidget(self.btn_add)

        self.btn_test_all = QToolButton()
        self.btn_test_all.setIcon(QApplication.style().standardIcon(QStyle.StandardPixmap.SP_MediaPlay))
        self.btn_test_all.setToolTip("Test All SMTPs")
        self.btn_test_all.clicked.connect(self.test_all_smtps)
        layout.addWidget(self.btn_test_all)

        self.btn_save = QToolButton()
        self.btn_save.setIcon(QApplication.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        self.btn_save.setToolTip("Save SMTP File")
        self.btn_save.clicked.connect(self.save_smtp_file)
        layout.addWidget(self.btn_save)

        self.btn_delete = QToolButton()
        self.btn_delete.setIcon(QApplication.style().standardIcon(QStyle.StandardPixmap.SP_TrashIcon))
        self.btn_delete.setToolTip("Delete Selected SMTP")
        self.btn_delete.clicked.connect(self.delete_selected_smtp)
        layout.addWidget(self.btn_delete)

        self.btn_remove_duplicates = QToolButton()
        self.btn_remove_duplicates.setIcon(QApplication.style().standardIcon(QStyle.StandardPixmap.SP_DialogResetButton))
        self.btn_remove_duplicates.setToolTip("Remove Duplicate SMTPs")
        self.btn_remove_duplicates.clicked.connect(self.remove_duplicates)
        layout.addWidget(self.btn_remove_duplicates)

        return layout

    def toggle_scroll_position(self):
        scrollbar = self.smtp_table.horizontalScrollBar()
        if scrollbar.value() < scrollbar.maximum():
            scrollbar.setValue(scrollbar.maximum())
        else:
            scrollbar.setValue(0)

    def update_last_modified(self, row, column):
        now = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
        self.smtp_table.blockSignals(True)
        self.smtp_table.setItem(row, 9, QTableWidgetItem(now))
        self.smtp_table.blockSignals(False)




    def validate_smtp_data(self, data):
        if not data["Host"] or not data["User"] or not data["Password"]:
            return False, "Host, User, and Password are required."
        if not data["Port"].isdigit():
            return False, "Port must be a number."
        if "@" not in data["From Email"]:
            return False, "Invalid email format."
        return True, ""

# Remaining methods omitted here for space - added in next chunk
    def test_all_smtps(self):
        current_item = self.smtp_lists.currentItem()
        if not current_item:
            QMessageBox.warning(self, "No File Selected", "Please select a list before testing.")
            return

        list_name = current_item.text()
        if list_name not in self.thread_pools:
            self.thread_pools[list_name] = QThreadPool()

        pool = self.thread_pools[list_name]
        self.smtp_table.blockSignals(True)

        for row in range(self.smtp_table.rowCount()):
            smtp_data = {
                "Host": self.smtp_table.item(row, 0).text(),
                "Port": self.smtp_table.item(row, 1).text(),
                "Security": self.smtp_table.item(row, 2).text(),
                "User": self.smtp_table.item(row, 3).text(),
                "Password": self.smtp_table.item(row, 4).text()
            }
            task = TestSMTPTask(row, smtp_data, self.on_test_complete)
            pool.start(task)

        self.smtp_table.blockSignals(False)

    def on_test_complete(self, row, result):
        self.smtp_table.setItem(row, 7, QTableWidgetItem(result["status"]))
        now_str = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
        self.smtp_table.setItem(row, 9, QTableWidgetItem(now_str))
        self.smtp_table.setItem(row, 8, QTableWidgetItem(now_str))
        self.save_smtp_file()

    def import_smtp_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open SMTP Excel File", DATA_DIR, "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            wb = load_workbook(filename=file_path)
            ws = wb.active
            self.smtp_table.setRowCount(0)
            now_str = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_data = list(row) + ["", now_str, now_str]
                row_position = self.smtp_table.rowCount()
                self.smtp_table.insertRow(row_position)
                for col_index, value in enumerate(row_data):
                    self.smtp_table.setItem(row_position, col_index, QTableWidgetItem(str(value)))
            self.save_smtp_file()
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to load file:\n{str(e)}")
            traceback.print_exc()

    def save_smtp_file(self):
        current_item = self.smtp_lists.currentItem()
        if not current_item:
            QMessageBox.warning(self, "No File Selected", "Please select a file from the list to save.")
            return
        file_path = os.path.join(DATA_DIR, current_item.text())
        wb = Workbook()
        ws = wb.active
        headers = [self.smtp_table.horizontalHeaderItem(i).text() for i in range(self.smtp_table.columnCount())]
        ws.append(headers)
        for row in range(self.smtp_table.rowCount()):
            row_data = []
            for col in range(self.smtp_table.columnCount()):
                item = self.smtp_table.item(row, col)
                row_data.append(item.text() if item else "")
            ws.append(row_data)
        try:
            wb.save(file_path)
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save file:\n{str(e)}")

    def delete_selected_smtp(self):
        selected = self.smtp_table.currentRow()
        if selected >= 0:
            self.smtp_table.removeRow(selected)

    def test_selected_smtp(self):
        row = self.smtp_table.currentRow()
        if row < 0:
            return
        smtp_data = {
            "Host": self.smtp_table.item(row, 0).text(),
            "Port": self.smtp_table.item(row, 1).text(),
            "Security": self.smtp_table.item(row, 2).text(),
            "User": self.smtp_table.item(row, 3).text(),
            "Password": self.smtp_table.item(row, 4).text()
        }
        current_item = self.smtp_lists.currentItem()
        if not current_item:
            QMessageBox.warning(self, "No List Selected", "Please select a list before testing.")
            return
        list_name = current_item.text()
        if list_name not in self.thread_pools:
            self.thread_pools[list_name] = QThreadPool()
        task = TestSMTPTask(row, smtp_data, self.on_test_complete)
        self.thread_pools[list_name].start(task)

    def add_new_smtp(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add New SMTP")
        layout = QVBoxLayout(dialog)
        fields = ["Host", "Port", "Security", "User", "Password", "From Name", "From Email"]
        inputs = {}
        for field in fields:
            row = QHBoxLayout()
            label = QLabel(field + ":")
            line_edit = QLineEdit()
            inputs[field] = line_edit
            row.addWidget(label)
            row.addWidget(line_edit)
            layout.addLayout(row)
        buttons = QHBoxLayout()
        ok_btn = QPushButton("Add")
        cancel_btn = QPushButton("Cancel")
        ok_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        buttons.addWidget(ok_btn)
        buttons.addWidget(cancel_btn)
        layout.addLayout(buttons)
        if dialog.exec():
            smtp_data = {field: inputs[field].text() for field in fields}
            valid, msg = self.validate_smtp_data(smtp_data)
            if not valid:
                QMessageBox.warning(self, "Validation Failed", msg)
                return
            now_str = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
            row_position = self.smtp_table.rowCount()
            self.smtp_table.insertRow(row_position)
            for i, field in enumerate(fields):
                self.smtp_table.setItem(row_position, i, QTableWidgetItem(smtp_data[field]))
            self.smtp_table.setItem(row_position, 7, QTableWidgetItem(""))
            self.smtp_table.setItem(row_position, 8, QTableWidgetItem(now_str))
            self.smtp_table.setItem(row_position, 9, QTableWidgetItem(now_str))

    def create_new_list(self):
        text, ok = QInputDialog.getText(self, "Create New SMTP List", "Enter list name:")
        if ok and text:
            filename = os.path.join(DATA_DIR, f"{text}.xlsx")
            if os.path.exists(filename):
                QMessageBox.warning(self, "File Exists", "A list with this name already exists.")
            else:
                wb = Workbook()
                ws = wb.active
                ws.append([
                    "Host", "Port", "Security", "User", "Password", "From Name",
                    "From Email", "Status", "Imported Date", "Last Update"
                ])
                wb.save(filename)
                self.load_smtp_files()

    def delete_selected_list(self):
        item = self.smtp_lists.currentItem()
        if item:
            filepath = os.path.join(DATA_DIR, item.text())
            if os.path.exists(filepath):
                os.remove(filepath)
                self.load_smtp_files()
                self.smtp_table.setRowCount(0)

    def load_smtp_files(self):
        self.smtp_lists.clear()
        for file in os.listdir(DATA_DIR):
            if file.endswith(".xlsx"):
                self.smtp_lists.addItem(file)

    def load_smtp_list(self, item):
        file_path = os.path.join(DATA_DIR, item.text())
        try:
            wb = load_workbook(filename=file_path)
            ws = wb.active
            self.smtp_table.setRowCount(0)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_position = self.smtp_table.rowCount()
                self.smtp_table.insertRow(row_position)
                for col_index, value in enumerate(row):
                    self.smtp_table.setItem(row_position, col_index, QTableWidgetItem(str(value)))
        except Exception as e:
            QMessageBox.critical(self, "Load Error", f"Failed to load SMTP list:\n{str(e)}")

    def filter_list_items(self, text):
        for i in range(self.smtp_lists.count()):
            item = self.smtp_lists.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def show_list_context_menu(self, pos):
        menu = QMenu()
        rename_action = menu.addAction("Rename List")
        selected_item = self.smtp_lists.itemAt(pos)
        if selected_item:
            action = menu.exec(self.smtp_lists.mapToGlobal(pos))
            if action == rename_action:
                old_name = selected_item.text()
                new_name, ok = QInputDialog.getText(self, "Rename List", "Enter new name:", text=old_name.replace(".xlsx", ""))
                if ok and new_name:
                    old_path = os.path.join(DATA_DIR, old_name)
                    new_path = os.path.join(DATA_DIR, new_name + ".xlsx")
                    if os.path.exists(new_path):
                        QMessageBox.warning(self, "Rename Error", "A file with this name already exists.")
                    else:
                        os.rename(old_path, new_path)
                        self.load_smtp_files()

    def show_context_menu(self, pos):
        menu = QMenu()
        test_action = menu.addAction("Test this SMTP")
        delete_action = menu.addAction("Delete this SMTP")
        action = menu.exec(self.smtp_table.mapToGlobal(pos))
        if action == test_action:
            self.test_selected_smtp()
        elif action == delete_action:
            self.delete_selected_smtp()

    def remove_duplicates(self):
        seen = set()
        rows_to_remove = []
        for row in range(self.smtp_table.rowCount()):
            identifier = tuple(self.smtp_table.item(row, i).text() for i in [0, 1, 3])  # Host, Port, User
            if identifier in seen:
                rows_to_remove.append(row)
            else:
                seen.add(identifier)
        for row in reversed(rows_to_remove):
            self.smtp_table.removeRow(row)
        QMessageBox.information(self, "Duplicates Removed", f"{len(rows_to_remove)} duplicate row(s) removed.")
