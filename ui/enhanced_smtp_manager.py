# Enhanced SMTP Manager with Original Functionality Plus Background Workers

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QListWidget, QTableWidget, QTableWidgetItem,
    QToolButton, QComboBox, QLabel, QFileDialog, QMessageBox, QMenu, QInputDialog,
    QAbstractItemView, QApplication, QStyle, QHeaderView, QDialog, QLineEdit, QPushButton,
    QFormLayout, QDialogButtonBox, QProgressBar, QFrame, QSplitter, QTextEdit, QScrollArea,
    QSizePolicy, QSpinBox
)
from PyQt6.QtCore import Qt, QDateTime, QThreadPool, QRunnable, pyqtSlot, pyqtSignal, QThread, QTimer
from PyQt6.QtGui import QFont, QAction
import os
from openpyxl import load_workbook, Workbook
import traceback
import json

# Import the SMTP testing function
try:
    from engine.smtp_worker import test_smtp
except ImportError:
    # Fallback test function
    def test_smtp(smtp_data):
        import smtplib
        try:
            if smtp_data.get('Security', '').lower() == 'ssl':
                server = smtplib.SMTP_SSL(smtp_data['Host'], int(smtp_data['Port']))
            else:
                server = smtplib.SMTP(smtp_data['Host'], int(smtp_data['Port']))
                if smtp_data.get('Security', '').lower() == 'tls':
                    server.starttls()
            
            server.login(smtp_data['User'], smtp_data['Password'])
            server.quit()
            return True
        except Exception:
            return False

DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '../data/smtps'))

class CreateListDialog(QDialog):
    """Professional dialog for creating new SMTP lists with descriptions."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Create New SMTP List")
        self.setModal(True)
        self.setFixedSize(400, 250)
        
        layout = QVBoxLayout(self)
        
        # List name
        layout.addWidget(QLabel("List Name:"))
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("Enter SMTP list name...")
        layout.addWidget(self.name_edit)
        
        # Description
        layout.addWidget(QLabel("Description (Optional):"))
        self.desc_edit = QTextEdit()
        self.desc_edit.setPlaceholderText("Enter description for this SMTP list...")
        self.desc_edit.setMaximumHeight(80)
        layout.addWidget(self.desc_edit)
        
        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        self.name_edit.textChanged.connect(self._validate_input)
        self._validate_input()
    
    def _validate_input(self):
        name = self.name_edit.text().strip()
        valid = bool(name and not name.endswith('.xlsx') and len(name) > 0)
        self.findChild(QDialogButtonBox).button(QDialogButtonBox.StandardButton.Ok).setEnabled(valid)
    
    def get_list_data(self):
        name = self.name_edit.text().strip()
        if not name.endswith('.xlsx'):
            name += '.xlsx'
        return {
            'name': name,
            'description': self.desc_edit.toPlainText().strip()
        }

class SMTPDialog(QDialog):
    """Enhanced dialog for adding/editing SMTP configurations."""
    
    def __init__(self, parent=None, smtp_data=None, title="Add SMTP"):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setFixedSize(500, 400)
        
        layout = QVBoxLayout(self)
        
        # Form layout
        form_layout = QFormLayout()
        
        # SMTP fields
        self.host_edit = QLineEdit()
        self.host_edit.setPlaceholderText("smtp.gmail.com")
        form_layout.addRow("Host:", self.host_edit)
        
        self.port_edit = QLineEdit()
        self.port_edit.setPlaceholderText("587")
        form_layout.addRow("Port:", self.port_edit)
        
        self.security_combo = QComboBox()
        self.security_combo.addItems(["TLS", "SSL", "None"])
        form_layout.addRow("Security:", self.security_combo)
        
        self.user_edit = QLineEdit()
        self.user_edit.setPlaceholderText("your-email@gmail.com")
        form_layout.addRow("Username:", self.user_edit)
        
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_edit.setPlaceholderText("your-password or app-password")
        form_layout.addRow("Password:", self.password_edit)
        
        self.from_name_edit = QLineEdit()
        self.from_name_edit.setPlaceholderText("Your Name")
        form_layout.addRow("From Name:", self.from_name_edit)
        
        self.from_email_edit = QLineEdit()
        self.from_email_edit.setPlaceholderText("your-email@gmail.com")
        form_layout.addRow("From Email:", self.from_email_edit)
        
        layout.addLayout(form_layout)
        
        # Test button
        self.test_btn = QPushButton("🔧 Test Connection")
        self.test_btn.clicked.connect(self._test_connection)
        layout.addWidget(self.test_btn)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Status label
        self.status_label = QLabel("")
        layout.addWidget(self.status_label)
        
        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        # Load existing data if provided
        if smtp_data:
            self._load_smtp_data(smtp_data)
    
    def _load_smtp_data(self, smtp_data):
        """Load existing SMTP data into the form."""
        self.host_edit.setText(smtp_data.get('Host', ''))
        self.port_edit.setText(smtp_data.get('Port', ''))
        
        security = smtp_data.get('Security', 'TLS')
        index = self.security_combo.findText(security)
        if index >= 0:
            self.security_combo.setCurrentIndex(index)
        
        self.user_edit.setText(smtp_data.get('User', ''))
        self.password_edit.setText(smtp_data.get('Password', ''))
        self.from_name_edit.setText(smtp_data.get('From Name', ''))
        self.from_email_edit.setText(smtp_data.get('From Email', ''))
    
    def _test_connection(self):
        """Test the SMTP connection."""
        smtp_data = self.get_smtp_data()
        
        # Validate required fields
        if not all([smtp_data['Host'], smtp_data['Port'], smtp_data['User'], smtp_data['Password']]):
            QMessageBox.warning(self, "Validation Error", "Please fill in all required fields (Host, Port, User, Password).")
            return
        
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # Indeterminate
        self.status_label.setText("Testing connection...")
        self.test_btn.setEnabled(False)
        
        # Create worker for testing
        self.test_worker = SMTPTestWorker(smtp_data)
        self.test_worker.test_completed.connect(self._on_test_completed)
        self.test_worker.start()
    
    def _on_test_completed(self, success, message):
        """Handle test completion."""
        self.progress_bar.setVisible(False)
        self.test_btn.setEnabled(True)
        
        if success:
            self.status_label.setText("✅ Connection successful!")
            self.status_label.setStyleSheet("color: green;")
        else:
            self.status_label.setText(f"❌ Connection failed: {message}")
            self.status_label.setStyleSheet("color: red;")
    
    def get_smtp_data(self):
        """Get SMTP data from the form."""
        return {
            'Host': self.host_edit.text().strip(),
            'Port': self.port_edit.text().strip(),
            'Security': self.security_combo.currentText(),
            'User': self.user_edit.text().strip(),
            'Password': self.password_edit.text().strip(),
            'From Name': self.from_name_edit.text().strip(),
            'From Email': self.from_email_edit.text().strip()
        }

class SMTPTestWorker(QThread):
    """Background worker for SMTP testing."""
    
    test_completed = pyqtSignal(bool, str)
    
    def __init__(self, smtp_data):
        super().__init__()
        self.smtp_data = smtp_data
    
    def run(self):
        try:
            result = test_smtp(self.smtp_data)
            if result:
                self.test_completed.emit(True, "Connection successful")
            else:
                self.test_completed.emit(False, "Authentication failed")
        except Exception as e:
            self.test_completed.emit(False, str(e))

class TestSMTPTask(QRunnable):
    """Background task for testing multiple SMTP configurations."""
    
    def __init__(self, row, smtp_data, callback):
        super().__init__()
        self.row = row
        self.smtp_data = smtp_data
        self.callback = callback

    @pyqtSlot()
    def run(self):
        result = test_smtp(self.smtp_data)
        self.callback(self.row, result)

class EnhancedSMTPManager(QWidget):
    """Enhanced SMTP Manager with original functionality plus improvements."""
    
    counts_changed = pyqtSignal()
    
    def __init__(self, parent=None, config=None):
        super().__init__(parent)
        self.setWindowTitle("SMTP Manager")
        self.config = config
        
        # Threading
        self.thread_pool = QThreadPool()
        self.thread_pools = {}  # key: list file name, value: QThreadPool
        
        # Auto-save timer
        self.auto_save_timer = QTimer()
        self.auto_save_timer.timeout.connect(self._auto_save_data)
        self.auto_save_timer.setSingleShot(True)
        
        self._setup_ui()
        self._load_smtp_files()
        self._apply_styling()
    
    def _setup_ui(self):
        """Setup the user interface."""
        main_layout = QHBoxLayout(self)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Left panel - Lists
        left_panel = self._create_left_panel()
        main_layout.addWidget(left_panel, 1)
        
        # Splitter for resizable panels
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Right panel - Data view
        right_panel = self._create_right_panel()
        splitter.addWidget(right_panel)
        
        # Description panel
        desc_panel = self._create_description_panel()
        splitter.addWidget(desc_panel)
        
        splitter.setSizes([600, 200])
        main_layout.addWidget(splitter, 3)
    
    def _create_left_panel(self):
        """Create the left panel with list management."""
        panel = QFrame()
        panel.setFrameStyle(QFrame.Shape.StyledPanel)
        panel.setMaximumWidth(350)
        panel.setMinimumWidth(250)
        
        layout = QVBoxLayout(panel)
        
        # Header
        header = QLabel("📧 SMTP Lists")
        header.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setStyleSheet("QLabel { background: #e74c3c; color: white; padding: 8px; border-radius: 4px; }")
        layout.addWidget(header)
        
        # Search
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("🔍 Search SMTP Lists")
        self.search_bar.textChanged.connect(self.filter_list_items)
        layout.addWidget(self.search_bar)
        
        # Lists
        self.smtp_lists = QListWidget()
        self.smtp_lists.itemClicked.connect(self.load_smtp_list)
        self.smtp_lists.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.smtp_lists.customContextMenuRequested.connect(self.show_list_context_menu)
        layout.addWidget(self.smtp_lists)
        
        # Buttons
        buttons_layout = QVBoxLayout()
        
        self.btn_create_list = QPushButton("➕ New List")
        self.btn_create_list.clicked.connect(self.create_new_list)
        self.btn_create_list.setStyleSheet("QPushButton { background: #27ae60; color: white; padding: 8px; border-radius: 4px; }")
        buttons_layout.addWidget(self.btn_create_list)
        
        self.btn_delete_list = QPushButton("🗑 Delete List")
        self.btn_delete_list.clicked.connect(self.delete_selected_list)
        self.btn_delete_list.setStyleSheet("QPushButton { background: #e74c3c; color: white; padding: 8px; border-radius: 4px; }")
        buttons_layout.addWidget(self.btn_delete_list)
        
        layout.addLayout(buttons_layout)
        
        return panel
    
    def _create_right_panel(self):
        """Create the right panel with SMTP table."""
        panel = QFrame()
        panel.setFrameStyle(QFrame.Shape.StyledPanel)
        
        layout = QVBoxLayout(panel)
        
        # Header
        header_layout = QHBoxLayout()
        self.table_header = QLabel("📊 SMTP Configurations")
        self.table_header.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        header_layout.addWidget(self.table_header)
        
        header_layout.addStretch()
        
        # Thread selector
        header_layout.addWidget(QLabel("Threads:"))
        self.thread_selector = QComboBox()
        self.thread_selector.addItems([str(i) for i in [1,2,3,4,5,6,7,8,9,10,20,50,100,200,500]])
        self.thread_selector.setCurrentText("1")
        header_layout.addWidget(self.thread_selector)
        
        layout.addLayout(header_layout)
        
        # Toolbar
        toolbar_layout = self._create_toolbar()
        layout.addLayout(toolbar_layout)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Status label
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("QLabel { color: #7f8c8d; font-style: italic; }")
        layout.addWidget(self.status_label)
        
        # SMTP table
        self.smtp_table = QTableWidget()
        self.smtp_table.setColumnCount(10)
        self.smtp_table.setHorizontalHeaderLabels([
            "Host", "Port", "Security", "User", "Password", "From Name",
            "From Email", "Status", "Imported Date", "Last Update"
        ])
        self.smtp_table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self.smtp_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.smtp_table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.smtp_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.smtp_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.smtp_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.smtp_table.customContextMenuRequested.connect(self.show_context_menu)
        self.smtp_table.cellChanged.connect(self.update_last_modified)
        self.smtp_table.cellChanged.connect(lambda: self.auto_save_timer.start(1000))
        layout.addWidget(self.smtp_table)
        
        # Scroll button
        self.scroll_button = QPushButton("⇄ View Hidden Columns")
        self.scroll_button.clicked.connect(self.toggle_scroll_position)
        layout.addWidget(self.scroll_button)
        
        return panel
    
    def _create_description_panel(self):
        """Create the description panel."""
        panel = QFrame()
        panel.setFrameStyle(QFrame.Shape.StyledPanel)
        panel.setMinimumWidth(200)
        panel.setMaximumWidth(300)
        
        layout = QVBoxLayout(panel)
        
        # Header
        header = QLabel("📝 Description")
        header.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setStyleSheet("QLabel { background: #95a5a6; color: white; padding: 6px; border-radius: 4px; }")
        layout.addWidget(header)
        
        # Scrollable description area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        self.description_label = QLabel("Select a list to view its description.")
        self.description_label.setWordWrap(True)
        self.description_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.description_label.setStyleSheet("QLabel { padding: 10px; background: white; }")
        
        scroll_area.setWidget(self.description_label)
        layout.addWidget(scroll_area)
        
        # Edit button
        self.btn_edit_desc = QPushButton("✏️ Edit Description")
        self.btn_edit_desc.clicked.connect(self._edit_description)
        self.btn_edit_desc.setEnabled(False)
        layout.addWidget(self.btn_edit_desc)
        
        return panel
    
    def _create_toolbar(self):
        """Create the toolbar."""
        toolbar = QHBoxLayout()
        
        self.btn_add_smtp = QPushButton("➕ Add SMTP")
        self.btn_add_smtp.clicked.connect(self.add_new_smtp)
        self.btn_add_smtp.setStyleSheet("QPushButton { background: #27ae60; color: white; padding: 6px 12px; border-radius: 4px; }")
        toolbar.addWidget(self.btn_add_smtp)
        
        self.btn_edit_smtp = QPushButton("✏️ Edit SMTP")
        self.btn_edit_smtp.clicked.connect(self.edit_selected_smtp)
        toolbar.addWidget(self.btn_edit_smtp)
        
        self.btn_delete_smtp = QPushButton("🗑 Delete")
        self.btn_delete_smtp.clicked.connect(self.delete_selected_smtp)
        self.btn_delete_smtp.setStyleSheet("QPushButton { background: #e74c3c; color: white; padding: 6px 12px; border-radius: 4px; }")
        toolbar.addWidget(self.btn_delete_smtp)
        
        toolbar.addWidget(QLabel("|"))
        
        self.btn_test_smtp = QPushButton("🔧 Test Selected")
        self.btn_test_smtp.clicked.connect(self.test_selected_smtp)
        self.btn_test_smtp.setStyleSheet("QPushButton { background: #3498db; color: white; padding: 6px 12px; border-radius: 4px; }")
        toolbar.addWidget(self.btn_test_smtp)
        
        self.btn_test_all = QPushButton("🔧 Test All")
        self.btn_test_all.clicked.connect(self.test_all_smtp)
        toolbar.addWidget(self.btn_test_all)
        
        toolbar.addWidget(QLabel("|"))
        
        self.btn_import = QPushButton("📥 Import")
        self.btn_import.clicked.connect(self.import_smtp_file)
        self.btn_import.setStyleSheet("QPushButton { background: #9b59b6; color: white; padding: 6px 12px; border-radius: 4px; }")
        toolbar.addWidget(self.btn_import)
        
        self.btn_export = QPushButton("📤 Export")
        self.btn_export.clicked.connect(self.export_smtp_file)
        self.btn_export.setStyleSheet("QPushButton { background: #f39c12; color: white; padding: 6px 12px; border-radius: 4px; }")
        toolbar.addWidget(self.btn_export)
        
        toolbar.addStretch()
        
        return toolbar
    
    def _apply_styling(self):
        """Apply professional styling."""
        self.setStyleSheet("""
            QFrame {
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: #ecf0f1;
            }
            QListWidget {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background: white;
                padding: 5px;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #ecf0f1;
            }
            QListWidget::item:selected {
                background: #e74c3c;
                color: white;
            }
            QTableWidget {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background: white;
                gridline-color: #ecf0f1;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QLineEdit {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                padding: 6px;
            }
            QLineEdit:focus {
                border: 2px solid #e74c3c;
            }
            QPushButton {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                padding: 6px 12px;
                background: #ecf0f1;
            }
            QPushButton:hover {
                background: #d5dbdb;
            }
            QPushButton:pressed {
                background: #bdc3c7;
            }
    
    # ========== List Management Methods ==========
    
    def _load_smtp_files(self):
        """Load all SMTP list files."""
        os.makedirs(DATA_DIR, exist_ok=True)
        self.smtp_lists.clear()
        
        for filename in os.listdir(DATA_DIR):
            if filename.endswith('.xlsx'):
                self.smtp_lists.addItem(filename)
        
        self.counts_changed.emit()
    
    def filter_list_items(self):
        """Filter list items based on search text."""
        search_text = self.search_bar.text().lower()
        for i in range(self.smtp_lists.count()):
            item = self.smtp_lists.item(i)
            item.setHidden(search_text not in item.text().lower())
    
    def create_new_list(self):
        """Create a new SMTP list."""
        dialog = CreateListDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_list_data()
            list_name = data['name']
            description = data['description']
            
            # Check if list already exists
            list_file = os.path.join(DATA_DIR, list_name)
            if os.path.exists(list_file):
                QMessageBox.warning(self, "List Exists", f"A list named '{list_name}' already exists.")
                return
            
            try:
                # Create Excel file with headers
                wb = Workbook()
                ws = wb.active
                ws.append(["Host", "Port", "Security", "User", "Password", "From Name", "From Email", "Status", "Imported Date", "Last Update"])
                wb.save(list_file)
                
                # Save description
                if description:
                    desc_file = os.path.join(DATA_DIR, list_name.replace('.xlsx', '_description.txt'))
                    with open(desc_file, 'w', encoding='utf-8') as f:
                        f.write(description)
                
                self._load_smtp_files()
                
                # Select the new list
                for i in range(self.smtp_lists.count()):
                    if self.smtp_lists.item(i).text() == list_name:
                        self.smtp_lists.setCurrentRow(i)
                        self.load_smtp_list(self.smtp_lists.item(i))
                        break
                
                QMessageBox.information(self, "Success", f"SMTP list '{list_name}' created successfully.")
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to create list: {str(e)}")
    
    def delete_selected_list(self):
        """Delete the selected SMTP list."""
        current_item = self.smtp_lists.currentItem()
        if not current_item:
            QMessageBox.warning(self, "No Selection", "Please select a list to delete.")
            return
        
        list_name = current_item.text()
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete the SMTP list '{list_name}'?\n\nThis action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Delete main file
                list_file = os.path.join(DATA_DIR, list_name)
                if os.path.exists(list_file):
                    os.remove(list_file)
                
                # Delete description file
                desc_file = os.path.join(DATA_DIR, list_name.replace('.xlsx', '_description.txt'))
                if os.path.exists(desc_file):
                    os.remove(desc_file)
                
                self._load_smtp_files()
                self._clear_table()
                
                QMessageBox.information(self, "Success", f"SMTP list '{list_name}' deleted successfully.")
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete list: {str(e)}")
    
    def load_smtp_list(self, item):
        """Load the selected SMTP list."""
        if not item:
            return
        
        list_name = item.text()
        file_path = os.path.join(DATA_DIR, list_name)
        
        try:
            self.progress_bar.setVisible(True)
            self.status_label.setText("Loading SMTP configurations...")
            
            wb = load_workbook(filename=file_path)
            ws = wb.active
            
            self.smtp_table.setRowCount(0)
            
            # Load data (skip header row)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if not any(row):  # Skip empty rows
                    continue
                
                row_position = self.smtp_table.rowCount()
                self.smtp_table.insertRow(row_position)
                
                for col_index, value in enumerate(row):
                    if col_index < self.smtp_table.columnCount():
                        item = QTableWidgetItem(str(value) if value is not None else "")
                        self.smtp_table.setItem(row_position, col_index, item)
            
            # Load description
            self._load_description(list_name)
            
            # Update UI
            self.table_header.setText(f"📊 {list_name} - SMTP Configurations")
            self.btn_edit_desc.setEnabled(True)
            
            self.progress_bar.setVisible(False)
            self.status_label.setText(f"Loaded {self.smtp_table.rowCount()} SMTP configurations")
            
        except Exception as e:
            self.progress_bar.setVisible(False)
            self.status_label.setText("Error loading file")
            QMessageBox.critical(self, "Load Error", f"Failed to load SMTP list:\n{str(e)}")
            traceback.print_exc()
    
    def _load_description(self, list_name):
        """Load and display list description."""
        desc_file = os.path.join(DATA_DIR, list_name.replace('.xlsx', '_description.txt'))
        
        if os.path.exists(desc_file):
            try:
                with open(desc_file, 'r', encoding='utf-8') as f:
                    description = f.read().strip()
                
                if description:
                    self.description_label.setText(description)
                else:
                    self.description_label.setText("No description available.")
            except Exception as e:
                self.description_label.setText(f"Error loading description: {str(e)}")
        else:
            self.description_label.setText("No description available.")
    
    def _edit_description(self):
        """Edit the description of the current list."""
        current_item = self.smtp_lists.currentItem()
        if not current_item:
            return
        
        list_name = current_item.text()
        desc_file = os.path.join(DATA_DIR, list_name.replace('.xlsx', '_description.txt'))
        current_desc = ""
        
        if os.path.exists(desc_file):
            try:
                with open(desc_file, 'r', encoding='utf-8') as f:
                    current_desc = f.read().strip()
            except:
                pass
        
        new_desc, ok = QInputDialog.getMultiLineText(
            self, "Edit Description", 
            f"Description for '{list_name}':",
            current_desc
        )
        
        if ok:
            try:
                if new_desc.strip():
                    with open(desc_file, 'w', encoding='utf-8') as f:
                        f.write(new_desc.strip())
                else:
                    # Remove description file if empty
                    if os.path.exists(desc_file):
                        os.remove(desc_file)
                
                self._load_description(list_name)
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save description: {str(e)}")
    
    def _clear_table(self):
        """Clear the SMTP table."""
        self.smtp_table.setRowCount(0)
        self.table_header.setText("📊 SMTP Configurations")
        self.description_label.setText("Select a list to view its description.")
        self.btn_edit_desc.setEnabled(False)
    
    # ========== SMTP Management Methods ==========
    
    def add_new_smtp(self):
        """Add a new SMTP configuration."""
        dialog = SMTPDialog(self, title="Add New SMTP")
        if dialog.exec() == QDialog.DialogCode.Accepted:
            smtp_data = dialog.get_smtp_data()
            
            # Validate required fields
            if not all([smtp_data['Host'], smtp_data['Port'], smtp_data['User'], smtp_data['Password']]):
                QMessageBox.warning(self, "Validation Error", "Please fill in all required fields (Host, Port, User, Password).")
                return
            
            # Add to table
            now_str = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
            row_position = self.smtp_table.rowCount()
            self.smtp_table.insertRow(row_position)
            
            columns = ["Host", "Port", "Security", "User", "Password", "From Name", "From Email", "Status", "Imported Date", "Last Update"]
            values = [smtp_data.get(col, "") for col in columns[:7]]
            values.extend(["Untested", now_str, now_str])
            
            for col_index, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                self.smtp_table.setItem(row_position, col_index, item)
            
            self.save_smtp_file()
    
    def edit_selected_smtp(self):
        """Edit the selected SMTP configuration."""
        row = self.smtp_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "No Selection", "Please select an SMTP configuration to edit.")
            return
        
        # Get current data
        smtp_data = {}
        columns = ["Host", "Port", "Security", "User", "Password", "From Name", "From Email"]
        for i, col in enumerate(columns):
            item = self.smtp_table.item(row, i)
            smtp_data[col] = item.text() if item else ""
        
        dialog = SMTPDialog(self, smtp_data, "Edit SMTP")
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_data = dialog.get_smtp_data()
            
            # Update table
            for i, col in enumerate(columns):
                item = QTableWidgetItem(new_data.get(col, ""))
                self.smtp_table.setItem(row, i, item)
            
            # Update last modified
            now_str = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
            self.smtp_table.setItem(row, 9, QTableWidgetItem(now_str))
            
            self.save_smtp_file()
    
    def delete_selected_smtp(self):
        """Delete the selected SMTP configuration."""
        selected_rows = set()
        for item in self.smtp_table.selectedItems():
            selected_rows.add(item.row())
        
        if not selected_rows:
            QMessageBox.warning(self, "No Selection", "Please select SMTP configuration(s) to delete.")
            return
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete {len(selected_rows)} SMTP configuration(s)?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Delete rows in reverse order to maintain indices
            for row in sorted(selected_rows, reverse=True):
                self.smtp_table.removeRow(row)
            
            self.save_smtp_file()
    
    def test_selected_smtp(self):
        """Test the selected SMTP configuration."""
        row = self.smtp_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "No Selection", "Please select an SMTP configuration to test.")
            return
        
        smtp_data = {
            "Host": self.smtp_table.item(row, 0).text() if self.smtp_table.item(row, 0) else "",
            "Port": self.smtp_table.item(row, 1).text() if self.smtp_table.item(row, 1) else "",
            "Security": self.smtp_table.item(row, 2).text() if self.smtp_table.item(row, 2) else "",
            "User": self.smtp_table.item(row, 3).text() if self.smtp_table.item(row, 3) else "",
            "Password": self.smtp_table.item(row, 4).text() if self.smtp_table.item(row, 4) else ""
        }
        
        current_item = self.smtp_lists.currentItem()
        if not current_item:
            QMessageBox.warning(self, "No List Selected", "Please select a list before testing.")
            return
        
        list_name = current_item.text()
        if list_name not in self.thread_pools:
            self.thread_pools[list_name] = QThreadPool()
        
        # Update status to testing
        self.smtp_table.setItem(row, 7, QTableWidgetItem("Testing..."))
        
        task = TestSMTPTask(row, smtp_data, self.on_test_complete)
        self.thread_pools[list_name].start(task)
    
    def test_all_smtp(self):
        """Test all SMTP configurations in the current list."""
        if self.smtp_table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "No SMTP configurations to test.")
            return
        
        current_item = self.smtp_lists.currentItem()
        if not current_item:
            QMessageBox.warning(self, "No List Selected", "Please select a list before testing.")
            return
        
        reply = QMessageBox.question(
            self, "Test All SMTP",
            f"Test all {self.smtp_table.rowCount()} SMTP configurations?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            list_name = current_item.text()
            if list_name not in self.thread_pools:
                self.thread_pools[list_name] = QThreadPool()
            
            # Set max threads
            max_threads = int(self.thread_selector.currentText())
            self.thread_pools[list_name].setMaxThreadCount(max_threads)
            
            # Test all configurations
            for row in range(self.smtp_table.rowCount()):
                smtp_data = {
                    "Host": self.smtp_table.item(row, 0).text() if self.smtp_table.item(row, 0) else "",
                    "Port": self.smtp_table.item(row, 1).text() if self.smtp_table.item(row, 1) else "",
                    "Security": self.smtp_table.item(row, 2).text() if self.smtp_table.item(row, 2) else "",
                    "User": self.smtp_table.item(row, 3).text() if self.smtp_table.item(row, 3) else "",
                    "Password": self.smtp_table.item(row, 4).text() if self.smtp_table.item(row, 4) else ""
                }
                
                # Update status to testing
                self.smtp_table.setItem(row, 7, QTableWidgetItem("Testing..."))
                
                task = TestSMTPTask(row, smtp_data, self.on_test_complete)
                self.thread_pools[list_name].start(task)
            
            self.status_label.setText(f"Testing {self.smtp_table.rowCount()} SMTP configurations...")
    
    def on_test_complete(self, row, result):
        """Handle SMTP test completion."""
        if row < self.smtp_table.rowCount():
            status = "✅ Working" if result else "❌ Failed"
            self.smtp_table.setItem(row, 7, QTableWidgetItem(status))
            
            # Update last tested time
            now_str = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
            self.smtp_table.setItem(row, 9, QTableWidgetItem(now_str))
        
        # Check if all tests are complete
        testing_count = 0
        for i in range(self.smtp_table.rowCount()):
            item = self.smtp_table.item(i, 7)
            if item and item.text() == "Testing...":
                testing_count += 1
        
        if testing_count == 0:
            self.status_label.setText("All SMTP tests completed")
            self.save_smtp_file()  # Auto-save after testing
    
    def update_last_modified(self, row, column):
        """Update the last modified timestamp when a cell is changed."""
        now_str = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
        self.smtp_table.setItem(row, 9, QTableWidgetItem(now_str))
    
    # ========== Import/Export Methods ==========
    
    def import_smtp_file(self):
        """Import SMTP configurations from an Excel file."""
        current_item = self.smtp_lists.currentItem()
        if not current_item:
            QMessageBox.warning(self, "No List Selected", "Please select a list before importing.")
            return
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import SMTP File", DATA_DIR, "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            self.progress_bar.setVisible(True)
            self.status_label.setText("Importing SMTP configurations...")
            
            # Clear existing data
            reply = QMessageBox.question(
                self, "Import Data",
                "Replace existing SMTP configurations with imported data?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.smtp_table.setRowCount(0)
            
            if file_path.endswith('.csv'):
                # Handle CSV import
                import csv
                with open(file_path, 'r', encoding='utf-8', newline='') as csvfile:
                    reader = csv.reader(csvfile)
                    headers = next(reader, [])  # Skip header row
                    
                    now_str = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
                    for row_data in reader:
                        if not any(row_data):
                            continue
                        
                        # Pad row to match table columns
                        padded_row = row_data + [""] * (7 - len(row_data))
                        padded_row = padded_row[:7]  # Take only first 7 columns
                        padded_row.extend(["Imported", now_str, now_str])
                        
                        row_position = self.smtp_table.rowCount()
                        self.smtp_table.insertRow(row_position)
                        for col_index, value in enumerate(padded_row):
                            self.smtp_table.setItem(row_position, col_index, QTableWidgetItem(str(value)))
            else:
                # Handle Excel import
                wb = load_workbook(filename=file_path)
                ws = wb.active
                
                now_str = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    
                    # Pad row to match table columns
                    row_data = list(row) + [""] * (7 - len(row))
                    row_data = row_data[:7]  # Take only first 7 columns
                    row_data.extend(["Imported", now_str, now_str])
                    
                    row_position = self.smtp_table.rowCount()
                    self.smtp_table.insertRow(row_position)
                    for col_index, value in enumerate(row_data):
                        self.smtp_table.setItem(row_position, col_index, QTableWidgetItem(str(value) if value else ""))
            
            self.save_smtp_file()
            
            self.progress_bar.setVisible(False)
            self.status_label.setText(f"Imported {self.smtp_table.rowCount()} SMTP configurations")
            
            QMessageBox.information(self, "Import Complete", f"Successfully imported SMTP configurations.")
            
        except Exception as e:
            self.progress_bar.setVisible(False)
            self.status_label.setText("Import failed")
            QMessageBox.critical(self, "Import Error", f"Failed to import file:\n{str(e)}")
            traceback.print_exc()
    
    def export_smtp_file(self):
        """Export SMTP configurations to an Excel file."""
        if self.smtp_table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "No SMTP configurations to export.")
            return
        
        current_item = self.smtp_lists.currentItem()
        default_name = f"{current_item.text().replace('.xlsx', '')}_export.xlsx" if current_item else "smtp_export.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export SMTP File", default_name, "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)"
        )
        
        if file_path:
            try:
                self.progress_bar.setVisible(True)
                self.status_label.setText("Exporting SMTP configurations...")
                
                if file_path.endswith('.csv'):
                    # Export as CSV
                    import csv
                    with open(file_path, 'w', encoding='utf-8', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        
                        # Write headers
                        headers = [self.smtp_table.horizontalHeaderItem(i).text() for i in range(self.smtp_table.columnCount())]
                        writer.writerow(headers)
                        
                        # Write data
                        for row in range(self.smtp_table.rowCount()):
                            row_data = []
                            for col in range(self.smtp_table.columnCount()):
                                item = self.smtp_table.item(row, col)
                                row_data.append(item.text() if item else "")
                            writer.writerow(row_data)
                else:
                    # Export as Excel
                    wb = Workbook()
                    ws = wb.active
                    
                    # Write headers
                    headers = [self.smtp_table.horizontalHeaderItem(i).text() for i in range(self.smtp_table.columnCount())]
                    ws.append(headers)
                    
                    # Write data
                    for row in range(self.smtp_table.rowCount()):
                        row_data = []
                        for col in range(self.smtp_table.columnCount()):
                            item = self.smtp_table.item(row, col)
                            row_data.append(item.text() if item else "")
                        ws.append(row_data)
                    
                    wb.save(file_path)
                
                self.progress_bar.setVisible(False)
                self.status_label.setText("Export completed")
                
                QMessageBox.information(self, "Export Complete", f"SMTP configurations exported to '{file_path}'")
                
            except Exception as e:
                self.progress_bar.setVisible(False)
                self.status_label.setText("Export failed")
                QMessageBox.critical(self, "Export Error", f"Failed to export file:\n{str(e)}")
    
    def save_smtp_file(self):
        """Save the current SMTP configurations."""
        current_item = self.smtp_lists.currentItem()
        if not current_item:
            return
        
        file_path = os.path.join(DATA_DIR, current_item.text())
        
        try:
            wb = Workbook()
            ws = wb.active
            
            # Write headers
            headers = [self.smtp_table.horizontalHeaderItem(i).text() for i in range(self.smtp_table.columnCount())]
            ws.append(headers)
            
            # Write data
            for row in range(self.smtp_table.rowCount()):
                row_data = []
                for col in range(self.smtp_table.columnCount()):
                    item = self.smtp_table.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)
            
            wb.save(file_path)
            
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save file:\n{str(e)}")
    
    def _auto_save_data(self):
        """Auto-save the current data."""
        self.save_smtp_file()
    
    # ========== Utility Methods ==========
    
    def toggle_scroll_position(self):
        """Toggle horizontal scroll position of the table."""
        scroll_bar = self.smtp_table.horizontalScrollBar()
        if scroll_bar.value() == 0:
            scroll_bar.setValue(scroll_bar.maximum())
        else:
            scroll_bar.setValue(0)
    
    def show_context_menu(self, position):
        """Show context menu for the table."""
        menu = QMenu()
        
        menu.addAction("➕ Add SMTP", self.add_new_smtp)
        
        if self.smtp_table.currentRow() >= 0:
            menu.addAction("✏️ Edit SMTP", self.edit_selected_smtp)
            menu.addAction("🔧 Test Selected", self.test_selected_smtp)
            menu.addAction("🗑 Delete Selected", self.delete_selected_smtp)
        
        menu.addSeparator()
        menu.addAction("🔧 Test All", self.test_all_smtp)
        menu.addSeparator()
        menu.addAction("📥 Import", self.import_smtp_file)
        menu.addAction("📤 Export", self.export_smtp_file)
        
        menu.exec(self.smtp_table.mapToGlobal(position))
    
    def show_list_context_menu(self, position):
        """Show context menu for the lists."""
        item = self.smtp_lists.itemAt(position)
        
        menu = QMenu()
        
        if item:
            menu.addAction("📂 Open", lambda: self.load_smtp_list(item))
            menu.addSeparator()
            menu.addAction("✏️ Edit Description", self._edit_description)
            menu.addSeparator()
            menu.addAction("🗑️ Delete", self.delete_selected_list)
        
        menu.addSeparator()
        menu.addAction("➕ New List", self.create_new_list)
        menu.addAction("🔄 Refresh", self._load_smtp_files)
        
        menu.exec(self.smtp_lists.mapToGlobal(position))
    
    # ========== Public Methods ==========
    
    def get_list_count(self):
        """Get the number of SMTP lists."""
        return self.smtp_lists.count()
    
    def get_selected_list(self):
        """Get the currently selected list name."""
        current_item = self.smtp_lists.currentItem()
        return current_item.text() if current_item else None
    
    def refresh_lists(self):
        """Refresh the lists display."""
        self._load_smtp_files()
    
    def apply_theme(self):
        """Apply theme from assets if available."""
        qss_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../assets/themes/Default.qss'))
        if os.path.exists(qss_path):
            try:
                with open(qss_path, 'r', encoding='utf-8') as f:
                    self.setStyleSheet(f.read())
            except Exception as e:
                print(f"Failed to load theme: {e}")