# workers/data_operation_worker.py
"""
Unified background worker for all data operations (import, export, save, load)
Prevents UI freezing during large file operations
"""

from PyQt6.QtCore import QThread, pyqtSignal, QMutex, QWaitCondition
import os
import traceback
from openpyxl import load_workbook, Workbook
import csv

class DataOperationWorker(QThread):
    """Universal worker thread for all data operations across managers"""
    
    # Signals for different operation types
    progress_updated = pyqtSignal(int, int, str)  # current, total, operation_name
    operation_completed = pyqtSignal(dict)  # result data
    error_occurred = pyqtSignal(str, str)  # error_message, operation_type
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.operation_type = None
        self.file_path = None
        self.data_to_save = None
        self.save_path = None
        self.operation_params = {}
        self.stop_requested = False
        self.mutex = QMutex()
        self.wait_condition = QWaitCondition()
    
    def set_operation(self, operation_type: str, **kwargs):
        """Configure the operation to perform"""
        self.mutex.lock()
        try:
            self.operation_type = operation_type
            self.operation_params = kwargs
            self.stop_requested = False
            
            # Extract common parameters
            self.file_path = kwargs.get('file_path')
            self.data_to_save = kwargs.get('data_to_save')
            self.save_path = kwargs.get('save_path')
        finally:
            self.mutex.unlock()
    
    def request_stop(self):
        """Request the operation to stop"""
        self.mutex.lock()
        try:
            self.stop_requested = True
            self.wait_condition.wakeAll()
        finally:
            self.mutex.unlock()
    
    def run(self):
        """Main thread execution"""
        try:
            if self.operation_type == "import_excel":
                self._import_excel()
            elif self.operation_type == "import_csv":
                self._import_csv()
            elif self.operation_type == "import_txt":
                self._import_txt()
            elif self.operation_type == "save_excel":
                self._save_excel()
            elif self.operation_type == "save_csv":
                self._save_csv()
            elif self.operation_type == "save_txt":
                self._save_txt()
            elif self.operation_type == "load_data":
                self._load_data()
            else:
                self.error_occurred.emit(f"Unknown operation type: {self.operation_type}", self.operation_type)
                
        except Exception as e:
            error_msg = f"Operation failed: {str(e)}\n{traceback.format_exc()}"
            self.error_occurred.emit(error_msg, self.operation_type or "unknown")
    
    def _check_stop_requested(self):
        """Check if stop was requested"""
        self.mutex.lock()
        try:
            return self.stop_requested
        finally:
            self.mutex.unlock()
    
    def _import_excel(self):
        """Import data from Excel file"""
        file_path = self.file_path
        manager_type = self.operation_params.get('manager_type', 'generic')
        
        if not os.path.exists(file_path):
            self.error_occurred.emit(f"File not found: {file_path}", "import_excel")
            return
        
        try:
            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook.active
            
            # Get total rows for progress
            total_rows = worksheet.max_row
            self.progress_updated.emit(0, total_rows, f"Reading {manager_type} Excel file")
            
            headers = []
            data = []
            
            # Read headers
            if total_rows > 0:
                headers = [cell.value or f"Column_{i+1}" for i, cell in enumerate(worksheet[1])]
            
            # Read data rows
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 1):
                if self._check_stop_requested():
                    break
                
                # Convert row to list, handling None values
                row_data = [cell if cell is not None else "" for cell in row]
                data.append(row_data)
                
                # Update progress every 100 rows
                if row_idx % 100 == 0:
                    self.progress_updated.emit(row_idx, total_rows-1, f"Reading row {row_idx}")
            
            workbook.close()
            
            # Emit completion with data
            result = {
                'headers': headers,
                'data': data,
                'total_rows': len(data),
                'file_path': file_path,
                'manager_type': manager_type
            }
            self.operation_completed.emit(result)
            
        except Exception as e:
            self.error_occurred.emit(f"Excel import failed: {str(e)}", "import_excel")
    
    def _import_csv(self):
        """Import data from CSV file"""
        file_path = self.file_path
        manager_type = self.operation_params.get('manager_type', 'generic')
        
        try:
            data = []
            headers = []
            
            with open(file_path, 'r', encoding='utf-8', newline='') as csvfile:
                # Detect dialect
                sample = csvfile.read(1024)
                csvfile.seek(0)
                dialect = csv.Sniffer().sniff(sample)
                
                reader = csv.reader(csvfile, dialect)
                
                # Read headers
                headers = next(reader, [])
                
                # Count total lines for progress
                csvfile.seek(0)
                total_lines = sum(1 for line in csvfile) - 1  # Subtract header
                csvfile.seek(0)
                reader = csv.reader(csvfile, dialect)
                next(reader)  # Skip header
                
                self.progress_updated.emit(0, total_lines, f"Reading {manager_type} CSV file")
                
                # Read data
                for row_idx, row in enumerate(reader):
                    if self._check_stop_requested():
                        break
                    
                    data.append(row)
                    
                    if row_idx % 100 == 0:
                        self.progress_updated.emit(row_idx, total_lines, f"Reading row {row_idx}")
            
            result = {
                'headers': headers,
                'data': data,
                'total_rows': len(data),
                'file_path': file_path,
                'manager_type': manager_type
            }
            self.operation_completed.emit(result)
            
        except Exception as e:
            self.error_occurred.emit(f"CSV import failed: {str(e)}", "import_csv")
    
    def _import_txt(self):
        """Import data from text file (for subjects, messages)"""
        file_path = self.file_path
        manager_type = self.operation_params.get('manager_type', 'generic')
        
        try:
            data = []
            
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            total_lines = len(lines)
            self.progress_updated.emit(0, total_lines, f"Reading {manager_type} text file")
            
            for idx, line in enumerate(lines):
                if self._check_stop_requested():
                    break
                
                line = line.strip()
                if line:  # Skip empty lines
                    data.append(line)
                
                if idx % 100 == 0:
                    self.progress_updated.emit(idx, total_lines, f"Reading line {idx}")
            
            result = {
                'data': data,
                'total_rows': len(data),
                'file_path': file_path,
                'manager_type': manager_type
            }
            self.operation_completed.emit(result)
            
        except Exception as e:
            self.error_occurred.emit(f"Text import failed: {str(e)}", "import_txt")
    
    def _save_excel(self):
        """Save data to Excel file"""
        data = self.data_to_save
        save_path = self.save_path
        headers = self.operation_params.get('headers', [])
        
        try:
            workbook = Workbook()
            worksheet = workbook.active
            
            total_rows = len(data)
            self.progress_updated.emit(0, total_rows, "Saving Excel file")
            
            # Write headers
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col_idx, value=header)
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):  # Start from row 2 (after headers)
                if self._check_stop_requested():
                    break
                
                for col_idx, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                
                if row_idx % 100 == 0:
                    self.progress_updated.emit(row_idx-1, total_rows, f"Saving row {row_idx-1}")
            
            workbook.save(save_path)
            
            result = {
                'success': True,
                'file_path': save_path,
                'rows_saved': len(data)
            }
            self.operation_completed.emit(result)
            
        except Exception as e:
            self.error_occurred.emit(f"Excel save failed: {str(e)}", "save_excel")
    
    def _save_txt(self):
        """Save data to text file (for subjects, messages)"""
        data = self.data_to_save
        save_path = self.save_path
        
        try:
            total_lines = len(data)
            self.progress_updated.emit(0, total_lines, "Saving text file")
            
            with open(save_path, 'w', encoding='utf-8') as f:
                for idx, line in enumerate(data):
                    if self._check_stop_requested():
                        break
                    
                    f.write(str(line) + '\n')
                    
                    if idx % 100 == 0:
                        self.progress_updated.emit(idx, total_lines, f"Saving line {idx}")
            
            result = {
                'success': True,
                'file_path': save_path,
                'lines_saved': len(data)
            }
            self.operation_completed.emit(result)
            
        except Exception as e:
            self.error_occurred.emit(f"Text save failed: {str(e)}", "save_txt")
    
    def _load_data(self):
        """Load existing data from file"""
        file_path = self.file_path
        manager_type = self.operation_params.get('manager_type', 'generic')
        
        try:
            if file_path.endswith('.xlsx'):
                self._import_excel()
            elif file_path.endswith('.csv'):
                self._import_csv()
            elif file_path.endswith('.txt'):
                self._import_txt()
            else:
                self.error_occurred.emit(f"Unsupported file type: {file_path}", "load_data")
        except Exception as e:
            self.error_occurred.emit(f"Load data failed: {str(e)}", "load_data")