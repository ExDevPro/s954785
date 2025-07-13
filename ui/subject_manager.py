import os
import shutil
import csv
from openpyxl import load_workbook, Workbook
import re
import traceback # Import traceback for detailed error logging

from PyQt6.QtWidgets import (
    QWidget, QLabel, QListWidget, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit,
    QHBoxLayout, QVBoxLayout,
    QFileDialog, QMessageBox, QInputDialog, QMenu,
    QProgressBar, QAbstractItemView, QApplication
)
from PyQt6.QtGui import QAction, QCursor
# *** REMOVED: qRegisterMetaType and QMetaType imports ***
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QObject

BASE_PATH = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
DATA_DIR = os.path.join(BASE_PATH, 'data', 'subjects')
CHUNK_SIZE = 100 # Number of subjects per page

# --- Helper Function to Count Lines ---
def count_lines_in_file(filepath):
    """Counts non-empty lines in a text file."""
    count = 0
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            count = sum(1 for line in f if line.strip())
    except Exception:
        pass # Ignore errors during counting for the header
    return count

# --- Background Thread for Importing ---

# *** REMOVED: SubjectImportResult class definition ***

class SubjectDataThread(QThread):
    """Worker thread for loading/importing subject data."""
    # *** MODIFIED: Signal now emits a dictionary ***
    import_finished = pyqtSignal(dict) # Emits {'subjects': [], 'error': str|None, 'original_filename': str}

    def __init__(self, file_path, parent=None):
        super().__init__(parent)
        self.file_path = file_path

    def run(self):
        """Reads subjects from TXT, CSV, or XLSX."""
        subjects = []
        error_msg = None
        original_filename = os.path.basename(self.file_path)
        result_dict = {} # Dictionary to hold results

        try:
            print(f"Import Thread: Starting import for {self.file_path}")
            ext = os.path.splitext(self.file_path)[1].lower()

            if ext == '.txt':
                with open(self.file_path, 'r', encoding='utf-8') as f:
                    subjects = [line.strip() for line in f if line.strip()]
            elif ext == '.csv':
                with open(self.file_path, 'r', encoding='utf-8', newline='') as f:
                    reader = csv.reader(f)
                    try:
                        first_row = next(reader)
                        header_likely = first_row and first_row[0].strip()
                        remaining_subjects = [row[0].strip() for row in reader if row and len(row) > 0 and row[0].strip()]
                        if not header_likely:
                            if first_row and len(first_row) > 0 and first_row[0].strip():
                                subjects = [first_row[0].strip()] + remaining_subjects
                            else:
                                subjects = remaining_subjects
                        else:
                            subjects = remaining_subjects
                    except StopIteration:
                        pass # Empty file
            elif ext == '.xlsx':
                wb = load_workbook(filename=self.file_path, read_only=True)
                ws = wb.active
                subjects = [str(row[0].value).strip() for row in ws.iter_rows(min_row=2, max_col=1) if row[0].value and str(row[0].value).strip()]
                wb.close()
            else:
                error_msg = f"Unsupported file type: {ext}"

            if not error_msg:
                 print(f"Import Thread: Successfully read {len(subjects)} subjects from {original_filename}")

        except FileNotFoundError:
            error_msg = f"File not found: {self.file_path}"
        except Exception as e:
            error_msg = f"Error importing file {original_filename}:\n{traceback.format_exc()}"
            print(f"Import Thread: Error - {error_msg}")

        # *** MODIFIED: Populate dictionary instead of custom object ***
        result_dict['subjects'] = subjects
        result_dict['error'] = error_msg
        result_dict['original_filename'] = original_filename

        self.import_finished.emit(result_dict) # Emit the dictionary
        print(f"Import Thread: Finished for {original_filename}. Emitting signal.")


# --- Subject Manager UI ---
class SubjectManager(QWidget):
    counts_changed = pyqtSignal(int, int)

    def __init__(self, parent=None):
        super().__init__(parent)
        os.makedirs(DATA_DIR, exist_ok=True)
        self.current_list_path = None
        self.all_data = []
        self.current_offset = 0
        self.total_rows = 0
        self.import_thread = None

        self._build_ui()
        self._refresh_list()

    def _build_ui(self):
        layout = QHBoxLayout(self)

        # --- Left Pane ---
        left_layout = QVBoxLayout()
        left_layout.setSpacing(5)
        self.header_label = QLabel("0 lists – 0 subjects")
        self.header_label.setObjectName("liveHeaderLabel")
        self.header_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        left_layout.addWidget(self.header_label)
        list_button_layout = QHBoxLayout()
        btn_new = QPushButton("➕ New List")
        btn_new.setToolTip("Create a new empty subject list file")
        btn_new.clicked.connect(self._new_list)
        btn_del = QPushButton("🗑 Delete")
        btn_del.setToolTip("Delete the selected subject list file")
        btn_del.clicked.connect(self._delete_list)
        list_button_layout.addWidget(btn_new)
        list_button_layout.addWidget(btn_del)
        left_layout.addLayout(list_button_layout)
        self.list_widget = QListWidget()
        self.list_widget.setObjectName("subjectListWidget")
        self.list_widget.currentTextChanged.connect(self._on_list_selection_changed)
        left_layout.addWidget(self.list_widget)

        # --- Right Pane ---
        right_layout = QVBoxLayout()
        right_layout.setSpacing(5)
        action_layout = QHBoxLayout()
        btn_import = QPushButton("⬆️ Import")
        btn_import.setToolTip("Import subjects from TXT, CSV, or XLSX")
        btn_import.clicked.connect(self._import)
        btn_save = QPushButton("💾 Save")
        btn_save.setToolTip("Save changes to the current list")
        btn_save.clicked.connect(self._save)
        self.btn_save = btn_save
        self.btn_save.setEnabled(False)
        btn_dedupe = QPushButton("✨ De-dupe")
        btn_dedupe.setToolTip("Remove duplicate subjects from the current list")
        btn_dedupe.clicked.connect(self._remove_duplicates)
        btn_nav_prev = QPushButton("◀ Prev")
        btn_nav_prev.setToolTip("Previous page")
        btn_nav_prev.clicked.connect(self._prev_chunk)
        self.btn_prev = btn_nav_prev
        self.btn_prev.setEnabled(False)
        btn_nav_next = QPushButton("Next ▶")
        btn_nav_next.setToolTip("Next page")
        btn_nav_next.clicked.connect(self._next_chunk)
        self.btn_next = btn_nav_next
        self.btn_next.setEnabled(False)
        action_layout.addWidget(btn_import)
        action_layout.addWidget(btn_save)
        action_layout.addWidget(btn_dedupe)
        action_layout.addStretch(1)
        action_layout.addWidget(btn_nav_prev)
        action_layout.addWidget(btn_nav_next)
        right_layout.addLayout(action_layout)
        search_status_layout = QHBoxLayout()
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("🔍 Type to filter subjects...")
        self.filter_input.textChanged.connect(self._filter_table)
        search_status_layout.addWidget(self.filter_input)
        self.status_label = QLabel("Rows 0-0 of 0")
        self.status_label.setObjectName("statusBarLabel")
        search_status_layout.addWidget(self.status_label)
        right_layout.addLayout(search_status_layout)
        self.table = QTableWidget()
        self.table.setColumnCount(1)
        self.table.setHorizontalHeaderLabels(["Subject"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked | QAbstractItemView.EditTrigger.SelectedClicked | QAbstractItemView.EditTrigger.AnyKeyPressed)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.itemChanged.connect(self._on_item_changed)
        right_layout.addWidget(self.table)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setMaximum(0)
        self.progress_bar.setMinimum(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("Importing...")
        right_layout.addWidget(self.progress_bar)
        layout.addLayout(left_layout, 1)
        layout.addLayout(right_layout, 3)

    # --- List Management Methods --- (No changes needed in these methods)
    def _update_header_counts(self):
        list_files = []
        total_subject_count = 0
        try:
             list_files = [f for f in os.listdir(DATA_DIR) if f.lower().endswith('.txt') and os.path.isfile(os.path.join(DATA_DIR, f))]
             list_count = len(list_files)
             for filename in list_files:
                 filepath = os.path.join(DATA_DIR, filename)
                 total_subject_count += count_lines_in_file(filepath)
        except Exception as e:
             print(f"Error calculating header counts: {e}")
             list_count = 0
             total_subject_count = 0
        self.header_label.setText(f"{list_count} lists – {total_subject_count} subjects")
        self.counts_changed.emit(list_count, total_subject_count)

    def _refresh_list(self):
        self.list_widget.blockSignals(True)
        current_selection_name = self.list_widget.currentItem().text() if self.list_widget.currentItem() else None
        self.list_widget.clear()
        found_selection = False
        try:
            if os.path.isdir(DATA_DIR):
                for fname in sorted(os.listdir(DATA_DIR)):
                    if fname.lower().endswith('.txt') and os.path.isfile(os.path.join(DATA_DIR, fname)):
                        list_name = os.path.splitext(fname)[0]
                        self.list_widget.addItem(list_name)
                        if list_name == current_selection_name:
                            found_selection = True
        except Exception as e:
             QMessageBox.critical(self, "Error Refreshing Lists", f"Could not read subject directory:\n{e}")
        if found_selection and current_selection_name:
             items = self.list_widget.findItems(current_selection_name, Qt.MatchFlag.MatchExactly)
             if items:
                 if self.list_widget.currentItem() != items[0]:
                     self.list_widget.setCurrentItem(items[0])
        self.list_widget.blockSignals(False)
        self._update_header_counts()

    def _new_list(self):
        name, ok = QInputDialog.getText(self, "New Subject List", "Enter list name:")
        if ok and name and name.strip():
            clean_name = re.sub(r'[<>:"/\\|?*]', '_', name.strip())
            if not clean_name:
                 QMessageBox.warning(self, "Invalid Name", "Please enter a valid list name.")
                 return
            path = os.path.join(DATA_DIR, f"{clean_name}.txt")
            if os.path.exists(path):
                QMessageBox.warning(self, "Exists", f"List '{clean_name}' already exists.")
            else:
                try:
                    with open(path, 'w', encoding='utf-8') as f: f.write("")
                    print(f"Created new list file: {path}")
                    self._refresh_list()
                    items = self.list_widget.findItems(clean_name, Qt.MatchFlag.MatchExactly)
                    if items: self.list_widget.setCurrentItem(items[0])
                except Exception as e: QMessageBox.critical(self, "Error", f"Could not create list file:\n{e}")

    def _delete_list(self):
        item = self.list_widget.currentItem()
        if not item: QMessageBox.warning(self, "No Selection", "Please select a list to delete."); return
        name = item.text()
        path = os.path.join(DATA_DIR, f"{name}.txt")
        if QMessageBox.question(self, "Confirm Delete", f"Are you sure you want to permanently delete the subject list '{name}'?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            try:
                if os.path.exists(path):
                     os.remove(path)
                     print(f"Deleted list file: {path}")
                     if self.current_list_path == path:
                          self.table.setRowCount(0); self.all_data = []; self.current_list_path = None
                          self.current_offset = 0; self.total_rows = 0; self.btn_save.setEnabled(False)
                          self.btn_prev.setEnabled(False); self.btn_next.setEnabled(False)
                          self.filter_input.clear(); self.status_label.setText("Rows 0-0 of 0")
                     self._refresh_list()
                else: QMessageBox.warning(self, "Not Found", f"List file '{name}.txt' not found."); self._refresh_list()
            except Exception as e: QMessageBox.critical(self, "Error Deleting", f"Could not delete list file:\n{e}")

    def _on_list_selection_changed(self, list_name: str):
        if not list_name:
            self.table.setRowCount(0); self.all_data = []; self.current_list_path = None
            self.btn_save.setEnabled(False); self.status_label.setText("Rows 0-0 of 0")
            self.btn_prev.setEnabled(False); self.btn_next.setEnabled(False); print("List selection cleared.")
            return
        new_path = os.path.join(DATA_DIR, f"{list_name}.txt")
        if new_path == self.current_list_path: return
        print(f"Selection changed to: {list_name}")
        self._load_list(list_name)

    # --- Data Loading and Display Methods --- (No changes needed)
    def _load_list(self, list_name: str):
        path = os.path.join(DATA_DIR, f"{list_name}.txt")
        if not os.path.exists(path):
            QMessageBox.warning(self, "Load Error", f"File not found: {path}")
            self.current_list_path = None; self.all_data = []; self.table.setRowCount(0)
            self.btn_save.setEnabled(False); self.status_label.setText("Rows 0-0 of 0")
            self._refresh_list(); return
        self.current_list_path = path; self.current_offset = 0; self.all_data = []
        try:
            with open(path, 'r', encoding='utf-8') as f: self.all_data = [line.strip() for line in f if line.strip()]
            self.total_rows = len(self.all_data); print(f"Loaded {self.total_rows} subjects from {list_name}.txt")
            self._render_page(); self.btn_save.setEnabled(False); self.filter_input.clear()
        except Exception as e:
            QMessageBox.critical(self, "Error Loading List", f"Could not load list from file:\n{path}\n\n{e}")
            self.current_list_path = None; self.all_data = []; self.total_rows = 0
            self.table.setRowCount(0); self.btn_save.setEnabled(False); self.status_label.setText("Rows 0-0 of 0")

    def _render_page(self):
        self.table.setRowCount(0); self.table.blockSignals(True)
        start_index = self.current_offset; end_index = min(start_index + CHUNK_SIZE, self.total_rows)
        page_data = self.all_data[start_index:end_index]; self.table.setRowCount(len(page_data))
        for r_idx, subject in enumerate(page_data): self.table.setItem(r_idx, 0, QTableWidgetItem(subject))
        self.table.blockSignals(False); start_display = start_index + 1 if self.total_rows > 0 else 0
        self.status_label.setText(f"Rows {start_display}–{end_index} of {self.total_rows}")
        self.btn_prev.setEnabled(self.current_offset > 0); self.btn_next.setEnabled(end_index < self.total_rows)
        if self.filter_input.text(): self._filter_table(self.filter_input.text())
        else:
             for r in range(self.table.rowCount()): self.table.setRowHidden(r, False)

    def _next_chunk(self):
        if self.current_offset + CHUNK_SIZE < self.total_rows: self.current_offset += CHUNK_SIZE; self._render_page()
    def _prev_chunk(self):
        if self.current_offset > 0: self.current_offset = max(0, self.current_offset - CHUNK_SIZE); self._render_page()
    def _filter_table(self, text: str):
        filter_text = text.lower().strip(); self.table.blockSignals(True); visible_count = 0
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0); row_matches = False
            if not filter_text: row_matches = True
            elif item and item.text():
                if filter_text in item.text().lower(): row_matches = True
            self.table.setRowHidden(r, not row_matches)
            if row_matches: visible_count += 1
        self.table.blockSignals(False)

    # --- Action Methods --- (No changes needed for _on_item_changed, _save, _remove_duplicates, _import)
    def _on_item_changed(self, item: QTableWidgetItem):
        if item and self.table.signalsBlocked() is False:
             data_idx = self.current_offset + item.row()
             if 0 <= data_idx < self.total_rows:
                  new_text = item.text().strip()
                  if self.all_data[data_idx] != new_text:
                      self.all_data[data_idx] = new_text; self.btn_save.setEnabled(True)
                      print(f"Item changed at index {data_idx}: {self.all_data[data_idx]}")

    def _save(self):
        if not self.current_list_path: QMessageBox.warning(self, "No List Loaded", "Please load a subject list first."); return
        if not self.btn_save.isEnabled(): QMessageBox.information(self, "No Changes", "No changes to save."); return
        try:
            print(f"Saving {len(self.all_data)} subjects to {self.current_list_path}")
            subjects_to_save = [s for s in self.all_data if s]
            with open(self.current_list_path, 'w', encoding='utf-8') as f:
                for subject in subjects_to_save: f.write(subject + "\n")
            self.total_rows = len(subjects_to_save); self.all_data = subjects_to_save
            QMessageBox.information(self, "Saved", "Subject list saved successfully.")
            self.btn_save.setEnabled(False); self._update_header_counts(); self._render_page()
        except Exception as e: QMessageBox.critical(self, "Error Saving", f"Could not save list:\n{e}")

    def _remove_duplicates(self):
        if not self.current_list_path: QMessageBox.warning(self, "No List Loaded", "Load a list to remove duplicates."); return
        initial_count = len(self.all_data)
        if initial_count == 0: QMessageBox.information(self, "Empty List", "The list is empty, nothing to de-duplicate."); return
        seen = set(); unique_subjects = []
        for subject in self.all_data:
            lower_subject = subject.lower()
            if lower_subject not in seen: seen.add(lower_subject); unique_subjects.append(subject)
        duplicates_removed = initial_count - len(unique_subjects)
        if duplicates_removed > 0:
            print(f"Removed {duplicates_removed} duplicate subjects.")
            self.all_data = unique_subjects; self.total_rows = len(self.all_data)
            self._render_page(); self.btn_save.setEnabled(True)
            QMessageBox.information(self, "Duplicates Removed", f"{duplicates_removed} duplicate subject(s) removed.\nRemember to save the changes.")
        else: QMessageBox.information(self, "No Duplicates", "No duplicate subjects found.")

    def _import(self):
        if self.import_thread and self.import_thread.isRunning(): QMessageBox.warning(self, "Import Running", "An import operation is already in progress."); return
        # Check if a list is selected
        if not self.current_list_path:
            QMessageBox.warning(self, "No List Selected", "Please create or select a subject list first before importing.")
            return
        file_path, _ = QFileDialog.getOpenFileName(self, "Import Subjects", "", "Subject Files (*.txt *.csv *.xlsx);;Text Files (*.txt);;CSV Files (*.csv);;Excel Files (*.xlsx)")
        if not file_path: return
        self.progress_bar.setVisible(True); self.progress_bar.setFormat("Importing..."); self.progress_bar.setMaximum(0)
        print(f"Starting import thread for: {file_path}")
        self.import_thread = SubjectDataThread(file_path, self)
        self.import_thread.import_finished.connect(self._on_import_finished)
        self.import_thread.start()


    # *** MODIFIED: Slot signature and data access ***
    def _on_import_finished(self, result_dict: dict):
        """Handles the results from the import thread."""
        print("Import thread finished. Processing results...")
        self.progress_bar.setVisible(False)
        self.import_thread = None

        # Access data using dictionary keys
        error = result_dict.get('error')
        imported_subjects = result_dict.get('subjects', [])
        original_filename = result_dict.get('original_filename', 'Unknown File')

        if error:
            QMessageBox.critical(self, "Import Error", error)
            return

        if not imported_subjects:
             QMessageBox.information(self, "Import Complete", "No subjects found in the selected file.")
             return

        print(f"Received {len(imported_subjects)} subjects from import.")

        # --- Logic for Merging or Creating New List --- (Same logic, uses dict)
        if self.current_list_path and os.path.exists(self.current_list_path):
            reply = QMessageBox.question(self, "Merge Subjects", f"Add {len(imported_subjects)} subjects from '{original_filename}' to the current list '{os.path.basename(self.current_list_path)}'?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.Yes)
            if reply == QMessageBox.StandardButton.Yes:
                current_lower = {s.lower() for s in self.all_data}; added_count = 0
                for subj in imported_subjects:
                     if subj.lower() not in current_lower: self.all_data.append(subj); added_count += 1
                if added_count > 0:
                     self.total_rows = len(self.all_data); print(f"Merged {added_count} new subjects into the list.")
                     self._render_page(); self.btn_save.setEnabled(True)
                     QMessageBox.information(self, "Merge Complete", f"{added_count} new subjects added.\nRemember to save the changes.")
                else: QMessageBox.information(self, "Merge Complete", "No new subjects were added (they might already exist).")
            else: print("Merge cancelled by user.")
        else:
            base_name = os.path.splitext(original_filename)[0]
            new_list_name = re.sub(r'[<>:"/\\|?*]', '_', base_name.strip())
            if not new_list_name: new_list_name = "Imported_Subjects"
            counter = 1; final_name = new_list_name
            while os.path.exists(os.path.join(DATA_DIR, f"{final_name}.txt")): final_name = f"{new_list_name}_{counter}"; counter += 1
            new_list_path = os.path.join(DATA_DIR, f"{final_name}.txt")
            try:
                with open(new_list_path, 'w', encoding='utf-8') as f:
                    for subject in imported_subjects: f.write(subject + "\n")
                print(f"Created and saved new list: {new_list_path}")
                self._refresh_list()
                items = self.list_widget.findItems(final_name, Qt.MatchFlag.MatchExactly)
                if items: self.list_widget.setCurrentItem(items[0])
                else: print(f"WARN: Could not find newly created list '{final_name}' in widget, loading manually."); self._load_list(final_name)
                QMessageBox.information(self, "Import Successful", f"Created new list '{final_name}' with {len(imported_subjects)} subjects.")
            except Exception as e: QMessageBox.critical(self, "Error Creating List", f"Could not create or save new list file '{final_name}.txt':\n{e}")

    # --- Context Menu Methods --- (No changes needed)
    def _show_context_menu(self, pos):
        menu = QMenu(self); selected_rows = self.table.selectionModel().selectedRows(); num_selected = len(selected_rows)
        action_delete = QAction(f"Delete {num_selected} Row(s)" if num_selected > 1 else "Delete Row", self)
        action_delete.setEnabled(num_selected > 0); action_delete.triggered.connect(self._delete_selected_rows); menu.addAction(action_delete)
        action_copy = QAction("Copy Subject(s)", self); action_copy.setEnabled(num_selected > 0)
        action_copy.triggered.connect(self._copy_selected_rows); menu.addAction(action_copy); menu.exec(self.table.mapToGlobal(pos))
    def _delete_selected_rows(self):
        selected_indices = sorted([index.row() for index in self.table.selectionModel().selectedRows()], reverse=True);
        if not selected_indices: return; print(f"Deleting rows (view indices): {selected_indices}")
        indices_to_remove_data = sorted([self.current_offset + view_idx for view_idx in selected_indices], reverse=True); removed_count = 0
        for data_idx in indices_to_remove_data:
            if 0 <= data_idx < len(self.all_data): del self.all_data[data_idx]; removed_count += 1
        if removed_count > 0:
            self.total_rows -= removed_count; print(f"Removed {removed_count} items from self.all_data. New total: {self.total_rows}")
            self.btn_save.setEnabled(True); self._render_page()
    def _copy_selected_rows(self):
        selected_indices = sorted([index.row() for index in self.table.selectionModel().selectedRows()])
        if not selected_indices: return; subjects_to_copy = []
        for view_idx in selected_indices:
             item = self.table.item(view_idx, 0);
             if item and item.text(): subjects_to_copy.append(item.text())
        if subjects_to_copy: QApplication.clipboard().setText("\n".join(subjects_to_copy)); print(f"Copied {len(subjects_to_copy)} subjects to clipboard.")

    # --- Cleanup --- (No changes needed)
    def closeEvent(self, event):
        if self.import_thread and self.import_thread.isRunning():
            print("Attempting to stop import thread on close...")
            self.import_thread.terminate(); self.import_thread.wait(1000); print("Import thread stopped.")
        event.accept()