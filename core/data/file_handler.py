"""
File handler for Bulk Email Sender.

This module provides file operations including:
- Safe file reading/writing
- CSV processing
- Excel file handling
- Import/export functionality
"""

import os
import csv
import json
import time
from typing import List, Dict, Any, Optional, Union, Iterator
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

from core.utils.logger import get_module_logger
from core.utils.exceptions import DataFileError, ValidationError
from core.utils.helpers import get_safe_filename, ensure_directory, safe_json_load, safe_json_save
from core.security.path_sanitizer import validate_attachment_path, validate_data_directory
from core.validation.data_validator import DataValidator

logger = get_module_logger(__name__)


class FileHandler:
    """Handles file operations for the application."""
    
    def __init__(self, base_data_dir: Optional[str] = None):
        """
        Initialize file handler.
        
        Args:
            base_data_dir: Base directory for data files
        """
        self.base_data_dir = base_data_dir or "data"
        self.validator = DataValidator()
        
        # Ensure base directory exists
        ensure_directory(self.base_data_dir)
    
    def read_csv_file(self, filepath: str, encoding: str = 'utf-8') -> Iterator[Dict[str, Any]]:
        """
        Read CSV file and yield rows as dictionaries.
        
        Args:
            filepath: Path to CSV file
            encoding: File encoding
            
        Yields:
            Dictionary representing each row
            
        Raises:
            DataFileError: If file cannot be read
        """
        try:
            validated_path = validate_attachment_path(filepath)
            
            with open(validated_path, 'r', newline='', encoding=encoding) as csvfile:
                # Detect dialect
                sample = csvfile.read(1024)
                csvfile.seek(0)
                
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel
                
                reader = csv.DictReader(csvfile, dialect=dialect)
                
                for row_num, row in enumerate(reader, start=1):
                    # Clean up row data
                    cleaned_row = {}
                    for key, value in row.items():
                        if key is not None:
                            clean_key = str(key).strip()
                            clean_value = str(value).strip() if value is not None else ""
                            if clean_key:
                                cleaned_row[clean_key] = clean_value
                    
                    if cleaned_row:  # Only yield non-empty rows
                        cleaned_row['_row_number'] = row_num
                        yield cleaned_row
        
        except UnicodeDecodeError as e:
            raise DataFileError(f"File encoding error: {e}", filepath=filepath)
        except Exception as e:
            raise DataFileError(f"Failed to read CSV file: {e}", filepath=filepath)
    
    def write_csv_file(self, filepath: str, data: List[Dict[str, Any]], 
                      fieldnames: Optional[List[str]] = None) -> bool:
        """
        Write data to CSV file.
        
        Args:
            filepath: Output file path
            data: List of dictionaries to write
            fieldnames: Optional list of field names for ordering
            
        Returns:
            True if successful
            
        Raises:
            DataFileError: If file cannot be written
        """
        try:
            if not data:
                raise DataFileError("No data to write", filepath=filepath)
            
            # Ensure output directory exists
            ensure_directory(os.path.dirname(filepath))
            
            # Determine fieldnames if not provided
            if fieldnames is None:
                fieldnames = list(data[0].keys())
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for row in data:
                    # Filter row to only include specified fieldnames
                    filtered_row = {field: row.get(field, '') for field in fieldnames}
                    writer.writerow(filtered_row)
            
            logger.info(f"CSV file written successfully: {filepath} ({len(data)} rows)")
            return True
        
        except Exception as e:
            raise DataFileError(f"Failed to write CSV file: {e}", filepath=filepath)
    
    def read_excel_file(self, filepath: str, sheet_name: Optional[str] = None) -> Iterator[Dict[str, Any]]:
        """
        Read Excel file and yield rows as dictionaries.
        
        Args:
            filepath: Path to Excel file
            sheet_name: Name of sheet to read (first sheet if None)
            
        Yields:
            Dictionary representing each row
            
        Raises:
            DataFileError: If file cannot be read
        """
        try:
            validated_path = validate_attachment_path(filepath)
            
            workbook = openpyxl.load_workbook(validated_path, read_only=True)
            
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise DataFileError(f"Sheet '{sheet_name}' not found in workbook", filepath=filepath)
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # Get header row
            headers = []
            first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not first_row:
                raise DataFileError("Excel file is empty", filepath=filepath)
            
            for cell in first_row:
                header = str(cell).strip() if cell is not None else ""
                headers.append(header)
            
            # Read data rows
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {}
                
                for i, cell_value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        value = str(cell_value).strip() if cell_value is not None else ""
                        row_data[headers[i]] = value
                
                # Only yield rows with data
                if any(value for value in row_data.values()):
                    row_data['_row_number'] = row_num
                    yield row_data
            
            workbook.close()
        
        except FileNotFoundError:
            raise DataFileError(f"Excel file not found: {filepath}", filepath=filepath)
        except Exception as e:
            raise DataFileError(f"Failed to read Excel file: {e}", filepath=filepath)
    
    def write_excel_file(self, filepath: str, data: List[Dict[str, Any]], 
                        sheet_name: str = "Data") -> bool:
        """
        Write data to Excel file.
        
        Args:
            filepath: Output file path
            data: List of dictionaries to write
            sheet_name: Name of the worksheet
            
        Returns:
            True if successful
            
        Raises:
            DataFileError: If file cannot be written
        """
        try:
            if not data:
                raise DataFileError("No data to write", filepath=filepath)
            
            # Ensure output directory exists
            ensure_directory(os.path.dirname(filepath))
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            
            # Get all fieldnames
            fieldnames = list(data[0].keys())
            
            # Write headers
            for col, fieldname in enumerate(fieldnames, start=1):
                worksheet.cell(row=1, column=col, value=fieldname)
            
            # Write data
            for row_num, row_data in enumerate(data, start=2):
                for col, fieldname in enumerate(fieldnames, start=1):
                    value = row_data.get(fieldname, '')
                    worksheet.cell(row=row_num, column=col, value=value)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(filepath)
            workbook.close()
            
            logger.info(f"Excel file written successfully: {filepath} ({len(data)} rows)")
            return True
        
        except Exception as e:
            raise DataFileError(f"Failed to write Excel file: {e}", filepath=filepath)
    
    def read_json_file(self, filepath: str) -> Dict[str, Any]:
        """
        Read JSON file.
        
        Args:
            filepath: Path to JSON file
            
        Returns:
            Parsed JSON data
            
        Raises:
            DataFileError: If file cannot be read
        """
        try:
            validated_path = validate_attachment_path(filepath)
            data = safe_json_load(validated_path)
            
            if data is None:
                raise DataFileError("Failed to parse JSON file", filepath=filepath)
            
            return data
        
        except Exception as e:
            raise DataFileError(f"Failed to read JSON file: {e}", filepath=filepath)
    
    def write_json_file(self, filepath: str, data: Dict[str, Any], indent: int = 2) -> bool:
        """
        Write data to JSON file.
        
        Args:
            filepath: Output file path
            data: Data to write
            indent: JSON indentation
            
        Returns:
            True if successful
            
        Raises:
            DataFileError: If file cannot be written
        """
        try:
            # Ensure output directory exists
            ensure_directory(os.path.dirname(filepath))
            
            success = safe_json_save(data, filepath, indent)
            if not success:
                raise DataFileError("Failed to save JSON file", filepath=filepath)
            
            logger.info(f"JSON file written successfully: {filepath}")
            return True
        
        except Exception as e:
            raise DataFileError(f"Failed to write JSON file: {e}", filepath=filepath)
    
    def import_leads_from_file(self, filepath: str, email_column: str = 'email') -> List[Dict[str, Any]]:
        """
        Import leads from CSV or Excel file.
        
        Args:
            filepath: Path to file
            email_column: Name of email column
            
        Returns:
            List of validated lead dictionaries
            
        Raises:
            DataFileError: If import fails
        """
        file_ext = Path(filepath).suffix.lower()
        
        try:
            # Read file based on extension
            if file_ext == '.csv':
                rows = list(self.read_csv_file(filepath))
            elif file_ext in ['.xlsx', '.xls']:
                rows = list(self.read_excel_file(filepath))
            else:
                raise DataFileError(f"Unsupported file format: {file_ext}", filepath=filepath)
            
            if not rows:
                raise DataFileError("File contains no data", filepath=filepath)
            
            # Check if email column exists
            first_row = rows[0]
            if email_column not in first_row:
                available_columns = list(first_row.keys())
                raise DataFileError(
                    f"Email column '{email_column}' not found. Available columns: {available_columns}",
                    filepath=filepath
                )
            
            # Convert rows to lead format
            leads = []
            errors = []
            
            for row in rows:
                try:
                    # Map common fields
                    lead_data = {
                        'email': row.get(email_column, '').strip(),
                        'first_name': row.get('first_name', row.get('First Name', '')).strip(),
                        'last_name': row.get('last_name', row.get('Last Name', '')).strip(),
                        'company': row.get('company', row.get('Company', '')).strip(),
                        'title': row.get('title', row.get('Title', '')).strip(),
                        'phone': row.get('phone', row.get('Phone', '')).strip(),
                    }
                    
                    # Add custom fields for any additional columns
                    custom_fields = {}
                    for key, value in row.items():
                        if key not in ['email', 'first_name', 'last_name', 'company', 'title', 'phone', '_row_number']:
                            clean_key = str(key).strip()
                            clean_value = str(value).strip() if value else ""
                            if clean_key and clean_value:
                                custom_fields[clean_key] = clean_value
                    
                    if custom_fields:
                        lead_data['custom_fields'] = custom_fields
                    
                    # Validate lead data
                    validated_lead = self.validator.validate_lead_data(lead_data)
                    leads.append(validated_lead)
                
                except ValidationError as e:
                    row_num = row.get('_row_number', 'unknown')
                    errors.append(f"Row {row_num}: {e.message}")
                except Exception as e:
                    row_num = row.get('_row_number', 'unknown')
                    errors.append(f"Row {row_num}: {str(e)}")
            
            # Log import results
            logger.info(f"Imported {len(leads)} leads from {filepath}")
            if errors:
                logger.warning(f"Import had {len(errors)} errors")
                for error in errors[:10]:  # Log first 10 errors
                    logger.warning(f"Import error: {error}")
            
            return leads
        
        except Exception as e:
            raise DataFileError(f"Failed to import leads: {e}", filepath=filepath)
    
    def export_leads_to_file(self, leads: List[Dict[str, Any]], filepath: str) -> bool:
        """
        Export leads to CSV or Excel file.
        
        Args:
            leads: List of lead dictionaries
            filepath: Output file path
            
        Returns:
            True if successful
            
        Raises:
            DataFileError: If export fails
        """
        if not leads:
            raise DataFileError("No leads to export", filepath=filepath)
        
        file_ext = Path(filepath).suffix.lower()
        
        try:
            # Flatten lead data for export
            export_data = []
            
            for lead in leads:
                row = {
                    'email': lead.get('email', ''),
                    'first_name': lead.get('first_name', ''),
                    'last_name': lead.get('last_name', ''),
                    'company': lead.get('company', ''),
                    'title': lead.get('title', ''),
                    'phone': lead.get('phone', ''),
                }
                
                # Add custom fields
                custom_fields = lead.get('custom_fields', {})
                for key, value in custom_fields.items():
                    row[f"custom_{key}"] = value
                
                # Add metadata if available
                if 'created_at' in lead:
                    row['created_at'] = lead['created_at']
                if 'tags' in lead and lead['tags']:
                    row['tags'] = ', '.join(lead['tags'])
                
                export_data.append(row)
            
            # Export based on file extension
            if file_ext == '.csv':
                return self.write_csv_file(filepath, export_data)
            elif file_ext in ['.xlsx']:
                return self.write_excel_file(filepath, export_data)
            else:
                raise DataFileError(f"Unsupported export format: {file_ext}", filepath=filepath)
        
        except Exception as e:
            raise DataFileError(f"Failed to export leads: {e}", filepath=filepath)
    
    def backup_data_file(self, filepath: str, backup_dir: Optional[str] = None) -> str:
        """
        Create a backup of a data file.
        
        Args:
            filepath: Path to file to backup
            backup_dir: Backup directory (auto-determined if None)
            
        Returns:
            Path to backup file
            
        Raises:
            DataFileError: If backup fails
        """
        try:
            if not os.path.exists(filepath):
                raise DataFileError(f"File to backup does not exist: {filepath}")
            
            if backup_dir is None:
                backup_dir = os.path.join(self.base_data_dir, 'backups')
            
            ensure_directory(backup_dir)
            
            # Generate backup filename with timestamp
            filename = os.path.basename(filepath)
            name, ext = os.path.splitext(filename)
            timestamp = time.strftime('%Y%m%d_%H%M%S')
            backup_filename = f"{name}_backup_{timestamp}{ext}"
            backup_path = os.path.join(backup_dir, backup_filename)
            
            # Copy file
            import shutil
            shutil.copy2(filepath, backup_path)
            
            logger.info(f"Created backup: {filepath} -> {backup_path}")
            return backup_path
        
        except Exception as e:
            raise DataFileError(f"Failed to create backup: {e}", filepath=filepath)
    
    def get_file_info(self, filepath: str) -> Dict[str, Any]:
        """
        Get information about a file.
        
        Args:
            filepath: Path to file
            
        Returns:
            File information dictionary
        """
        try:
            validated_path = validate_attachment_path(filepath)
            stat = os.stat(validated_path)
            
            return {
                'path': validated_path,
                'filename': os.path.basename(validated_path),
                'size_bytes': stat.st_size,
                'size_mb': stat.st_size / (1024 * 1024),
                'modified_time': stat.st_mtime,
                'extension': Path(validated_path).suffix.lower(),
                'exists': True
            }
        
        except Exception as e:
            return {
                'path': filepath,
                'exists': False,
                'error': str(e)
            }


# Convenience functions
def import_leads_from_csv(filepath: str, email_column: str = 'email') -> List[Dict[str, Any]]:
    """
    Convenience function to import leads from CSV file.
    
    Args:
        filepath: Path to CSV file
        email_column: Name of email column
        
    Returns:
        List of validated lead dictionaries
    """
    handler = FileHandler()
    return handler.import_leads_from_file(filepath, email_column)


def export_leads_to_csv(leads: List[Dict[str, Any]], filepath: str) -> bool:
    """
    Convenience function to export leads to CSV file.
    
    Args:
        leads: List of lead dictionaries
        filepath: Output file path
        
    Returns:
        True if successful
    """
    handler = FileHandler()
    return handler.export_leads_to_file(leads, filepath)